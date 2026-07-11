from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from vhlookup_core.excel_comments import make_comment
from vhlookup_core.inspection import InspectionResult


class InspectionReportWriter:
    def write_xlsx(self, result: InspectionResult, path: str | Path) -> Path:
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            source_sheet_names = self._write_source_review_sheets(writer, result)
            self._guide_frame(result).to_excel(writer, sheet_name="먼저확인", index=False)
            result.files.to_excel(writer, sheet_name="파일점검", index=False)
            result.sheets.to_excel(writer, sheet_name="시트헤더점검", index=False)
            if not result.column_comparison.empty:
                result.column_comparison.to_excel(writer, sheet_name="열비교", index=False)
            if not result.column_profiles.empty:
                result.column_profiles.to_excel(writer, sheet_name="컬럼요약", index=False)
            if not result.column_mappings.empty:
                result.column_mappings.to_excel(writer, sheet_name="컬럼매칭점검", index=False)
                self._mapping_evidence_frame(result).to_excel(writer, sheet_name="자동추천근거", index=False)
            if not result.privacy_records.empty:
                result.privacy_records.to_excel(writer, sheet_name="개인정보점검", index=False)
            self._issue_frame(result).to_excel(writer, sheet_name="확인필요", index=False)
            pd.DataFrame(
                [{"항목": key, "값": value} for key, value in result.summary.items()]
            ).to_excel(writer, sheet_name="점검요약", index=False)
            self._style_workbook(writer.book)
            self._mark_source_review_sheets(writer.book, result, source_sheet_names)
        return output_path

    def _write_source_review_sheets(self, writer: pd.ExcelWriter, result: InspectionResult) -> dict[str, str]:
        sheet_names: dict[str, str] = {}
        used = set()
        for file_name, frame in result.source_tables.items():
            sheet_name = self._unique_sheet_name(f"원본확인_{Path(file_name).stem}", used)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            sheet_names[file_name] = sheet_name
        return sheet_names

    def _guide_frame(self, result: InspectionResult) -> pd.DataFrame:
        rows = [
            {
                "먼저 볼 내용": "색칠된 원본 확인",
                "현재 결과": f"{len(result.source_tables)}개 원본확인 시트",
                "바로 할 일": "결과 파일 앞쪽의 원본확인 시트에서 색칠된 셀과 메모를 먼저 확인하세요.",
                "관련 시트": "원본확인_*",
            },
            {
                "먼저 볼 내용": "개인정보 의심",
                "현재 결과": self._privacy_summary(result.privacy_records),
                "바로 할 일": "연락처, 이름, 이메일, 주소, 계좌, 주민번호 유형이 있으면 공유 전 필요 여부를 확인하세요.",
                "관련 시트": "개인정보점검",
            },
            {
                "먼저 볼 내용": "확인 필요",
                "현재 결과": f"{len(result.issues)}건",
                "바로 할 일": "오류나 경고가 있으면 확인필요 시트를 먼저 처리하세요.",
                "관련 시트": "확인필요",
            },
            {
                "먼저 볼 내용": "파일 처리",
                "현재 결과": f"{result.summary.get('file_count', 0)}개 파일",
                "바로 할 일": "실패 파일이 있으면 파일점검 시트에서 파일명과 오류를 확인하세요.",
                "관련 시트": "파일점검",
            },
            {
                "먼저 볼 내용": "파일별 열 차이",
                "현재 결과": self._column_difference_summary(result),
                "바로 할 일": "같은 양식으로 합칠 파일인데 누락 열이 있으면 원본 양식을 먼저 맞추세요.",
                "관련 시트": "열비교",
            },
            {
                "먼저 볼 내용": "헤더 자동탐지",
                "현재 결과": f"{len(result.sheets)}개 시트 점검",
                "바로 할 일": "헤더 행이 틀리게 잡힌 파일이 있으면 시트헤더점검에서 확인하세요.",
                "관련 시트": "시트헤더점검",
            },
            {
                "먼저 볼 내용": "컬럼별 값 분포",
                "현재 결과": f"{len(result.column_profiles)}개 컬럼 요약",
                "바로 할 일": "빈값이 많거나 숫자/날짜로 보이지 않는 컬럼을 확인하세요.",
                "관련 시트": "컬럼요약",
            },
            {
                "먼저 볼 내용": "컬럼 자동매칭",
                "현재 결과": f"{len(result.column_mappings)}건",
                "바로 할 일": "표준 양식과 맞춰야 하는 업무라면 자동 매칭 결과가 맞는지 확인하세요.",
                "관련 시트": "컬럼매칭점검, 자동추천근거",
            },
        ]
        return pd.DataFrame(rows)

    def _privacy_summary(self, frame: pd.DataFrame) -> str:
        if frame.empty:
            return "의심 항목 없음"
        rows = []
        grouped = frame.groupby("점검 유형", dropna=False)["감지 건수"].sum().sort_values(ascending=False)
        for category, count in grouped.items():
            rows.append(f"{category} {int(count)}건")
        return ", ".join(rows)

    def _column_difference_summary(self, result: InspectionResult) -> str:
        if result.column_comparison.empty:
            return "비교할 열 정보 없음"
        missing = result.column_comparison["누락 파일"].map(lambda value: bool(str(value).strip()))
        missing_count = int(missing.sum())
        if missing_count:
            return f"누락 파일이 있는 열 {missing_count}개"
        return f"공통/개별 열 {len(result.column_comparison)}개, 누락 열 없음"

    def _mark_source_review_sheets(self, workbook, result: InspectionResult, source_sheet_names: dict[str, str]) -> None:
        privacy_fill = PatternFill("solid", fgColor="FDE68A")
        issue_fill = PatternFill("solid", fgColor="FCA5A5")

        if not result.privacy_records.empty:
            for _, record in result.privacy_records.iterrows():
                file_name = str(record.get("파일명", ""))
                sheet_name = source_sheet_names.get(file_name)
                if not sheet_name or sheet_name not in workbook.sheetnames:
                    continue
                sheet = workbook[sheet_name]
                column_name = str(record.get("컬럼명", ""))
                column_index = self._find_column_index(sheet, column_name)
                if column_index is None:
                    continue
                message = (
                    f"개인정보 의심: {record.get('점검 유형', '')}\n"
                    f"위험도: {record.get('위험도', '')}\n"
                    f"감지 건수: {record.get('감지 건수', '')}\n"
                    f"{record.get('조치 안내', '')}"
                )
                self._mark_column(sheet, column_index, privacy_fill, message, max_cell_comments=30)

        for issue in result.issues:
            if not issue.file_name:
                continue
            sheet_name = source_sheet_names.get(issue.file_name)
            if not sheet_name or sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            message = f"확인 필요: {issue.message}"
            if issue.column_name:
                column_index = self._find_column_index(sheet, issue.column_name)
                if column_index is not None:
                    self._mark_column(sheet, column_index, issue_fill, message, max_cell_comments=10)
                    continue
            if issue.row_number:
                row_index = max(int(issue.row_number), 2)
                for cell in sheet[row_index]:
                    cell.fill = issue_fill
                    if cell.comment is None:
                        cell.comment = make_comment(message)

    def _mark_column(self, sheet, column_index: int, fill: PatternFill, message: str, max_cell_comments: int) -> None:
        header = sheet.cell(row=1, column=column_index)
        header.fill = fill
        header.comment = make_comment(message)
        comments_left = max_cell_comments
        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            if cell.value in (None, ""):
                continue
            cell.fill = fill
            if comments_left > 0 and cell.comment is None:
                cell.comment = make_comment(message)
                comments_left -= 1

    def _find_column_index(self, sheet, column_name: str) -> int | None:
        for cell in sheet[1]:
            if str(cell.value) == column_name:
                return int(cell.column)
        return None

    def _unique_sheet_name(self, value: object, used_sheet_names: set[str]) -> str:
        raw = str(value).strip() or "원본확인"
        for char in "[]:*?/\\":
            raw = raw.replace(char, "_")
        base = raw[:31] or "원본확인"
        sheet_name = base
        suffix = 2
        while sheet_name in used_sheet_names:
            suffix_text = f"_{suffix}"
            sheet_name = f"{base[:31 - len(suffix_text)]}{suffix_text}"
            suffix += 1
        used_sheet_names.add(sheet_name)
        return sheet_name

    def _mapping_evidence_frame(self, result: InspectionResult) -> pd.DataFrame:
        frame = result.column_mappings.copy()
        if frame.empty:
            return frame
        frame["확인 안내"] = frame.apply(
            lambda row: (
                "자동 매칭됨"
                if str(row.get("원본 컬럼", "")).strip() and float(row.get("신뢰도", 0) or 0) > 0
                else "원본에서 대응 컬럼을 찾지 못했습니다"
            ),
            axis=1,
        )
        return frame

    def _issue_frame(self, result: InspectionResult) -> pd.DataFrame:
        rows = [
            {
                "심각도": issue.severity,
                "확인 유형": issue.issue_type,
                "안내 문구": issue.message,
                "파일명": issue.file_name,
                "상세": issue.details,
            }
            for issue in result.issues
        ]
        return pd.DataFrame(rows, columns=["심각도", "확인 유형", "안내 문구", "파일명", "상세"])

    def _style_workbook(self, workbook) -> None:
        header_fill = PatternFill("solid", fgColor="1F6F5F")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            if sheet.max_row > 1 and sheet.max_column > 0:
                sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, min(len(value), 48))
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                sheet.column_dimensions[column_letter].width = max(10, min(max_length + 2, 50))
