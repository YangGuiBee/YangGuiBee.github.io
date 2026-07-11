from __future__ import annotations

import re
from dataclasses import dataclass
from numbers import Real
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from vhlookup_core.excel_comments import make_comment
from vhlookup_core.header import HeaderDetector
from vhlookup_core.loader import ExcelLoader
from vhlookup_core.models import JobResult
from vhlookup_core.normalization import is_blank, normalize_header
from vhlookup_core.sheet import SheetDetector


SPLIT_COLUMN_HINTS = (
    "부서",
    "기관",
    "학교",
    "소속",
    "팀",
    "과",
    "department",
    "dept",
    "division",
    "office",
    "state",
    "status",
    "position",
    "manager",
)

PIVOT_ROW_HINTS = (
    "부서",
    "기관",
    "학교",
    "소속",
    "팀",
    "과",
    "사업",
    "항목",
    "구분",
    "상태",
    "직급",
    "담당",
    "department",
    "dept",
    "division",
    "office",
    "category",
    "status",
    "type",
)

PIVOT_COLUMN_HINTS = (
    "월",
    "분기",
    "연도",
    "년도",
    "상태",
    "구분",
    "month",
    "quarter",
    "year",
    "status",
    "type",
)

PIVOT_VALUE_HINTS = (
    "금액",
    "예산",
    "수당",
    "단가",
    "급여",
    "지급",
    "집행",
    "수량",
    "인원",
    "현원",
    "정원",
    "건수",
    "점수",
    "salary",
    "amount",
    "budget",
    "pay",
    "cost",
    "price",
    "fee",
    "qty",
    "quantity",
    "count",
    "total",
)

NON_VALUE_HINTS = (
    "사번",
    "직원번호",
    "기관코드",
    "학교코드",
    "코드",
    "id",
    "번호",
    "연락처",
    "전화",
    "우편",
    "phone",
    "tel",
    "code",
)

AGGREGATION_ALIASES = {
    "건수": "건수",
    "count": "건수",
    "행개수": "건수",
    "개수": "건수",
    "합계": "합계",
    "sum": "합계",
    "total": "합계",
    "평균": "평균",
    "average": "평균",
    "mean": "평균",
    "최대": "최대",
    "max": "최대",
    "최소": "최소",
    "min": "최소",
}

AGGREGATION_FUNCTIONS = {
    "건수": "sum",
    "합계": "sum",
    "평균": "mean",
    "최대": "max",
    "최소": "min",
}

COUNT_HELPER_COLUMN = "__vhlookup_row_count__"
COUNT_DISPLAY_COLUMN = "건수"
PIVOT_DECIMAL_PLACES = 2
PIVOT_INTEGER_FORMAT = "#,##0"
PIVOT_DECIMAL_FORMAT = "#,##0.##"


@dataclass(frozen=True)
class SplitWorkbookResult:
    output_path: Path
    split_column: str
    sheet_count: int
    row_count: int


@dataclass(frozen=True)
class PivotSummaryResult:
    source_frame: pd.DataFrame
    pivot_frame: pd.DataFrame
    top_frame: pd.DataFrame
    guide_frame: pd.DataFrame
    invalid_rows: pd.DataFrame
    summary: dict[str, Any]


@dataclass(frozen=True)
class PivotWorkbookResult:
    output_path: Path
    row_column: str
    column_column: str | None
    value_column: str | None
    aggregation: str
    row_count: int
    invalid_value_count: int


class AdminWorkbookTools:
    def __init__(
        self,
        loader: ExcelLoader | None = None,
        header_detector: HeaderDetector | None = None,
        sheet_detector: SheetDetector | None = None,
    ) -> None:
        self.loader = loader or ExcelLoader()
        self.header_detector = header_detector or HeaderDetector()
        self.sheet_detector = sheet_detector or SheetDetector(self.header_detector)

    def load_table(self, path: str | Path) -> tuple[pd.DataFrame, dict[str, Any]]:
        source = self.loader.load(path)
        sheet = self.sheet_detector.select(source)
        detection = self.header_detector.detect(sheet)
        table = self.header_detector.apply(sheet, detection)
        metadata = {
            "file_name": Path(path).name,
            "sheet_name": sheet.name,
            "header_row_number": detection.header_row_number,
            "data_start_row_number": detection.data_start_row_number,
            "header_confidence": detection.confidence,
            "warnings": detection.warnings,
        }
        return table, metadata

    def clean_file(self, path: str | Path) -> JobResult:
        table, metadata = self.load_table(path)
        original_rows = len(table)
        original_columns = len(table.columns)
        cleaned = self._clean_frame(table)
        removed_rows = original_rows - len(cleaned)
        removed_columns = original_columns - len(cleaned.columns)
        duplicate_rows = int(cleaned.duplicated().sum()) if not cleaned.empty else 0
        blank_cells = int(cleaned.map(is_blank).sum().sum()) if not cleaned.empty else 0

        original_row_numbers = [metadata["data_start_row_number"] + int(index) for index in cleaned.index]
        cleaned = cleaned.reset_index(drop=True)
        self._insert_tracking_columns(cleaned, metadata, original_row_numbers)

        summary = {
            "workflow": "엑셀 파일 실무 정리",
            "row_count": len(cleaned),
            "original_row_count": original_rows,
            "original_column_count": original_columns,
            "removed_blank_rows": removed_rows,
            "removed_blank_columns": removed_columns,
            "duplicate_row_count": duplicate_rows,
            "blank_cell_count": blank_cells,
        }
        mapping_records = [
            {
                "역할": "파일 정리",
                "파일명": metadata["file_name"],
                "시트명": metadata["sheet_name"],
                "원본 컬럼": "",
                "표준 컬럼": "",
                "신뢰도": metadata["header_confidence"],
                "추천 방식": f"{metadata['header_row_number']}행을 헤더로 보고 정리본 생성",
                "검토 메모": "빈 행/빈 열 제거, 문자 앞뒤 공백 제거, 필터/틀고정 적용",
            },
            {
                "역할": "정리 요약",
                "파일명": metadata["file_name"],
                "시트명": metadata["sheet_name"],
                "원본 컬럼": "",
                "표준 컬럼": "",
                "신뢰도": 1.0,
                "추천 방식": "원본 파일은 수정하지 않고 새 결과 파일 생성",
                "검토 메모": f"빈 행 {removed_rows}개, 빈 열 {removed_columns}개, 중복 행 {duplicate_rows}개",
            },
        ]
        return JobResult(result_frame=cleaned, summary=summary, mapping_records=mapping_records)

    def infer_split_column(self, frame: pd.DataFrame) -> str:
        candidates: list[tuple[int, float, str]] = []
        row_count = max(len(frame), 1)
        for column in frame.columns:
            series = frame[column].dropna().map(lambda value: str(value).strip())
            series = series[series != ""]
            if series.empty:
                continue
            unique_count = int(series.nunique())
            if unique_count < 2 or unique_count > min(60, max(8, row_count // 2)):
                continue
            normalized = normalize_header(column)
            hint_score = 1 if any(hint in normalized for hint in SPLIT_COLUMN_HINTS) else 0
            ratio = unique_count / row_count
            candidates.append((hint_score, -ratio, str(column)))
        if candidates:
            candidates.sort(reverse=True)
            return candidates[0][2]
        for column in frame.columns:
            if frame[column].dropna().nunique() > 1:
                return str(column)
        raise ValueError("분류 기준으로 사용할 만한 열을 찾지 못했습니다.")

    def write_split_workbook(
        self,
        path: str | Path,
        output_path: str | Path,
        split_column: str | None = None,
    ) -> SplitWorkbookResult:
        table, metadata = self.load_table(path)
        cleaned = self._clean_frame(table).reset_index(drop=True)
        selected_column = split_column or self.infer_split_column(cleaned)
        if selected_column not in cleaned.columns:
            raise KeyError(f"분류 기준 열을 찾지 못했습니다: {selected_column}")

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        groups = list(cleaned.groupby(cleaned[selected_column].map(lambda value: "(빈값)" if is_blank(value) else str(value)), dropna=False))
        used_sheet_names = {"전체", "확인사항"}
        split_sheet_names: list[str] = []
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for value, group in groups:
                sheet_name = self._unique_sheet_name(value, used_sheet_names)
                group.to_excel(writer, sheet_name=sheet_name, index=False)
                split_sheet_names.append(sheet_name)
            cleaned.to_excel(writer, sheet_name="전체", index=False)
            guide = pd.DataFrame(
                [
                    {"항목": "먼저 볼 내용", "내용": "앞쪽의 분류별 시트에서 나뉜 결과를 바로 확인하세요."},
                    {"항목": "원본 파일", "내용": metadata["file_name"]},
                    {"항목": "원본 시트", "내용": metadata["sheet_name"]},
                    {"항목": "분류 기준 열", "내용": selected_column},
                    {"항목": "전체 행 수", "내용": len(cleaned)},
                    {"항목": "생성 시트 수", "내용": len(groups)},
                    {"항목": "주의", "내용": "원본 파일은 수정하지 않았습니다."},
                ]
            )
            guide.to_excel(writer, sheet_name="확인사항", index=False)
            self._style_workbook(writer.book)
        return SplitWorkbookResult(output, selected_column, len(groups), len(cleaned))

    def prepare_pivot_source(self, path: str | Path) -> tuple[pd.DataFrame, dict[str, Any]]:
        table, metadata = self.load_table(path)
        cleaned = self._clean_frame(table).reset_index(drop=True)
        cleaned.columns = self._unique_column_names(cleaned.columns)
        return cleaned, metadata

    def infer_pivot_defaults(self, frame: pd.DataFrame) -> dict[str, str]:
        cleaned = frame.copy()
        cleaned.columns = self._unique_column_names(cleaned.columns)
        row_column = self._infer_pivot_row_column(cleaned)
        value_column = self._infer_pivot_value_column(cleaned)
        column_column = self._infer_pivot_column_column(cleaned, row_column, value_column)
        return {
            "row_column": row_column,
            "column_column": column_column,
            "value_column": value_column,
            "aggregation": "합계" if value_column else "건수",
        }

    def build_pivot_summary(
        self,
        frame: pd.DataFrame,
        row_column: str,
        column_column: str | None = None,
        value_column: str | None = None,
        aggregation: str = "건수",
        metadata: dict[str, Any] | None = None,
    ) -> PivotSummaryResult:
        source = self._clean_frame(frame).reset_index(drop=True)
        source.columns = self._unique_column_names(source.columns)
        if not row_column:
            raise ValueError("행 기준 열을 선택하세요.")
        if row_column not in source.columns:
            raise KeyError(f"행 기준 열을 찾지 못했습니다: {row_column}")

        selected_column = self._none_if_empty(column_column)
        selected_value = self._none_if_empty(value_column)
        aggregation_label = self._normalize_aggregation(aggregation)
        if selected_column == row_column:
            raise ValueError("열 기준은 행 기준과 다른 열을 선택하세요.")
        if selected_column and selected_column not in source.columns:
            raise KeyError(f"열 기준 열을 찾지 못했습니다: {selected_column}")

        work = source.copy()
        invalid_rows = pd.DataFrame()
        if aggregation_label == "건수":
            work[COUNT_HELPER_COLUMN] = 1
            values_column = COUNT_HELPER_COLUMN
            output_value_column = COUNT_DISPLAY_COLUMN
        else:
            if not selected_value:
                raise ValueError("합계/평균/최대/최소를 만들려면 값 열을 선택하세요.")
            if selected_value not in source.columns:
                raise KeyError(f"값 열을 찾지 못했습니다: {selected_value}")
            numeric_values = self._coerce_numeric(work[selected_value])
            blank_mask = work[selected_value].map(is_blank)
            invalid_mask = ~blank_mask & numeric_values.isna()
            if invalid_mask.any():
                invalid_rows = source.loc[invalid_mask].copy()
                invalid_rows.insert(0, "오류내용", f"{selected_value} 열을 숫자로 읽을 수 없습니다.")
            work = work.loc[~invalid_mask].copy()
            work[selected_value] = numeric_values.loc[~invalid_mask]
            values_column = selected_value
            output_value_column = f"{selected_value}_{aggregation_label}"

        work[row_column] = work[row_column].map(self._pivot_label)
        if selected_column:
            work[selected_column] = work[selected_column].map(self._pivot_label)

        pivot_frame = self._make_pivot_frame(
            work,
            row_column=row_column,
            column_column=selected_column,
            values_column=values_column,
            output_value_column=output_value_column,
            aggregation=aggregation_label,
        )
        pivot_frame = self._clean_pivot_display_numbers(pivot_frame, label_columns={row_column})
        top_frame = self._make_top_frame(pivot_frame, row_column, selected_column, output_value_column)
        top_frame = self._clean_pivot_display_numbers(top_frame, label_columns={"순위", row_column})
        guide_frame = self._pivot_guide_frame(
            metadata=metadata,
            row_column=row_column,
            column_column=selected_column,
            value_column=None if aggregation_label == "건수" else selected_value,
            aggregation=aggregation_label,
            source_rows=len(source),
            invalid_rows=len(invalid_rows),
        )
        summary = {
            "workflow": "피벗 요약표 만들기",
            "row_column": row_column,
            "column_column": selected_column or "",
            "value_column": selected_value or COUNT_DISPLAY_COLUMN,
            "aggregation": aggregation_label,
            "row_count": len(source),
            "summary_row_count": len(pivot_frame),
            "invalid_value_count": len(invalid_rows),
        }
        return PivotSummaryResult(source, pivot_frame, top_frame, guide_frame, invalid_rows, summary)

    def write_pivot_workbook(
        self,
        path: str | Path,
        output_path: str | Path,
        row_column: str,
        column_column: str | None = None,
        value_column: str | None = None,
        aggregation: str = "건수",
    ) -> PivotWorkbookResult:
        source, metadata = self.prepare_pivot_source(path)
        result = self.build_pivot_summary(
            source,
            row_column=row_column,
            column_column=column_column,
            value_column=value_column,
            aggregation=aggregation,
            metadata=metadata,
        )
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            result.pivot_frame.to_excel(writer, sheet_name="피벗요약", index=False)
            self._pivot_review_frame(result).to_excel(writer, sheet_name="확인사항", index=False)
            self._style_workbook(writer.book)
            self._format_pivot_number_cells(writer.book, result)
            self._mark_pivot_summary(writer.book, result)
        return PivotWorkbookResult(
            output_path=output,
            row_column=str(result.summary["row_column"]),
            column_column=str(result.summary["column_column"]) or None,
            value_column=str(result.summary["value_column"]) or None,
            aggregation=str(result.summary["aggregation"]),
            row_count=int(result.summary["row_count"]),
            invalid_value_count=int(result.summary["invalid_value_count"]),
        )

    def _pivot_review_frame(self, result: PivotSummaryResult) -> pd.DataFrame:
        rows: list[dict[str, object]] = []
        columns = ["구분", "기준/컬럼", "내용", "확인할 점"]

        for _, guide_row in result.guide_frame.iterrows():
            rows.append(
                {
                    "구분": "요약 기준",
                    "기준/컬럼": guide_row.get("항목", ""),
                    "내용": guide_row.get("내용", ""),
                    "확인할 점": "",
                }
            )

        for _, top_row in result.top_frame.head(20).iterrows():
            label = top_row.get(result.summary.get("row_column", ""), "")
            rows.append(
                {
                    "구분": "상위 목록",
                    "기준/컬럼": label,
                    "내용": "; ".join(f"{column}={value}" for column, value in top_row.items()),
                    "확인할 점": "상위 항목이 예상과 맞는지 확인하세요.",
                }
            )

        for _, invalid_row in result.invalid_rows.head(50).iterrows():
            rows.append(
                {
                    "구분": "숫자 오류",
                    "기준/컬럼": result.summary.get("value_column", ""),
                    "내용": "; ".join(f"{column}={value}" for column, value in invalid_row.items()),
                    "확인할 점": "숫자로 집계해야 하는 값인지 원본을 확인하세요.",
                }
            )

        for key, value in result.summary.items():
            rows.append(
                {
                    "구분": "처리 요약",
                    "기준/컬럼": key,
                    "내용": value,
                    "확인할 점": "",
                }
            )

        if not rows:
            rows.append({"구분": "확인", "기준/컬럼": "", "내용": "추가 확인 항목이 없습니다.", "확인할 점": "첫 시트를 확인하세요."})
        return pd.DataFrame(rows, columns=columns)

    def _clean_frame(self, frame: pd.DataFrame) -> pd.DataFrame:
        cleaned = frame.copy()
        cleaned = cleaned.dropna(axis=0, how="all").dropna(axis=1, how="all")
        for column in cleaned.columns:
            if pd.api.types.is_object_dtype(cleaned[column]) or pd.api.types.is_string_dtype(cleaned[column]):
                cleaned[column] = cleaned[column].map(lambda value: value.strip() if isinstance(value, str) else value)
        return cleaned

    def _unique_column_names(self, columns) -> list[str]:
        used: dict[str, int] = {}
        names: list[str] = []
        for column in columns:
            base = str(column).strip() or "빈열"
            count = used.get(base, 0) + 1
            used[base] = count
            names.append(base if count == 1 else f"{base}_{count}")
        return names

    def _infer_pivot_row_column(self, frame: pd.DataFrame) -> str:
        candidates: list[tuple[int, float, str]] = []
        row_count = max(len(frame), 1)
        for column in frame.columns:
            series = frame[column].dropna().map(lambda value: str(value).strip())
            series = series[series != ""]
            if series.empty:
                continue
            unique_count = int(series.nunique())
            if unique_count < 1 or unique_count > min(80, max(8, int(row_count * 0.8))):
                continue
            normalized = normalize_header(column)
            hint_score = self._hint_score(normalized, PIVOT_ROW_HINTS)
            id_penalty = self._hint_score(normalized, NON_VALUE_HINTS)
            ratio = unique_count / row_count
            candidates.append((hint_score - id_penalty, -ratio, str(column)))
        if candidates:
            candidates.sort(reverse=True)
            return candidates[0][2]
        if len(frame.columns) > 0:
            return str(frame.columns[0])
        raise ValueError("피벗 기준으로 사용할 열을 찾지 못했습니다.")

    def _infer_pivot_column_column(
        self,
        frame: pd.DataFrame,
        row_column: str,
        value_column: str,
    ) -> str:
        candidates: list[tuple[int, float, str]] = []
        row_count = max(len(frame), 1)
        for column in frame.columns:
            if str(column) in {row_column, value_column}:
                continue
            series = frame[column].dropna().map(lambda value: str(value).strip())
            series = series[series != ""]
            if series.empty:
                continue
            unique_count = int(series.nunique())
            if unique_count < 2 or unique_count > min(24, max(6, row_count // 2)):
                continue
            normalized = normalize_header(column)
            hint_score = self._hint_score(normalized, PIVOT_COLUMN_HINTS)
            if hint_score <= 0:
                continue
            candidates.append((hint_score, -unique_count / row_count, str(column)))
        if candidates:
            candidates.sort(reverse=True)
            return candidates[0][2]
        return ""

    def _infer_pivot_value_column(self, frame: pd.DataFrame) -> str:
        candidates: list[tuple[int, float, str]] = []
        for column in frame.columns:
            normalized = normalize_header(column)
            non_blank = frame[column][~frame[column].map(is_blank)]
            if non_blank.empty:
                continue
            numeric = self._coerce_numeric(non_blank)
            numeric_ratio = float(numeric.notna().sum()) / max(len(non_blank), 1)
            if numeric_ratio < 0.75:
                continue
            hint_score = self._hint_score(normalized, PIVOT_VALUE_HINTS)
            id_penalty = self._hint_score(normalized, NON_VALUE_HINTS)
            candidates.append((hint_score - id_penalty, numeric_ratio, str(column)))
        if candidates:
            candidates.sort(reverse=True)
            if candidates[0][0] >= 0:
                return candidates[0][2]
        return ""

    def _hint_score(self, normalized_column: str, hints: tuple[str, ...]) -> int:
        for index, hint in enumerate(hints):
            if hint in normalized_column:
                return len(hints) - index
        return 0

    def _coerce_numeric(self, series: pd.Series) -> pd.Series:
        def clean(value: object) -> object:
            if is_blank(value):
                return pd.NA
            if isinstance(value, (int, float)):
                return value
            text = str(value).strip()
            text = re.sub(r"[\s,원₩$]", "", text)
            return text

        return pd.to_numeric(series.map(clean), errors="coerce")

    def _normalize_aggregation(self, aggregation: str) -> str:
        key = normalize_header(aggregation)
        if key in AGGREGATION_ALIASES:
            return AGGREGATION_ALIASES[key]
        raise ValueError(f"지원하지 않는 집계 방식입니다: {aggregation}")

    def _none_if_empty(self, value: str | None) -> str | None:
        if value is None:
            return None
        stripped = str(value).strip()
        if not stripped or stripped in {"(선택 안 함)", "(행 개수)"}:
            return None
        return stripped

    def _pivot_label(self, value: object) -> str:
        if is_blank(value):
            return "(빈값)"
        return str(value).strip()

    def _make_pivot_frame(
        self,
        work: pd.DataFrame,
        row_column: str,
        column_column: str | None,
        values_column: str,
        output_value_column: str,
        aggregation: str,
    ) -> pd.DataFrame:
        if work.empty:
            return pd.DataFrame([{"안내": "요약할 데이터가 없습니다."}])
        aggfunc = AGGREGATION_FUNCTIONS[aggregation]
        if column_column:
            pivot = pd.pivot_table(
                work,
                index=row_column,
                columns=column_column,
                values=values_column,
                aggfunc=aggfunc,
                fill_value=0,
                margins=True,
                margins_name="합계",
                dropna=False,
            ).reset_index()
            pivot.columns = [str(column) for column in pivot.columns]
            if "합계" in pivot.columns:
                total_rows = pivot[pivot[row_column] == "합계"]
                body = pivot[pivot[row_column] != "합계"].copy()
                body = body.sort_values("합계", ascending=False, kind="stable")
                pivot = pd.concat([body, total_rows], ignore_index=True)
            return pivot

        grouped = work.groupby(row_column, dropna=False)[values_column].agg(aggfunc).reset_index(name=output_value_column)
        grouped = grouped.sort_values(output_value_column, ascending=False, kind="stable")
        total_value = work[values_column].agg(aggfunc)
        total = pd.DataFrame([{row_column: "합계", output_value_column: total_value}])
        return pd.concat([grouped, total], ignore_index=True)

    def _clean_pivot_display_numbers(self, frame: pd.DataFrame, label_columns: set[str]) -> pd.DataFrame:
        if frame.empty:
            return frame
        output = frame.copy()
        for column in output.columns:
            if str(column) in label_columns:
                continue
            output[column] = output[column].map(self._clean_pivot_number)
        return output

    def _clean_pivot_number(self, value: object) -> object:
        if is_blank(value):
            return value
        if isinstance(value, Real) and not isinstance(value, bool):
            rounded = round(float(value), PIVOT_DECIMAL_PLACES)
            if rounded.is_integer():
                return int(rounded)
            return rounded
        return value

    def _make_top_frame(
        self,
        pivot_frame: pd.DataFrame,
        row_column: str,
        column_column: str | None,
        output_value_column: str,
    ) -> pd.DataFrame:
        if pivot_frame.empty or row_column not in pivot_frame.columns:
            return pd.DataFrame([{"안내": "상위목록을 만들 데이터가 없습니다."}])
        body = pivot_frame[pivot_frame[row_column] != "합계"].copy()
        if body.empty:
            return pd.DataFrame([{"안내": "상위목록을 만들 데이터가 없습니다."}])
        measure_column = "합계" if column_column and "합계" in body.columns else output_value_column
        if measure_column not in body.columns:
            measure_column = next((column for column in body.columns if column != row_column), row_column)
        body = body.sort_values(measure_column, ascending=False, kind="stable").head(30).reset_index(drop=True)
        body.insert(0, "순위", range(1, len(body) + 1))
        return body

    def _pivot_guide_frame(
        self,
        metadata: dict[str, Any] | None,
        row_column: str,
        column_column: str | None,
        value_column: str | None,
        aggregation: str,
        source_rows: int,
        invalid_rows: int,
    ) -> pd.DataFrame:
        metadata = metadata or {}
        rows = [
            {"항목": "원본 파일", "내용": metadata.get("file_name", "")},
            {"항목": "원본 시트", "내용": metadata.get("sheet_name", "")},
            {"항목": "행 기준", "내용": row_column},
            {"항목": "열 기준", "내용": column_column or "사용 안 함"},
            {"항목": "값 열", "내용": value_column or COUNT_DISPLAY_COLUMN},
            {"항목": "집계 방식", "내용": aggregation},
            {"항목": "원본 행 수", "내용": source_rows},
            {"항목": "숫자 오류 행", "내용": invalid_rows},
            {"항목": "주의", "내용": "원본 파일은 수정하지 않았습니다. 이 파일은 실행 시점의 요약표입니다."},
        ]
        return pd.DataFrame(rows)

    def _pivot_check_frame(self, result: PivotSummaryResult) -> pd.DataFrame:
        invalid_count = int(result.summary.get("invalid_value_count", 0))
        base_rows = [
            {
                "먼저 볼 내용": "피벗 요약표",
                "현재 결과": f"{result.summary.get('summary_row_count', 0)}행",
                "바로 할 일": "첫 번째 `피벗요약` 시트에서 요약 결과를 확인하세요.",
                "관련 시트": "피벗요약",
            },
            {
                "먼저 볼 내용": "요약 기준",
                "현재 결과": (
                    f"행 기준: {result.summary.get('row_column', '')}, "
                    f"열 기준: {result.summary.get('column_column', '') or '사용 안 함'}, "
                    f"값: {result.summary.get('value_column', '')}, "
                    f"집계: {result.summary.get('aggregation', '')}"
                ),
                "바로 할 일": "기준이 의도와 다르면 프로그램에서 드롭박스를 다시 선택해 실행하세요.",
                "관련 시트": "기준설명",
            },
        ]
        if invalid_count:
            base_rows.append(
                {
                    "먼저 볼 내용": "숫자 오류 행",
                    "현재 결과": f"{invalid_count}건",
                    "바로 할 일": "`오류행만` 시트에서 숫자로 읽히지 않은 값을 확인하세요.",
                    "관련 시트": "오류행만",
                }
            )
        return pd.DataFrame(base_rows)

    def _mark_pivot_summary(self, workbook, result: PivotSummaryResult) -> None:
        if "피벗요약" not in workbook.sheetnames:
            return
        sheet = workbook["피벗요약"]
        fill = PatternFill("solid", fgColor="DBEAFE")
        message = (
            f"피벗 요약 기준\n"
            f"행 기준: {result.summary.get('row_column', '')}\n"
            f"열 기준: {result.summary.get('column_column', '') or '사용 안 함'}\n"
            f"값: {result.summary.get('value_column', '')}\n"
            f"집계: {result.summary.get('aggregation', '')}"
        )
        for cell in sheet[1]:
            cell.fill = fill
            if cell.comment is None:
                cell.comment = make_comment(message)

    def _format_pivot_number_cells(self, workbook, result: PivotSummaryResult) -> None:
        for sheet_name in ("피벗요약", "상위목록"):
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, int):
                        cell.number_format = PIVOT_INTEGER_FORMAT
                    elif isinstance(cell.value, float):
                        cell.number_format = PIVOT_DECIMAL_FORMAT

    def _unique_sheet_name(self, value: object, used_sheet_names: set[str]) -> str:
        raw = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(value)).strip() or "빈값"
        base = raw[:31] or "분류"
        sheet_name = base
        suffix = 2
        while sheet_name in used_sheet_names:
            suffix_text = f"_{suffix}"
            sheet_name = f"{base[:31 - len(suffix_text)]}{suffix_text}"
            suffix += 1
        used_sheet_names.add(sheet_name)
        return sheet_name

    def _insert_tracking_columns(
        self,
        frame: pd.DataFrame,
        metadata: dict[str, Any],
        original_row_numbers: list[int],
    ) -> None:
        columns = set(map(str, frame.columns))
        if {"원본 파일명", "원본 시트명", "원본 행 번호"} & columns:
            names = ("처리 원본 파일명", "처리 원본 시트명", "처리 원본 행 번호")
        else:
            names = ("원본 파일명", "원본 시트명", "원본 행 번호")
        frame.insert(0, names[2], original_row_numbers)
        frame.insert(0, names[1], metadata["sheet_name"])
        frame.insert(0, names[0], metadata["file_name"])

    def _style_workbook(self, workbook) -> None:
        header_fill = PatternFill("solid", fgColor="1F6F5F")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            if sheet.max_row > 1 and sheet.max_column > 0:
                sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, min(len(value), 42))
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                sheet.column_dimensions[column_letter].width = max(10, min(max_length + 2, 44))
