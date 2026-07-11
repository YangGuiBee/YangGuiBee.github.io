from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from vhlookup_core.excel_comments import make_comment
from vhlookup_core.models import JobResult
from vhlookup_core.privacy import PrivacyScanner


SHEET_GUIDE = "먼저확인"
SHEET_RESULT = "결과"
SHEET_REVIEW = "확인사항"
SHEET_ISSUES = "확인필요"
SHEET_ERROR_ROWS = "오류행만"
SHEET_SUMMARY = "처리요약"
SHEET_MAPPING = "매핑표"
SHEET_AUTO_EVIDENCE = "자동추천근거"
SHEET_PRIVACY = "개인정보점검"

SEVERITY_LABELS = {
    "info": "안내",
    "warning": "확인",
    "error": "오류",
}

ISSUE_LABELS = {
    "auto_detection_warning": "자동탐지 확인",
    "auto_inference_warning": "자동추천 확인",
    "file_processing_failed": "파일 처리 실패",
    "format_mismatch": "형식 불일치",
    "match_failed": "기준표에 없음",
    "missing_required_key": "필수 키 누락",
    "reference_duplicate_key": "기준표 중복",
    "reference_duplicate_key_blocked": "중복 키 병합 중단",
    "target_duplicate_key": "대상표 중복",
    "reference_missing_required_key": "기준표 키 누락",
    "target_missing_required_key": "대상표 키 누락",
    "missing_in_target": "대상표에 없음",
    "missing_in_reference": "기준표에 없음",
    "column_present_in_other_file": "다른 파일에만 있는 열",
    "reference_only_unmatched": "한쪽 파일에만 있는 행",
    "required_column_missing": "필수 컬럼 누락",
    "required_value_missing": "필수값 누락",
    "numeric_value_invalid": "숫자 오류",
    "date_value_invalid": "날짜 오류",
    "duplicate_business_key": "중복 제출 의심",
}

ISSUE_ACTIONS = {
    "auto_detection_warning": "헤더 행과 컬럼 매칭이 맞는지 미리보기에서 확인하세요.",
    "auto_inference_warning": "자동 추천된 키 컬럼과 가져올 컬럼이 맞는지 미리보기에서 확인하세요.",
    "file_processing_failed": "암호, 손상, 지원하지 않는 파일 형식인지 확인한 뒤 다시 처리하세요.",
    "format_mismatch": "사번, 날짜, 금액처럼 앞자리 0이나 표시 형식이 다른 컬럼을 같은 형식으로 맞추세요.",
    "match_failed": "붙일 파일에 해당 행이 없으면 빈 셀로 남는 것이 정상입니다. 기준값이 맞는지 확인하세요.",
    "missing_required_key": "키 컬럼의 빈 값을 채운 뒤 다시 실행하세요.",
    "reference_duplicate_key": "기준표에서 같은 키가 여러 건인 행을 먼저 정리하세요.",
    "reference_duplicate_key_blocked": "중복 키는 잘못 붙을 수 있어 자동 병합하지 않았습니다.",
    "target_duplicate_key": "대상표의 중복 행이 정상인지 확인하세요.",
    "reference_missing_required_key": "기준표 키 컬럼의 빈 값을 채우세요.",
    "target_missing_required_key": "대상표 키 컬럼의 빈 값을 채우세요.",
    "missing_in_target": "제출 누락, 교육 미이수, 지급대상 제외 여부를 확인하세요.",
    "missing_in_reference": "기준명단 누락 또는 잘못 제출된 대상인지 확인하세요.",
    "column_present_in_other_file": "이 열이 없는 파일의 행은 빈칸이 정상입니다. 필요한 값인지 확인하세요.",
    "reference_only_unmatched": "한쪽 파일에만 있는 행입니다. 누락인지, 추가 대상인지 확인하세요.",
    "required_column_missing": "제출 양식에 필수 컬럼이 있는지 확인하세요.",
    "required_value_missing": "해당 행의 빈 필수값을 채운 뒤 다시 수합하세요.",
    "numeric_value_invalid": "금액/수량에는 숫자만 남기고 단위나 설명은 비고로 옮기세요.",
    "date_value_invalid": "날짜를 YYYY-MM-DD 또는 기관에서 정한 날짜 형식으로 맞추세요.",
    "duplicate_business_key": "같은 기관/사업/항목이 중복 제출된 것인지 확인하세요.",
}

SUMMARY_LABELS = {
    "workflow": "업무 템플릿",
    "file_count": "처리 파일 수",
    "successful_file_count": "성공 파일 수",
    "row_count": "결과 행 수",
    "issue_count": "확인 필요 건수",
    "validation_issue_count": "값 검증 오류 건수",
    "mapped_column_count": "자동 매칭 컬럼 수",
    "target_rows": "대상표 행 수",
    "reference_rows": "기준표 행 수",
    "matched_rows": "매칭 성공 행 수",
    "auto_key_columns": "자동 선택 기준표 키",
    "auto_target_key_columns": "자동 선택 대상표 키",
    "auto_value_columns": "자동 선택 가져올 컬럼",
    "missing_in_target_rows": "대상표에 없는 건수",
    "missing_in_reference_rows": "기준표에 없는 건수",
    "reference_only_rows": "한쪽 파일에만 있는 행 수",
}


class ReportWriter:
    def write_xlsx(
        self,
        result: JobResult,
        path: str | Path,
        include_sensitive_details: bool = False,
        mark_result_cells: bool = True,
        include_privacy_scan: bool = True,
    ) -> Path:
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            result.result_frame.to_excel(writer, sheet_name=SHEET_RESULT, index=False)
            privacy_records = result.privacy_records or (
                PrivacyScanner().scan_frame(result.result_frame) if include_privacy_scan else []
            )
            self._review_frame(result, privacy_records, include_sensitive_details).to_excel(
                writer, sheet_name=SHEET_REVIEW, index=False
            )
            self._style_workbook(writer.book)
            if mark_result_cells:
                self._mark_result_sheet(writer.book, result, privacy_records)
        return output_path

    def _review_frame(
        self,
        result: JobResult,
        privacy_records: list[dict[str, object]],
        include_sensitive_details: bool,
    ) -> pd.DataFrame:
        rows: list[dict[str, object]] = []
        columns = ["구분", "파일명", "시트명", "기준/컬럼", "판단 내용", "확인할 점"]

        for record in result.mapping_records:
            basis_parts = [
                str(record.get(key, "")).strip()
                for key in ("기준표 컬럼", "대상표 컬럼", "원본 컬럼", "표준 컬럼")
                if str(record.get(key, "")).strip()
            ]
            rows.append(
                {
                    "구분": record.get("역할", "자동 매칭"),
                    "파일명": record.get("파일명", ""),
                    "시트명": record.get("시트명", ""),
                    "기준/컬럼": " -> ".join(basis_parts),
                    "판단 내용": record.get("추천 방식", ""),
                    "확인할 점": record.get("검토 메모", ""),
                }
            )

        if result.mapping is not None:
            for source, target in result.mapping.source_to_target.items():
                rows.append(
                    {
                        "구분": "컬럼 매칭",
                        "파일명": "",
                        "시트명": "",
                        "기준/컬럼": f"{source} -> {target}",
                        "판단 내용": "컬럼명 동의어/유사도 기준으로 맞췄습니다.",
                        "확인할 점": f"신뢰도 {result.mapping.confidence_by_target.get(target, '')}",
                    }
                )

        for issue in result.issues:
            row = {
                "구분": ISSUE_LABELS.get(issue.issue_type, issue.issue_type),
                "파일명": issue.file_name or "",
                "시트명": issue.sheet_name or "",
                "기준/컬럼": self._issue_column_text(issue) or "",
                "판단 내용": issue.message,
                "확인할 점": ISSUE_ACTIONS.get(issue.issue_type, "원본 자료와 매핑 설정을 확인하세요."),
            }
            if include_sensitive_details:
                row["상세"] = json.dumps(issue.details, ensure_ascii=False, sort_keys=True)
            rows.append(row)

        for record in privacy_records:
            rows.append(
                {
                    "구분": "개인정보 의심",
                    "파일명": "",
                    "시트명": "",
                    "기준/컬럼": record.get("컬럼명", ""),
                    "판단 내용": f"{record.get('점검 유형', '')} {record.get('감지 건수', '')}건",
                    "확인할 점": record.get("조치 안내", "공유 전에 필요 여부를 확인하세요."),
                }
            )

        for key, value in result.summary.items():
            rows.append(
                {
                    "구분": "처리 요약",
                    "파일명": "",
                    "시트명": "",
                    "기준/컬럼": SUMMARY_LABELS.get(key, key),
                    "판단 내용": value,
                    "확인할 점": "",
                }
            )

        if not rows:
            rows.append(
                {
                    "구분": "확인",
                    "파일명": "",
                    "시트명": "",
                    "기준/컬럼": "",
                    "판단 내용": "추가 확인 항목이 없습니다.",
                    "확인할 점": "첫 번째 결과 시트를 확인하세요.",
                }
            )
        if include_sensitive_details and "상세" not in columns:
            columns.append("상세")
        return pd.DataFrame(rows, columns=columns)

    def _guide_frame(self, result: JobResult) -> pd.DataFrame:
        issue_frame = self._issue_frame(result, include_sensitive_details=False)
        if issue_frame.empty:
            issue_count_rows = [{"확인 유형": "확인 필요 항목 없음", "건수": 0, "우선 조치": "결과 시트를 검토한 뒤 저장하거나 공유하세요."}]
        else:
            issue_count_rows = []
            counts = issue_frame.groupby("확인 유형").size().sort_values(ascending=False)
            for issue_type, count in counts.items():
                matching = issue_frame[issue_frame["확인 유형"] == issue_type]
                issue_count_rows.append(
                    {
                        "확인 유형": issue_type,
                        "건수": int(count),
                        "우선 조치": matching.iloc[0]["우선 조치"],
                    }
                )
        privacy_records = result.privacy_records or PrivacyScanner().scan_frame(result.result_frame)
        guide_rows = [
            {
                "먼저 볼 내용": "업무 결과",
                "현재 결과": f"{len(result.result_frame)}행, {len(result.result_frame.columns)}열",
                "바로 할 일": "첫 번째 `결과` 시트에서 업무 결과를 먼저 확인하세요.",
                "관련 시트": SHEET_RESULT,
            },
            {
                "먼저 볼 내용": "개인정보 의심",
                "현재 결과": self._privacy_summary_text(privacy_records),
                "바로 할 일": "공유 전 개인정보점검 시트에서 필요 여부를 확인하세요.",
                "관련 시트": SHEET_PRIVACY,
            },
            {
                "먼저 볼 내용": "확인 필요",
                "현재 결과": f"{len(result.issues)}건",
                "바로 할 일": "확인필요 시트에서 오류, 누락, 자동추천 확인 항목을 검토하세요.",
                "관련 시트": SHEET_ISSUES,
            },
            {
                "먼저 볼 내용": "자동 매칭/추천",
                "현재 결과": f"{len(result.mapping_records)}건",
                "바로 할 일": "컬럼명이 달랐던 업무는 자동추천근거 시트에서 맞게 붙었는지 확인하세요.",
                "관련 시트": SHEET_AUTO_EVIDENCE,
            },
        ]
        if issue_count_rows:
            guide_rows.extend(
                {
                    "먼저 볼 내용": row["확인 유형"],
                    "현재 결과": f"{row['건수']}건",
                    "바로 할 일": row["우선 조치"],
                    "관련 시트": SHEET_ISSUES,
                }
                for row in issue_count_rows
            )
        return pd.DataFrame(guide_rows)

    def _issue_frame(self, result: JobResult, include_sensitive_details: bool) -> pd.DataFrame:
        rows = []
        for issue in result.issues:
            row = {
                "심각도": SEVERITY_LABELS.get(issue.severity, issue.severity),
                "확인 유형": ISSUE_LABELS.get(issue.issue_type, issue.issue_type),
                "안내 문구": issue.message,
                "파일명": issue.file_name,
                "시트명": issue.sheet_name,
                "원본 행 번호": issue.row_number,
                "컬럼명": self._issue_column_text(issue),
                "우선 조치": ISSUE_ACTIONS.get(issue.issue_type, "원본 자료와 매핑 설정을 확인하세요."),
            }
            if include_sensitive_details:
                row["상세"] = json.dumps(issue.details, ensure_ascii=False, sort_keys=True)
            rows.append(row)
        columns = ["심각도", "확인 유형", "안내 문구", "파일명", "시트명", "원본 행 번호", "컬럼명", "우선 조치"]
        if include_sensitive_details:
            columns.append("상세")
        return pd.DataFrame(rows, columns=columns)

    def _error_rows_frame(self, result: JobResult) -> pd.DataFrame:
        if result.result_frame.empty or not result.issues:
            return pd.DataFrame()

        rows = []
        seen = set()
        for issue in result.issues:
            matching_indices = self._matching_result_indices(result.result_frame, issue)
            for index in matching_indices:
                key = (index, issue.issue_type, issue.column_name, issue.row_number)
                if key in seen:
                    continue
                seen.add(key)
                row = {
                    "심각도": SEVERITY_LABELS.get(issue.severity, issue.severity),
                    "확인 유형": ISSUE_LABELS.get(issue.issue_type, issue.issue_type),
                    "안내 문구": issue.message,
                    "파일명": issue.file_name,
                    "시트명": issue.sheet_name,
                    "원본 행 번호": issue.row_number,
                    "컬럼명": self._issue_column_text(issue),
                }
                row.update(result.result_frame.loc[index].to_dict())
                rows.append(row)
        return pd.DataFrame(rows)

    def _matching_result_indices(self, frame: pd.DataFrame, issue) -> list[int]:
        result_indices = issue.details.get("result_indices") if isinstance(issue.details, dict) else None
        if isinstance(result_indices, (list, tuple)):
            indices: list[int] = []
            for index in result_indices:
                try:
                    value = int(index)
                except (TypeError, ValueError):
                    continue
                if 0 <= value < len(frame):
                    indices.append(value)
            return indices
        if issue.row_number is None:
            return []

        mask = pd.Series(True, index=frame.index)
        matched_by_tracking = False
        if "원본 행 번호" in frame.columns:
            mask &= frame["원본 행 번호"].astype(str) == str(issue.row_number)
            matched_by_tracking = True
        if issue.file_name and "원본 파일명" in frame.columns:
            mask &= frame["원본 파일명"].astype(str) == str(issue.file_name)
        if issue.sheet_name and "원본 시트명" in frame.columns:
            mask &= frame["원본 시트명"].astype(str) == str(issue.sheet_name)
        if matched_by_tracking:
            indices = [int(index) for index in frame.index[mask]]
            if indices:
                return indices

        index = int(issue.row_number) - 2
        if 0 <= index < len(frame):
            return [index]
        return []

    def _summary_frame(self, result: JobResult) -> pd.DataFrame:
        rows = []
        for key, value in result.summary.items():
            rows.append({"항목": SUMMARY_LABELS.get(key, key), "값": value})
        if not rows:
            rows.append({"항목": "결과 행 수", "값": len(result.result_frame)})
        return pd.DataFrame(rows)

    def _mark_result_sheet(self, workbook, result: JobResult, privacy_records: list[dict[str, object]]) -> None:
        if SHEET_RESULT not in workbook.sheetnames:
            return
        sheet = workbook[SHEET_RESULT]
        privacy_fill = PatternFill("solid", fgColor="FDE68A")

        for record in privacy_records:
            column_index = self._find_column_index(sheet, str(record.get("컬럼명", "")))
            if column_index is None:
                continue
            message = (
                f"개인정보 의심: {record.get('점검 유형', '')}\n"
                f"위험도: {record.get('위험도', '')}\n"
                f"감지 건수: {record.get('감지 건수', '')}\n"
                f"{record.get('조치 안내', '')}"
            )
            self._mark_column(sheet, column_index, privacy_fill, message, max_cell_comments=30)

        for issue in result.issues:
            indices = self._matching_result_indices(result.result_frame, issue)
            columns_to_mark = self._issue_columns_to_mark(issue)
            message = (
                f"확인 필요: {ISSUE_LABELS.get(issue.issue_type, issue.issue_type)}\n"
                f"{issue.message}\n"
                f"조치: {ISSUE_ACTIONS.get(issue.issue_type, '원본 자료와 매핑 설정을 확인하세요.')}"
            )
            if indices:
                for index in indices:
                    excel_row = int(index) + 2
                    issue_fill = self._fill_for_issue(issue)
                    if columns_to_mark:
                        marked = False
                        for column_name in columns_to_mark:
                            column_index = self._find_column_index(sheet, column_name)
                            if column_index is None:
                                continue
                            self._mark_cell(sheet.cell(row=excel_row, column=column_index), issue_fill, message)
                            marked = True
                        if marked:
                            continue
                    if issue.column_name:
                        column_index = self._find_column_index(sheet, issue.column_name)
                        if column_index is not None:
                            self._mark_cell(sheet.cell(row=excel_row, column=column_index), issue_fill, message)
                            continue
                    for cell in sheet[excel_row]:
                        self._mark_cell(cell, issue_fill, message)
                continue
            if issue.column_name:
                column_index = self._find_column_index(sheet, issue.column_name)
                if column_index is not None:
                    self._mark_column(sheet, column_index, self._fill_for_issue(issue), message, max_cell_comments=10)

    def _fill_for_issue(self, issue) -> PatternFill:
        if issue.severity == "error":
            return PatternFill("solid", fgColor="FCA5A5")
        return PatternFill("solid", fgColor="FDE68A")

    def _issue_columns_to_mark(self, issue) -> list[str]:
        columns: list[str] = []
        if issue.column_name:
            columns.append(str(issue.column_name))
        result_columns = issue.details.get("result_columns") if isinstance(issue.details, dict) else None
        if isinstance(result_columns, (list, tuple)):
            columns.extend(str(column) for column in result_columns if str(column).strip())
        seen: set[str] = set()
        unique_columns: list[str] = []
        for column in columns:
            if column in seen:
                continue
            seen.add(column)
            unique_columns.append(column)
        return unique_columns

    def _issue_column_text(self, issue) -> str | None:
        columns = self._issue_columns_to_mark(issue)
        if not columns:
            return None
        return ", ".join(columns)

    def _privacy_summary_text(self, records: list[dict[str, object]]) -> str:
        if not records:
            return "의심 항목 없음"
        counts: dict[str, int] = {}
        for record in records:
            category = str(record.get("점검 유형", "") or "기타")
            try:
                count = int(record.get("감지 건수", 0) or 0)
            except (TypeError, ValueError):
                count = 0
            counts[category] = counts.get(category, 0) + count
        return ", ".join(f"{category} {count}건" for category, count in counts.items())

    def _mark_column(self, sheet, column_index: int, fill: PatternFill, message: str, max_cell_comments: int) -> None:
        self._mark_cell(sheet.cell(row=1, column=column_index), fill, message)
        comments_left = max_cell_comments
        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            if cell.value in (None, ""):
                continue
            cell.fill = fill
            if comments_left > 0 and cell.comment is None:
                cell.comment = make_comment(message)
                comments_left -= 1

    def _mark_cell(self, cell, fill: PatternFill, message: str) -> None:
        cell.fill = fill
        if cell.comment is None:
            cell.comment = make_comment(message)

    def _find_column_index(self, sheet, column_name: str) -> int | None:
        for cell in sheet[1]:
            if str(cell.value) == column_name:
                return int(cell.column)
        return None

    def _style_workbook(self, workbook) -> None:
        header_fill = PatternFill("solid", fgColor="1F6F5F")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            if sheet.max_row > 1 and sheet.max_column > 0:
                sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, min(len(value), 42))
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                sheet.column_dimensions[column_letter].width = max(10, min(max_length + 2, 44))
