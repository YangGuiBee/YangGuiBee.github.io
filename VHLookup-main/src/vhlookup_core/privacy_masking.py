from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from vhlookup_core.excel_comments import make_comment
from vhlookup_core.header import HeaderDetector
from vhlookup_core.loader import ExcelLoader
from vhlookup_core.normalization import display_value, is_blank, normalize_header
from vhlookup_core.sheet import SheetDetector


SHEET_MASKED = "마스킹결과"
SHEET_GUIDE = "먼저확인"
SHEET_RECORDS = "마스킹내역"

NAME_HEADERS = ("성명", "이름", "직원명", "담당자", "담당자명", "대상자명", "신청자명", "예금주")
PHONE_HEADERS = ("연락처", "전화번호", "휴대폰", "핸드폰", "내선번호", "phone", "mobile")
EMAIL_HEADERS = ("이메일", "메일", "email", "e-mail")
RRN_HEADERS = ("주민등록", "주민번호", "외국인등록", "외국인번호", "고유식별", "rrn")
ACCOUNT_HEADERS = ("계좌", "계좌번호", "account")
ADDRESS_HEADERS = ("주소", "거주지", "소재지", "address")
ID_HEADERS = ("사번", "직원번호", "학번", "회원번호", "employeeid", "employee_id")
GENDER_HEADERS = ("성별", "남녀", "gender", "sex")
AGE_HEADERS = ("나이", "연령", "만나이", "age")
BIRTHDATE_HEADERS = ("생년월일", "생년", "생일", "출생일", "출생연월일", "birth", "birthday", "dateofbirth", "dob")

RRN_RE = re.compile(r"(?<!\d)(\d{6})-?([1-8])(\d{6})(?!\d)")
PHONE_RE = re.compile(r"(?<!\d)(0\d{1,2})-?(\d{3,4})-?(\d{4})(?!\d)")
EMAIL_RE = re.compile(r"([\w.+-]+)@([\w-]+(?:\.[\w-]+)+)")
BIRTHDATE_KEYWORDS = r"(?:생년월일|생년|생일|출생일|출생연월일|birth\s*date|birthday|date\s*of\s*birth|dob)"
BIRTHDATE_VALUE_PATTERN = (
    r"(?:\d{4}[./-]\d{1,2}[./-]\d{1,2}"
    r"|\d{2}[./-]\d{1,2}[./-]\d{1,2}"
    r"|\d{2,4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일"
    r"|\d{8}"
    r"|\d{6})"
)
BIRTHDATE_PREFIX_CONTEXT_RE = re.compile(
    rf"({BIRTHDATE_KEYWORDS}\s*[:：]?\s*)({BIRTHDATE_VALUE_PATTERN})",
    re.IGNORECASE,
)
BIRTHDATE_SUFFIX_CONTEXT_RE = re.compile(
    rf"({BIRTHDATE_VALUE_PATTERN})\s*(생|출생)",
    re.IGNORECASE,
)
LONG_DIGIT_RE = re.compile(r"\d{5,}")


@dataclass(frozen=True)
class PrivacyMaskingResult:
    masked_frame: pd.DataFrame
    records: pd.DataFrame
    summary: dict[str, Any]


class PrivacyMaskingEngine:
    def __init__(
        self,
        loader: ExcelLoader | None = None,
        header_detector: HeaderDetector | None = None,
        sheet_detector: SheetDetector | None = None,
    ) -> None:
        self.loader = loader or ExcelLoader()
        self.header_detector = header_detector or HeaderDetector()
        self.sheet_detector = sheet_detector or SheetDetector(self.header_detector)

    def mask_file(self, path: str | Path) -> PrivacyMaskingResult:
        table, metadata = self._load_table(path)
        result = self.mask_frame(table)
        result.summary.update(metadata)
        return result

    def mask_frame(self, frame: pd.DataFrame) -> PrivacyMaskingResult:
        masked = frame.copy()
        records: list[dict[str, Any]] = []

        for column_index, column in enumerate(masked.columns):
            column_types = self._types_for_column(str(column))
            for row_index, value in masked[column].items():
                masked_value, applied_types = self._mask_value(value, column_types)
                if masked_value == value and not applied_types:
                    continue
                masked.at[row_index, column] = masked_value
                for masking_type in applied_types:
                    records.append(
                        {
                            "행 번호": int(row_index) + 2,
                            "컬럼명": str(column),
                            "마스킹 유형": masking_type,
                            "처리 내용": self._action_text(masking_type),
                            "_row_index": int(row_index),
                            "_column_index": int(column_index),
                        }
                    )

        record_frame = pd.DataFrame(records)
        public_records = (
            record_frame.drop(columns=["_row_index", "_column_index"])
            if not record_frame.empty
            else pd.DataFrame(columns=["행 번호", "컬럼명", "마스킹 유형", "처리 내용"])
        )
        summary = {
            "workflow": "개인정보 마스킹",
            "row_count": len(masked),
            "column_count": len(masked.columns),
            "masked_cell_count": int(record_frame[["_row_index", "_column_index"]].drop_duplicates().shape[0])
            if not record_frame.empty
            else 0,
            "masking_record_count": len(record_frame),
        }
        result = PrivacyMaskingResult(masked_frame=masked, records=public_records, summary=summary)
        object.__setattr__(result, "_internal_records", record_frame)
        return result

    def write_xlsx(self, path: str | Path, output_path: str | Path) -> Path:
        result = self.mask_file(path)
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            result.masked_frame.to_excel(writer, sheet_name=SHEET_MASKED, index=False)
            result.records.to_excel(writer, sheet_name=SHEET_RECORDS, index=False)
            self._style_workbook(writer.book)
            self._mark_masked_cells(writer.book, result)
        return output

    def _load_table(self, path: str | Path) -> tuple[pd.DataFrame, dict[str, Any]]:
        source = self.loader.load(path)
        sheet = self.sheet_detector.select(source)
        detection = self.header_detector.detect(sheet)
        table = self.header_detector.apply(sheet, detection)
        return table, {
            "file_name": Path(path).name,
            "sheet_name": sheet.name,
            "header_row_number": detection.header_row_number,
            "data_start_row_number": detection.data_start_row_number,
        }

    def _types_for_column(self, column: str) -> set[str]:
        normalized = normalize_header(column)
        types: set[str] = set()
        if self._matches(normalized, NAME_HEADERS):
            types.add("이름")
        if self._matches(normalized, PHONE_HEADERS):
            types.add("연락처")
        if self._matches(normalized, EMAIL_HEADERS):
            types.add("이메일")
        if self._matches(normalized, RRN_HEADERS):
            types.add("주민등록번호")
        if self._matches(normalized, ACCOUNT_HEADERS):
            types.add("계좌번호")
        if self._matches(normalized, ADDRESS_HEADERS):
            types.add("주소")
        if self._matches(normalized, ID_HEADERS):
            types.add("식별번호")
        if self._matches(normalized, GENDER_HEADERS):
            types.add("성별")
        if self._matches(normalized, AGE_HEADERS):
            types.add("나이")
        if self._matches(normalized, BIRTHDATE_HEADERS):
            types.add("생년월일")
        return types

    def _mask_value(self, value: object, column_types: set[str]) -> tuple[object, list[str]]:
        if is_blank(value):
            return value, []
        text = display_value(value)
        original = text
        applied: list[str] = []

        text, changed = self._mask_rrn_patterns(text)
        if changed:
            applied.append("주민등록번호")
        text, changed = self._mask_phone_patterns(text)
        if changed:
            applied.append("연락처")
        text, changed = self._mask_email_patterns(text)
        if changed:
            applied.append("이메일")
        text, changed = self._mask_birthdate_context_patterns(text)
        if changed:
            applied.append("생년월일")

        if "주민등록번호" in column_types:
            text, changed = self._mask_rrn_patterns(text, fallback=True)
            if changed:
                applied.append("주민등록번호")
        if "이름" in column_types:
            masked_name = mask_name(text)
            if masked_name != text:
                text = masked_name
                applied.append("이름")
        if "연락처" in column_types:
            masked_phone = mask_phone_text(text)
            if masked_phone != text:
                text = masked_phone
                applied.append("연락처")
        if "이메일" in column_types:
            text, changed = self._mask_email_patterns(text, fallback=True)
            if changed:
                applied.append("이메일")
        if "계좌번호" in column_types:
            masked_account = mask_long_digits(text)
            if masked_account != text:
                text = masked_account
                applied.append("계좌번호")
        if "주소" in column_types:
            masked_address = mask_address(text)
            if masked_address != text:
                text = masked_address
                applied.append("주소")
        if "식별번호" in column_types:
            masked_identifier = mask_identifier(text)
            if masked_identifier != text:
                text = masked_identifier
                applied.append("식별번호")
        if "성별" in column_types:
            masked_gender = mask_placeholder(text, "성별")
            if masked_gender != text:
                text = masked_gender
                applied.append("성별")
        if "나이" in column_types:
            masked_age = mask_placeholder(text, "나이")
            if masked_age != text:
                text = masked_age
                applied.append("나이")
        if "생년월일" in column_types:
            masked_birthdate = mask_placeholder(text, "생년월일")
            if masked_birthdate != text:
                text = masked_birthdate
                applied.append("생년월일")

        deduped = list(dict.fromkeys(applied))
        return (text if text != original else value), deduped

    def _mask_rrn_patterns(self, text: str, fallback: bool = False) -> tuple[str, bool]:
        masked, count = RRN_RE.subn(lambda match: "******-*******", text)
        if count or not fallback:
            return masked, bool(count)
        digits = re.sub(r"\D", "", text)
        if len(digits) == 13 and digits[6] in "12345678":
            return "******-*******", True
        return text, False

    def _mask_phone_patterns(self, text: str) -> tuple[str, bool]:
        masked, count = PHONE_RE.subn("***", text)
        return masked, bool(count)

    def _mask_email_patterns(self, text: str, fallback: bool = False) -> tuple[str, bool]:
        masked, count = EMAIL_RE.subn("***", text)
        if count or not fallback or "@" not in text:
            return masked, bool(count)
        local, _, domain = text.partition("@")
        if local and domain:
            return "***", True
        return text, False

    def _mask_birthdate_context_patterns(self, text: str) -> tuple[str, bool]:
        masked, prefix_count = BIRTHDATE_PREFIX_CONTEXT_RE.subn(lambda match: f"{match.group(1)}***", text)
        masked, suffix_count = BIRTHDATE_SUFFIX_CONTEXT_RE.subn(lambda match: f"***{match.group(2)}", masked)
        return masked, bool(prefix_count or suffix_count)

    def _guide_frame(self, result: PrivacyMaskingResult) -> pd.DataFrame:
        rows = [
            {"항목": "먼저 볼 내용", "내용": "첫 번째 마스킹결과 시트에서 개인정보가 가려진 결과를 확인하세요."},
            {"항목": "마스킹 셀 수", "내용": result.summary.get("masked_cell_count", 0)},
            {"항목": "마스킹 처리 건수", "내용": result.summary.get("masking_record_count", 0)},
            {"항목": "원본 파일", "내용": result.summary.get("file_name", "")},
            {"항목": "주의", "내용": "원본 파일은 수정하지 않았습니다. 결과 파일만 새로 만들었습니다."},
            {"항목": "추가 확인", "내용": "성별, 나이, 생년월일처럼 컬럼명으로 판단하는 항목은 마스킹결과 시트에서 사람이 한 번 더 확인하세요."},
        ]
        return pd.DataFrame(rows)

    def _mark_masked_cells(self, workbook, result: PrivacyMaskingResult) -> None:
        records = getattr(result, "_internal_records", pd.DataFrame())
        if records.empty or SHEET_MASKED not in workbook.sheetnames:
            return
        sheet = workbook[SHEET_MASKED]
        fill = PatternFill("solid", fgColor="FDE68A")
        grouped = records.groupby(["_row_index", "_column_index"])["마스킹 유형"].apply(lambda values: ", ".join(sorted(set(values))))
        for (row_index, column_index), masking_types in grouped.items():
            cell = sheet.cell(row=int(row_index) + 2, column=int(column_index) + 1)
            cell.fill = fill
            cell.comment = make_comment(f"개인정보 마스킹 적용\n유형: {masking_types}\n원본 값은 결과 파일에 저장하지 않았습니다.")

    def _style_workbook(self, workbook) -> None:
        header_fill = PatternFill("solid", fgColor="1F6F5F")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            if sheet.max_row > 1 and sheet.max_column > 0:
                sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, min(len(value), 42))
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                sheet.column_dimensions[column_letter].width = max(10, min(max_length + 2, 44))

    def _matches(self, normalized: str, headers: tuple[str, ...]) -> bool:
        return any(keyword and normalize_header(keyword) in normalized for keyword in headers)

    def _action_text(self, masking_type: str) -> str:
        actions = {
            "주민등록번호": "주민등록번호 또는 외국인등록번호 전체를 가렸습니다.",
            "이름": "이름 값을 전체 가렸습니다.",
            "연락처": "연락처 값을 전체 가렸습니다.",
            "이메일": "이메일 주소 전체를 가렸습니다.",
            "계좌번호": "계좌번호 값을 전체 가렸습니다.",
            "주소": "주소 값을 전체 가렸습니다.",
            "식별번호": "식별번호 앞부분을 가렸습니다.",
            "성별": "성별 값을 전체 가렸습니다.",
            "나이": "나이 또는 연령 값을 전체 가렸습니다.",
            "생년월일": "생년월일 값을 전체 가렸습니다.",
        }
        return actions.get(masking_type, "개인정보 의심 값을 가렸습니다.")


def mask_name(text: str) -> str:
    return mask_placeholder(text, "성명")


def mask_phone_text(text: str) -> str:
    masked, count = PHONE_RE.subn("***", text)
    if count:
        return masked
    digits = re.sub(r"\D", "", text)
    if 7 <= len(digits) <= 11:
        return "***"
    return text


def mask_long_digits(text: str) -> str:
    return "***" if LONG_DIGIT_RE.search(text) else text


def mask_address(text: str) -> str:
    return mask_placeholder(text, "주소")


def mask_identifier(text: str) -> str:
    if len(text) <= 2:
        return "*" * len(text)
    return f"{'*' * (len(text) - 2)}{text[-2:]}"


def mask_placeholder(text: str, label: str) -> str:
    return "***" if text.strip() else text
