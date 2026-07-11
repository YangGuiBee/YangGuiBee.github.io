from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from vhlookup_core.excel_comments import make_comment
from vhlookup_core.header import HeaderDetector
from vhlookup_core.keys import KeyRecommender, build_key_series, duplicate_ratio
from vhlookup_core.loader import ExcelLoader
from vhlookup_core.mapper import ColumnMapper
from vhlookup_core.normalization import display_value, is_blank, normalize_header, normalize_key_parts
from vhlookup_core.sheet import SheetDetector


@dataclass
class CellComment:
    row_index: int
    column_name: str
    message: str
    fill: str = "FDE68A"


@dataclass
class WorkbookDiffResult:
    before_frame: pd.DataFrame
    after_frame: pd.DataFrame
    diff_frame: pd.DataFrame
    column_frame: pd.DataFrame
    row_frame: pd.DataFrame
    summary: dict[str, Any]
    comments: list[CellComment] = field(default_factory=list)
    before_to_after: dict[str, str] = field(default_factory=dict)


class WorkbookDiffEngine:
    def __init__(
        self,
        loader: ExcelLoader | None = None,
        header_detector: HeaderDetector | None = None,
        sheet_detector: SheetDetector | None = None,
        mapper: ColumnMapper | None = None,
        key_recommender: KeyRecommender | None = None,
    ) -> None:
        self.loader = loader or ExcelLoader()
        self.header_detector = header_detector or HeaderDetector()
        self.sheet_detector = sheet_detector or SheetDetector(self.header_detector)
        self.mapper = mapper or ColumnMapper()
        self.key_recommender = key_recommender or KeyRecommender(self.mapper)

    def compare_files(
        self,
        before_file: str | Path,
        after_file: str | Path,
        scan_rows: int | str = 30,
        column_mapping_override: dict[str, str] | None = None,
        key_columns: tuple[str, ...] | None = None,
    ) -> WorkbookDiffResult:
        before_path = Path(before_file)
        after_path = Path(after_file)
        before_frame, before_sheet = self._load_table(before_path, scan_rows)
        after_frame, after_sheet = self._load_table(after_path, scan_rows)

        before_to_after = self.suggest_column_mapping(before_frame, after_frame)
        if column_mapping_override:
            before_to_after.update(
                {
                    before_column: after_column
                    for before_column, after_column in column_mapping_override.items()
                    if before_column in before_frame.columns and after_column in after_frame.columns
                }
            )
        after_aligned = pd.DataFrame(index=after_frame.index)
        for before_column, after_column in before_to_after.items():
            after_aligned[before_column] = after_frame[after_column]

        common_columns = [column for column in before_frame.columns if column in after_aligned.columns]
        before_only_columns = [column for column in before_frame.columns if column not in common_columns]
        after_only_columns = [column for column in after_frame.columns if column not in set(before_to_after.values())]
        resolved_key_columns = tuple(column for column in (key_columns or ()) if column in common_columns)
        if not resolved_key_columns:
            resolved_key_columns = self._choose_key_columns(before_frame, after_aligned, common_columns)
        compare_columns = [column for column in common_columns if column not in resolved_key_columns]

        diff_rows: list[dict[str, Any]] = []
        row_rows: list[dict[str, Any]] = []
        comments: list[CellComment] = []

        if resolved_key_columns:
            key_normalization = self._infer_key_normalization(before_frame, after_aligned, resolved_key_columns)
            self._compare_by_key(
                before_frame,
                after_frame,
                after_aligned,
                before_to_after,
                resolved_key_columns,
                key_normalization,
                compare_columns,
                diff_rows,
                row_rows,
                comments,
            )
            compare_basis = "키 컬럼: " + ", ".join(resolved_key_columns)
        else:
            self._compare_by_position(
                before_frame,
                after_frame,
                after_aligned,
                before_to_after,
                compare_columns,
                diff_rows,
                row_rows,
                comments,
            )
            compare_basis = "행 순서"

        for row_index in range(len(after_frame)):
            if row_index not in {comment.row_index for comment in comments}:
                continue

        column_rows = []
        for column in common_columns:
            column_rows.append(
                {
                    "상태": "양쪽 모두 있음",
                    "전 파일 컬럼": column,
                    "후 파일 컬럼": before_to_after.get(column, ""),
                }
            )
        for column in before_only_columns:
            column_rows.append({"상태": "후 파일에 컬럼 없음", "전 파일 컬럼": column, "후 파일 컬럼": ""})
        for column in after_only_columns:
            column_rows.append({"상태": "전 파일에 없던 컬럼", "전 파일 컬럼": "", "후 파일 컬럼": column})

        summary = {
            "workflow": "전/후 파일 검증",
            "before_file": str(before_path),
            "after_file": str(after_path),
            "before_sheet": before_sheet,
            "after_sheet": after_sheet,
            "compare_basis": compare_basis,
            "before_rows": len(before_frame),
            "after_rows": len(after_frame),
            "matched_column_count": len(common_columns),
            "changed_cell_count": sum(1 for row in diff_rows if row["상태"] == "값 다름"),
            "missing_row_count": sum(1 for row in row_rows if row["상태"] == "후 파일에 행 없음"),
            "added_row_count": sum(1 for row in row_rows if row["상태"] == "전 파일에 없던 행"),
            "before_only_column_count": len(before_only_columns),
            "after_only_column_count": len(after_only_columns),
        }
        return WorkbookDiffResult(
            before_frame=before_frame,
            after_frame=after_frame,
            diff_frame=pd.DataFrame(diff_rows),
            column_frame=pd.DataFrame(column_rows),
            row_frame=pd.DataFrame(row_rows),
            summary=summary,
            comments=comments,
            before_to_after=before_to_after,
        )

    def suggest_column_mapping(self, before_frame: pd.DataFrame, after_frame: pd.DataFrame) -> dict[str, str]:
        column_mapping = self.mapper.map_columns(list(after_frame.columns), list(before_frame.columns), threshold=0.72)
        return self._with_data_overlap_column_mapping(before_frame, after_frame, dict(column_mapping.target_to_source))

    def _load_table(self, path: Path, scan_rows: int | str) -> tuple[pd.DataFrame, str]:
        source = self.loader.load(path)
        sheet = self.sheet_detector.select(source)
        detection = self.header_detector.detect(sheet, scan_rows=scan_rows)
        return self.header_detector.apply(sheet, detection), sheet.name

    def _choose_key_columns(
        self,
        before_frame: pd.DataFrame,
        after_frame: pd.DataFrame,
        common_columns: list[str],
    ) -> tuple[str, ...]:
        for column in common_columns:
            normalized = normalize_header(column)
            if any(keyword in normalized for keyword in ("사번", "직원번호", "기관코드", "학교코드", "코드", "id", "번호")):
                if self._is_good_key(before_frame[column], after_frame[column]):
                    return (column,)
        if not common_columns:
            return ()
        candidates = self.key_recommender.recommend(
            before_frame.loc[:, common_columns],
            after_frame.loc[:, common_columns],
            limit=1,
        )
        if candidates and candidates[0].score >= 0.62:
            return candidates[0].reference_columns
        return ()

    def _is_good_key(self, before: pd.Series, after: pd.Series) -> bool:
        before_values = before.map(lambda value: normalize_key_parts([value], "loose_numeric"))
        after_values = after.map(lambda value: normalize_key_parts([value], "loose_numeric"))
        before_non_blank = before_values[before_values != ""]
        after_non_blank = after_values[after_values != ""]
        if before_non_blank.empty or after_non_blank.empty:
            return False
        if before_non_blank.duplicated().any() or after_non_blank.duplicated().any():
            return False
        overlap = len(set(before_non_blank).intersection(set(after_non_blank)))
        return overlap / max(min(before_non_blank.nunique(), after_non_blank.nunique()), 1) >= 0.5

    def _infer_key_normalization(
        self,
        before_frame: pd.DataFrame,
        after_aligned: pd.DataFrame,
        key_columns: tuple[str, ...],
    ) -> str:
        before_text = build_key_series(before_frame, key_columns)
        after_text = build_key_series(after_aligned, key_columns)
        before_loose = build_key_series(before_frame, key_columns, normalization="loose_numeric")
        after_loose = build_key_series(after_aligned, key_columns, normalization="loose_numeric")
        exact_overlap = len(set(before_text[before_text != ""]).intersection(set(after_text[after_text != ""])))
        loose_overlap = len(set(before_loose[before_loose != ""]).intersection(set(after_loose[after_loose != ""])))
        return "loose_numeric" if loose_overlap > exact_overlap else "text"

    def _with_data_overlap_column_mapping(
        self,
        before_frame: pd.DataFrame,
        after_frame: pd.DataFrame,
        mapping: dict[str, str],
    ) -> dict[str, str]:
        used_after_columns = set(mapping.values())
        for before_column in before_frame.columns:
            if before_column in mapping:
                continue
            best_after_column = ""
            best_score = 0.0
            for after_column in after_frame.columns:
                if after_column in used_after_columns:
                    continue
                score = self._column_data_score(before_frame[before_column], after_frame[after_column])
                if score > best_score:
                    best_score = score
                    best_after_column = str(after_column)
            if best_after_column and best_score >= 0.75:
                mapping[str(before_column)] = best_after_column
                used_after_columns.add(best_after_column)
        return mapping

    def _column_data_score(self, before: pd.Series, after: pd.Series) -> float:
        before_text = before.to_frame().apply(lambda row: normalize_key_parts(row.tolist()), axis=1)
        after_text = after.to_frame().apply(lambda row: normalize_key_parts(row.tolist()), axis=1)
        before_loose = before.to_frame().apply(lambda row: normalize_key_parts(row.tolist(), "loose_numeric"), axis=1)
        after_loose = after.to_frame().apply(lambda row: normalize_key_parts(row.tolist(), "loose_numeric"), axis=1)

        before_non_blank = before_text[before_text != ""]
        after_non_blank = after_text[after_text != ""]
        before_loose_non_blank = before_loose[before_loose != ""]
        after_loose_non_blank = after_loose[after_loose != ""]
        if before_non_blank.empty or after_non_blank.empty:
            return 0.0

        exact_overlap = len(set(before_non_blank).intersection(set(after_non_blank)))
        exact_denominator = max(min(before_non_blank.nunique(), after_non_blank.nunique()), 1)
        loose_overlap = len(set(before_loose_non_blank).intersection(set(after_loose_non_blank)))
        loose_denominator = max(min(before_loose_non_blank.nunique(), after_loose_non_blank.nunique()), 1)
        overlap_ratio = max(exact_overlap / exact_denominator, (loose_overlap / loose_denominator) * 0.92)
        completeness = min(
            len(before_non_blank) / max(len(before), 1),
            len(after_non_blank) / max(len(after), 1),
        )
        uniqueness = 1 - max(duplicate_ratio(before_non_blank), duplicate_ratio(after_non_blank))
        return 0.55 * overlap_ratio + 0.20 * completeness + 0.25 * uniqueness

    def _compare_by_key(
        self,
        before_frame: pd.DataFrame,
        after_frame: pd.DataFrame,
        after_aligned: pd.DataFrame,
        before_to_after: dict[str, str],
        key_columns: tuple[str, ...],
        key_normalization: str,
        compare_columns: list[str],
        diff_rows: list[dict[str, Any]],
        row_rows: list[dict[str, Any]],
        comments: list[CellComment],
    ) -> None:
        before_keys = build_key_series(before_frame, key_columns, key_normalization)
        after_keys = build_key_series(after_aligned, key_columns, key_normalization)
        before_map = {key: index for index, key in before_keys.items() if key}
        after_map = {key: index for index, key in after_keys.items() if key}
        all_keys = list(dict.fromkeys([*before_map.keys(), *after_map.keys()]))

        for key in all_keys:
            before_index = before_map.get(key)
            after_index = after_map.get(key)
            if before_index is None:
                row_rows.append({"상태": "전 파일에 없던 행", "비교 기준": key, "전 파일 행": "", "후 파일 행": int(after_index) + 2})
                after_column = str(after_frame.columns[0])
                diff_rows.append(
                    {
                        "상태": "전 파일에 없던 행",
                        "비교 기준": key,
                        "컬럼명": "",
                        "전 값": "",
                        "후 값": "",
                        "후 파일 셀": f"{after_column}{int(after_index) + 2}",
                    }
                )
                comments.append(
                    CellComment(
                        row_index=int(after_index),
                        column_name=after_column,
                        message="전 파일에 없던 행입니다.",
                        fill="BFDBFE",
                    )
                )
                continue
            if after_index is None:
                row_rows.append({"상태": "후 파일에 행 없음", "비교 기준": key, "전 파일 행": int(before_index) + 2, "후 파일 행": ""})
                diff_rows.append(
                    {
                        "상태": "후 파일에 행 없음",
                        "비교 기준": key,
                        "컬럼명": "",
                        "전 값": "",
                        "후 값": "",
                        "후 파일 셀": "",
                    }
                )
                continue
            row_rows.append(
                {"상태": "양쪽 모두 있음", "비교 기준": key, "전 파일 행": int(before_index) + 2, "후 파일 행": int(after_index) + 2}
            )
            self._compare_cells(
                before_frame,
                after_aligned,
                before_to_after,
                int(before_index),
                int(after_index),
                key,
                compare_columns,
                diff_rows,
                comments,
            )

    def _compare_by_position(
        self,
        before_frame: pd.DataFrame,
        after_frame: pd.DataFrame,
        after_aligned: pd.DataFrame,
        before_to_after: dict[str, str],
        compare_columns: list[str],
        diff_rows: list[dict[str, Any]],
        row_rows: list[dict[str, Any]],
        comments: list[CellComment],
    ) -> None:
        max_rows = max(len(before_frame), len(after_frame))
        for index in range(max_rows):
            key = f"{index + 2}행"
            if index >= len(before_frame):
                row_rows.append({"상태": "전 파일에 없던 행", "비교 기준": key, "전 파일 행": "", "후 파일 행": index + 2})
                after_column = str(after_frame.columns[0])
                diff_rows.append(
                    {
                        "상태": "전 파일에 없던 행",
                        "비교 기준": key,
                        "컬럼명": "",
                        "전 값": "",
                        "후 값": "",
                        "후 파일 셀": f"{after_column}{index + 2}",
                    }
                )
                comments.append(
                    CellComment(
                        row_index=index,
                        column_name=after_column,
                        message="전 파일에 없던 행입니다.",
                        fill="BFDBFE",
                    )
                )
                continue
            if index >= len(after_frame):
                row_rows.append({"상태": "후 파일에 행 없음", "비교 기준": key, "전 파일 행": index + 2, "후 파일 행": ""})
                diff_rows.append(
                    {
                        "상태": "후 파일에 행 없음",
                        "비교 기준": key,
                        "컬럼명": "",
                        "전 값": "",
                        "후 값": "",
                        "후 파일 셀": "",
                    }
                )
                continue
            row_rows.append({"상태": "양쪽 모두 있음", "비교 기준": key, "전 파일 행": index + 2, "후 파일 행": index + 2})
            self._compare_cells(
                before_frame,
                after_aligned,
                before_to_after,
                index,
                index,
                key,
                compare_columns,
                diff_rows,
                comments,
            )

    def _compare_cells(
        self,
        before_frame: pd.DataFrame,
        after_aligned: pd.DataFrame,
        before_to_after: dict[str, str],
        before_index: int,
        after_index: int,
        key: str,
        compare_columns: list[str],
        diff_rows: list[dict[str, Any]],
        comments: list[CellComment],
    ) -> None:
        for column in compare_columns:
            before_value = before_frame.at[before_index, column]
            after_value = after_aligned.at[after_index, column]
            if self._same_value(before_value, after_value):
                continue
            after_column = before_to_after.get(column, column)
            cell_address = f"{after_column}{after_index + 2}"
            before_text = display_value(before_value)
            after_text = display_value(after_value)
            diff_rows.append(
                {
                    "상태": "값 다름",
                    "비교 기준": key,
                    "컬럼명": column,
                    "전 값": before_text,
                    "후 값": after_text,
                    "후 파일 셀": cell_address,
                }
            )
            comments.append(
                CellComment(
                    row_index=after_index,
                    column_name=after_column,
                    message=f"전 값: {before_text}\n후 값: {after_text}",
                )
            )

    def _same_value(self, before_value: object, after_value: object) -> bool:
        if is_blank(before_value) and is_blank(after_value):
            return True
        return display_value(before_value) == display_value(after_value)


class WorkbookDiffReportWriter:
    ADDED_ROW_FILL = "BFDBFE"
    MISSING_ROW_FILL = "FECACA"

    def write_xlsx(self, result: WorkbookDiffResult, path: str | Path) -> Path:
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            self._annotated_after_frame(result).to_excel(writer, sheet_name="후파일_메모", index=False)
            self._review_frame(result).to_excel(writer, sheet_name="확인사항", index=False)
            self._apply_comments(writer.book, result)
            self._apply_row_highlights(writer.book, result)
            self._style_workbook(writer.book)
        return output_path

    def _annotated_after_frame(self, result: WorkbookDiffResult) -> pd.DataFrame:
        display_columns = list(result.after_frame.columns) or list(result.before_frame.columns)
        annotated = result.after_frame.loc[:, display_columns].copy() if display_columns else result.after_frame.copy()
        missing_rows = self._missing_row_records(result, display_columns)
        if not missing_rows:
            return annotated
        return pd.concat([annotated, pd.DataFrame(missing_rows, columns=display_columns)], ignore_index=True)

    def _missing_row_records(self, result: WorkbookDiffResult, display_columns: list[str]) -> list[dict[str, object]]:
        if result.before_frame.empty or result.row_frame.empty:
            return []

        missing = result.row_frame[result.row_frame["상태"] == "후 파일에 행 없음"]
        if missing.empty:
            return []

        after_to_before = {after_column: before_column for before_column, after_column in result.before_to_after.items()}
        rows: list[dict[str, object]] = []
        for _, missing_row in missing.iterrows():
            try:
                before_index = int(missing_row.get("전 파일 행", "")) - 2
            except (TypeError, ValueError):
                continue
            if not 0 <= before_index < len(result.before_frame):
                continue

            row: dict[str, object] = {}
            for display_column in display_columns:
                before_column = after_to_before.get(
                    display_column,
                    display_column if display_column in result.before_frame.columns else "",
                )
                row[display_column] = result.before_frame.at[before_index, before_column] if before_column else ""
            rows.append(row)
        return rows

    def _review_frame(self, result: WorkbookDiffResult) -> pd.DataFrame:
        rows: list[dict[str, object]] = []
        columns = ["구분", "상태", "비교 기준", "컬럼명", "전 값", "후 값", "위치", "확인할 점"]

        for _, diff_row in result.diff_frame.iterrows():
            if not str(diff_row.get("컬럼명", "") or "").strip():
                continue
            rows.append(
                {
                    "구분": "값 변경",
                    "상태": diff_row.get("상태", ""),
                    "비교 기준": diff_row.get("비교 기준", ""),
                    "컬럼명": diff_row.get("컬럼명", ""),
                    "전 값": diff_row.get("전 값", ""),
                    "후 값": diff_row.get("후 값", ""),
                    "위치": diff_row.get("후 파일 셀", ""),
                    "확인할 점": "첫 시트의 색칠된 셀 메모를 확인하세요.",
                }
            )

        for _, row_diff in result.row_frame.iterrows():
            status = str(row_diff.get("상태", ""))
            detail = ""
            if status == "후 파일에 행 없음":
                detail = self._before_row_detail(result, row_diff.get("전 파일 행", ""))
            rows.append(
                {
                    "구분": "행 추가/누락",
                    "상태": status,
                    "비교 기준": row_diff.get("비교 기준", ""),
                    "컬럼명": "",
                    "전 값": detail,
                    "후 값": "",
                    "위치": f"전 {row_diff.get('전 파일 행', '')} / 후 {row_diff.get('후 파일 행', '')}",
                    "확인할 점": "한쪽 파일에만 있는 행입니다. 누락인지 추가 대상인지 확인하세요.",
                }
            )

        for _, column_diff in result.column_frame.iterrows():
            rows.append(
                {
                    "구분": "컬럼 변경",
                    "상태": column_diff.get("상태", ""),
                    "비교 기준": "",
                    "컬럼명": f"{column_diff.get('전 파일 컬럼', '')} / {column_diff.get('후 파일 컬럼', '')}",
                    "전 값": "",
                    "후 값": "",
                    "위치": "",
                    "확인할 점": "컬럼명이 바뀐 것인지, 실제로 추가/삭제된 것인지 확인하세요.",
                }
            )

        for key, value in result.summary.items():
            rows.append(
                {
                    "구분": "점검 요약",
                    "상태": key,
                    "비교 기준": "",
                    "컬럼명": "",
                    "전 값": "",
                    "후 값": value,
                    "위치": "",
                    "확인할 점": "",
                }
            )

        if not rows:
            rows.append(
                {
                    "구분": "확인",
                    "상태": "차이 없음",
                    "비교 기준": "",
                    "컬럼명": "",
                    "전 값": "",
                    "후 값": "",
                    "위치": "",
                    "확인할 점": "첫 시트를 확인하세요.",
                }
            )
        return pd.DataFrame(rows, columns=columns)

    def _before_row_detail(self, result: WorkbookDiffResult, before_row_number: object) -> str:
        try:
            before_index = int(before_row_number) - 2
        except (TypeError, ValueError):
            return ""
        if not 0 <= before_index < len(result.before_frame):
            return ""
        row = result.before_frame.loc[before_index]
        parts = []
        for column, value in row.items():
            text = display_value(value)
            if text:
                parts.append(f"{column}={text}")
            if len(parts) >= 12:
                break
        return "; ".join(parts)

    def _guide_frame(self, result: WorkbookDiffResult) -> pd.DataFrame:
        return pd.DataFrame(
            [
                {
                    "먼저 볼 내용": "변경 표시된 후 파일",
                    "현재 결과": f"{result.summary.get('changed_cell_count', 0)}개 셀 변경",
                    "바로 할 일": "첫 번째 `후파일_메모` 시트에서 색칠된 셀과 메모를 확인하세요.",
                    "관련 시트": "후파일_메모",
                },
                {
                    "먼저 볼 내용": "비교 기준",
                    "현재 결과": result.summary.get("compare_basis", ""),
                    "바로 할 일": "기준열 자동 추천이 맞는지 확인하세요.",
                    "관련 시트": "점검요약",
                },
                {
                    "먼저 볼 내용": "행 추가/누락",
                    "현재 결과": f"{result.summary.get('missing_row_count', 0) + result.summary.get('added_row_count', 0)}건",
                    "바로 할 일": "후 파일에 새로 생기거나 빠진 행을 확인하세요. 빠진 행의 원본 내용은 `빠진행` 시트에 있습니다.",
                    "관련 시트": "행비교, 빠진행",
                },
                {
                    "먼저 볼 내용": "컬럼 추가/누락",
                    "현재 결과": f"{result.summary.get('before_only_column_count', 0) + result.summary.get('after_only_column_count', 0)}건",
                    "바로 할 일": "전/후 파일의 컬럼 구성이 달라진 부분을 확인하세요.",
                    "관련 시트": "컬럼비교",
                },
            ]
        )

    def _frame_or_empty(self, frame: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
        if frame.empty:
            return pd.DataFrame(columns=columns)
        return frame

    def _changed_rows_frame(self, result: WorkbookDiffResult) -> pd.DataFrame:
        if result.after_frame.empty:
            return pd.DataFrame()

        row_status: dict[int, set[str]] = {}
        for comment in result.comments:
            row_status.setdefault(comment.row_index, set()).add("셀 변경/추가")
        if not result.diff_frame.empty:
            for _, diff_row in result.diff_frame.iterrows():
                cell_ref = str(diff_row.get("후 파일 셀", ""))
                row_number = "".join(ch for ch in cell_ref if ch.isdigit())
                if row_number:
                    row_status.setdefault(int(row_number) - 2, set()).add(str(diff_row.get("상태", "")))
        if not row_status:
            return pd.DataFrame()

        rows = []
        for index in sorted(row_status):
            if 0 <= index < len(result.after_frame):
                row = {"차이 유형": ", ".join(sorted(row_status[index])), "후 파일 행": index + 2}
                row.update(result.after_frame.loc[index].to_dict())
                rows.append(row)
        return pd.DataFrame(rows)

    def _missing_rows_frame(self, result: WorkbookDiffResult) -> pd.DataFrame:
        if result.before_frame.empty or result.row_frame.empty:
            return pd.DataFrame()
        missing = result.row_frame[result.row_frame["상태"] == "후 파일에 행 없음"]
        if missing.empty:
            return pd.DataFrame()

        rows = []
        for _, missing_row in missing.iterrows():
            before_row_number = missing_row.get("전 파일 행", "")
            try:
                before_index = int(before_row_number) - 2
            except (TypeError, ValueError):
                continue
            if 0 <= before_index < len(result.before_frame):
                row = {
                    "차이 유형": "후 파일에 행 없음",
                    "비교 기준": missing_row.get("비교 기준", ""),
                    "전 파일 행": before_row_number,
                }
                row.update(result.before_frame.loc[before_index].to_dict())
                rows.append(row)
        return pd.DataFrame(rows)

    def _apply_comments(self, workbook, result: WorkbookDiffResult) -> None:
        if "후파일_메모" not in workbook.sheetnames:
            return
        sheet = workbook["후파일_메모"]
        column_positions = {cell.value: cell.column for cell in sheet[1]}
        for comment in result.comments:
            column = column_positions.get(comment.column_name)
            if not column:
                continue
            cell = sheet.cell(row=comment.row_index + 2, column=column)
            cell.comment = make_comment(comment.message)
            cell.fill = PatternFill("solid", fgColor=comment.fill)

    def _apply_row_highlights(self, workbook, result: WorkbookDiffResult) -> None:
        if "후파일_메모" not in workbook.sheetnames:
            return
        sheet = workbook["후파일_메모"]
        data_column_count = sheet.max_column

        for _, row_diff in result.row_frame.iterrows():
            status = str(row_diff.get("상태", ""))
            if status != "전 파일에 없던 행":
                continue
            try:
                row_number = int(row_diff.get("후 파일 행", ""))
            except (TypeError, ValueError):
                continue
            if row_number >= 2:
                self._mark_row(
                    sheet,
                    row_number,
                    data_column_count,
                    self.ADDED_ROW_FILL,
                    "전 파일에 없던 행입니다. 후 파일에서 새로 추가된 행인지 확인하세요.",
                )

        appended_row_number = len(result.after_frame) + 2
        for _, row_diff in result.row_frame.iterrows():
            status = str(row_diff.get("상태", ""))
            if status != "후 파일에 행 없음":
                continue
            self._mark_row(
                sheet,
                appended_row_number,
                data_column_count,
                self.MISSING_ROW_FILL,
                "후 파일에서 사라진 행입니다. 전 파일에는 있었지만 후 파일에는 없습니다.",
            )
            appended_row_number += 1

    def _mark_row(self, sheet, row_number: int, column_count: int, fill: str, message: str) -> None:
        row_fill = PatternFill("solid", fgColor=fill)
        for column_number in range(1, column_count + 1):
            sheet.cell(row=row_number, column=column_number).fill = row_fill
        first_cell = sheet.cell(row=row_number, column=1)
        if first_cell.comment is None:
            first_cell.comment = make_comment(message)

    def _style_workbook(self, workbook) -> None:
        header_fill = PatternFill("solid", fgColor="1F6F5F")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            if sheet.max_row > 1 and sheet.max_column > 0:
                sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, min(len(value), 48))
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                sheet.column_dimensions[column_letter].width = max(10, min(max_length + 2, 50))
