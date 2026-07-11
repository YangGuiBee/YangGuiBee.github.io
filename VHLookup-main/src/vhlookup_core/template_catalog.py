from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from vhlookup_core.templates import WorkflowTemplate, all_templates


def template_catalog_frame(templates: tuple[WorkflowTemplate, ...] | None = None) -> pd.DataFrame:
    rows = []
    for template in templates or all_templates():
        rows.append(
            {
                "템플릿 ID": template.id,
                "업무명": template.name,
                "모드": template.mode,
                "한 줄 설명": template.headline,
                "상세 설명": template.description,
                "표준 컬럼": ", ".join(template.standard_columns),
                "기준 키": ", ".join(template.key_columns),
                "대상 키": ", ".join(template.target_key_columns),
                "가져올 컬럼": ", ".join(template.value_columns),
                "필수 컬럼": ", ".join(template.required_columns),
                "숫자 컬럼": ", ".join(template.numeric_columns),
                "날짜 컬럼": ", ".join(template.date_columns),
                "중복 검사 키": ", ".join(template.duplicate_key_columns),
                "기본 결과 파일명": template.output_name,
            }
        )
    return pd.DataFrame(rows)


def template_usage_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "상황": "제출자료 폴더 수합",
                "추천 명령": "python -m vhlookup_cli.main consolidate --folder submissions --template school_submission_consolidation --out result.xlsx",
            },
            {
                "상황": "수합 전 사전점검",
                "추천 명령": "python -m vhlookup_cli.main inspect --path submissions --template school_submission_consolidation --out inspect.xlsx",
            },
            {
                "상황": "기준표 값 붙이기",
                "추천 명령": "python -m vhlookup_cli.main lookup --reference master.xlsx --target submitted.xlsx --template hr_training_lookup --out lookup.xlsx",
            },
            {
                "상황": "누락자 확인",
                "추천 명령": "python -m vhlookup_cli.main reconcile --reference expected.xlsx --target received.xlsx --template missing_people_reconciliation --out missing.xlsx",
            },
            {
                "상황": "월별 가로표 변환",
                "추천 명령": "python -m vhlookup_cli.main horizontal --file monthly.xlsx --out monthly_long.xlsx",
            },
        ]
    )


class TemplateCatalogWriter:
    def write_xlsx(self, path: str | Path) -> Path:
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            template_catalog_frame().to_excel(writer, sheet_name="업무템플릿", index=False)
            template_usage_frame().to_excel(writer, sheet_name="명령예시", index=False)
            self._style_workbook(writer.book)
        return output_path

    def _style_workbook(self, workbook) -> None:
        header_fill = PatternFill("solid", fgColor="1F6F5F")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            if sheet.max_row > 1 and sheet.max_column > 0:
                sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, min(len(value), 56))
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                sheet.column_dimensions[column_letter].width = max(10, min(max_length + 2, 58))
