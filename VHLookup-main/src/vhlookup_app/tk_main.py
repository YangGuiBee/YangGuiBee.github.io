from __future__ import annotations

import os
import sys
import traceback
from numbers import Integral, Real
from datetime import datetime
from pathlib import Path
from tkinter import Canvas, END, Listbox, StringVar, Text, Tk, Toplevel, filedialog, messagebox
from tkinter import ttk

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from vhlookup_core import (
    AutoLookupPlanner,
    AdminWorkbookTools,
    APP_DISPLAY_NAME,
    ConsolidationEngine,
    ExcelLoader,
    HeaderDetector,
    PrivacyMaskingEngine,
    ReportWriter,
    SheetDetector,
    WorkbookDiffEngine,
    WorkbookDiffReportWriter,
    get_template,
    template_catalog_frame,
    template_usage_frame,
)
from vhlookup_core.horizontal import HorizontalTableEngine
from vhlookup_core.merge import MergeEngine
from vhlookup_core.models import JobResult
from vhlookup_core.reconciliation import ReconciliationEngine


APP_TITLE = APP_DISPLAY_NAME


def app_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parents[2]


def resource_path(relative_path: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", app_base_dir()))
    return base / relative_path


def load_table(path: Path):
    loader = ExcelLoader()
    header_detector = HeaderDetector()
    sheet_detector = SheetDetector(header_detector)
    source = loader.load(path)
    sheet = sheet_detector.select(source)
    detection = header_detector.detect(sheet)
    return header_detector.apply(sheet, detection)


def attach_auto_summary(result: JobResult, plan, workflow: str) -> JobResult:
    result.mapping_records.extend(plan.evidence_rows)
    result.summary["workflow"] = workflow
    result.summary["auto_key_columns"] = " + ".join(plan.key_spec.reference_key_columns)
    result.summary["auto_target_key_columns"] = " + ".join(plan.key_spec.target_columns())
    if plan.value_columns:
        result.summary["auto_value_columns"] = ", ".join(plan.value_columns)
    return result


def style_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F6F5F")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row > 1 and sheet.max_column > 0:
            sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value), 48))
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            sheet.column_dimensions[column_letter].width = max(10, min(max_length + 2, 50))


DEMO_TOUR_ROWS = [
    {
        "순서": "01",
        "결과 파일": "01_제출자료_수합결과.xlsx",
        "업무 상황": "학교/부서/기관에서 받은 제출자료 수합",
        "열어볼 시트": "결과, 자동추천근거, 개인정보점검",
        "확인 포인트": "컬럼명을 자동으로 맞춘 통합본, 담당자/연락처 점검",
    },
    {
        "순서": "02",
        "결과 파일": "02_오류검증_제출자료.xlsx",
        "업무 상황": "제출자료 오류 검토",
        "열어볼 시트": "먼저확인, 확인필요, 개인정보점검",
        "확인 포인트": "필수값 누락, 숫자 오류, 날짜 오류, 중복 제출 의심",
    },
    {
        "순서": "03",
        "결과 파일": "03_교육이수_명단대조.xlsx",
        "업무 상황": "교육이수 명단에 직원 기본정보 붙이기",
        "열어볼 시트": "결과, 자동추천근거, 확인필요, 개인정보점검",
        "확인 포인트": "사번과 직원번호 자동 연결, 부서/직급/소속 붙이기",
    },
    {
        "순서": "04",
        "결과 파일": "04_교육누락자_대조결과.xlsx",
        "업무 상황": "교육 대상자 중 빠진 사람 찾기",
        "열어볼 시트": "결과, 확인필요, 개인정보점검",
        "확인 포인트": "기준명단에만 있음, 교육이수 명단에만 있음, 양쪽 모두 있음",
    },
    {
        "순서": "05",
        "결과 파일": "05_수당예산_대조결과.xlsx",
        "업무 상황": "수당/예산 지급대상자 기준표 대조",
        "열어볼 시트": "결과, 자동추천근거, 확인필요, 개인정보점검",
        "확인 포인트": "사번+지급월 복합 키로 단가/지급기준/예산과목 붙이기",
    },
    {
        "순서": "06",
        "결과 파일": "06_제출대상_누락확인.xlsx",
        "업무 상황": "제출 대상 기관 누락 확인",
        "열어볼 시트": "결과, 확인필요",
        "확인 포인트": "미제출 기관과 대상이 아닌 제출 기관 분리",
    },
    {
        "순서": "07",
        "결과 파일": "07_부서별현황_수합결과.xlsx",
        "업무 상황": "부서별 현황자료 수합",
        "열어볼 시트": "결과, 자동추천근거, 개인정보점검",
        "확인 포인트": "담당 부서/소속, 현원/인원수 같은 다른 표현 통합",
    },
    {
        "순서": "08",
        "결과 파일": "08_월별가로표_세로변환.xlsx",
        "업무 상황": "월별 가로표 세로 변환",
        "열어볼 시트": "결과",
        "확인 포인트": "1월, 2월, 3월 컬럼을 세로형 자료로 변환",
    },
]


DEMO_SECURITY_ROWS = [
    {"항목": "AI 사용", "내용": "사용 안 함"},
    {"항목": "외부 API", "내용": "호출 안 함"},
    {"항목": "클라우드 업로드", "내용": "없음"},
    {"항목": "원본 파일", "내용": "수정하지 않음"},
    {"항목": "결과 저장", "내용": "사용자 PC의 로컬 폴더에 새 xlsx 파일로 저장"},
    {"항목": "검토 방식", "내용": "먼저확인, 확인필요, 자동추천근거, 개인정보점검 시트로 사람이 확인 가능"},
]


def write_demo_tour(output_dir: Path) -> None:
    path = output_dir / "00_샘플_둘러보기.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(DEMO_TOUR_ROWS).to_excel(writer, sheet_name="샘플순서", index=False)
        pd.DataFrame(DEMO_SECURITY_ROWS).to_excel(writer, sheet_name="보안원칙", index=False)
        template_catalog_frame().to_excel(writer, sheet_name="업무템플릿", index=False)
        template_usage_frame().to_excel(writer, sheet_name="명령예시", index=False)
        pd.DataFrame(
            [
                {"단계": 1, "할 일": "00_샘플_둘러보기.xlsx에서 샘플순서 시트를 봅니다."},
                {"단계": 2, "할 일": "02_오류검증_제출자료.xlsx의 확인필요 시트를 봅니다."},
                {"단계": 3, "할 일": "03_교육이수_명단대조.xlsx의 자동추천근거 시트를 봅니다."},
                {"단계": 4, "할 일": "05_수당예산_대조결과.xlsx에서 복합 키 대조 결과를 봅니다."},
                {"단계": 5, "할 일": "01_제출자료_수합결과.xlsx의 개인정보점검 시트를 봅니다."},
            ]
        ).to_excel(writer, sheet_name="추천동선", index=False)
        style_workbook(writer.book)


def generate_demo_reports(output_dir: Path, samples_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    for old_output in output_dir.glob("*.xlsx"):
        old_output.unlink()

    writer = ReportWriter()
    write_demo_tour(output_dir)
    merge_samples = samples_dir / "03_merge_files"
    hr_samples = merge_samples / "column_merge_hr_training"
    extra_samples = samples_dir / "90_extra_cli_samples"

    consolidated = ConsolidationEngine().consolidate_folder(
        merge_samples / "row_merge_school_submissions",
        template="school_submission_consolidation",
    )
    writer.write_xlsx(
        consolidated,
        output_dir / "01_제출자료_수합결과.xlsx",
        mark_result_cells=False,
        include_privacy_scan=False,
    )

    invalid_consolidated = ConsolidationEngine().consolidate_folder(
        merge_samples / "row_merge_submission_errors",
        template="school_submission_consolidation",
    )
    writer.write_xlsx(
        invalid_consolidated,
        output_dir / "02_오류검증_제출자료.xlsx",
        mark_result_cells=False,
        include_privacy_scan=False,
    )

    employee_master = load_table(hr_samples / "hr_employee_master.csv")
    training = load_table(hr_samples / "hr_training_completion.csv")
    lookup = ConsolidationEngine().merge_files_by_columns(
        [hr_samples / "hr_employee_master.csv", hr_samples / "hr_training_completion.csv"]
    )
    writer.write_xlsx(lookup, output_dir / "03_교육이수_명단대조.xlsx", include_privacy_scan=False)

    reconciliation_plan = AutoLookupPlanner().infer_reconciliation_key_spec(employee_master, training)
    reconciliation = ReconciliationEngine().compare_lists(
        employee_master,
        training,
        reconciliation_plan.key_spec,
    )
    attach_auto_summary(reconciliation, reconciliation_plan, "빠진 사람/누락 제출자 찾기")
    writer.write_xlsx(reconciliation, output_dir / "04_교육누락자_대조결과.xlsx")

    allowance_template = get_template("allowance_budget_lookup")
    allowance_samples = merge_samples / "column_merge_allowance_budget"
    rate_reference = load_table(allowance_samples / "rate_reference.csv")
    payment_requests = load_table(allowance_samples / "payment_requests.csv")
    allowance_plan = AutoLookupPlanner().infer_lookup_plan(
        rate_reference,
        payment_requests,
        preferred_reference_key_columns=allowance_template.key_columns,
        preferred_target_key_columns=("직원번호", "지급월"),
        preferred_value_columns=allowance_template.value_columns,
    )
    allowance = MergeEngine().merge_lookup(
        rate_reference,
        payment_requests,
        allowance_plan.key_spec,
        value_columns=list(allowance_plan.value_columns),
    )
    attach_auto_summary(allowance, allowance_plan, allowance_template.name)
    writer.write_xlsx(allowance, output_dir / "05_수당예산_대조결과.xlsx")

    submitter_samples = extra_samples / "submission_reconciliation"
    expected = load_table(submitter_samples / "expected_submitters.csv")
    received = load_table(submitter_samples / "received_submitters.csv")
    submitter_plan = AutoLookupPlanner().infer_reconciliation_key_spec(
        expected,
        received,
        preferred_reference_key_columns=("기관코드",),
        preferred_target_key_columns=("제출기관코드",),
    )
    submitter_reconciliation = ReconciliationEngine().compare_lists(
        expected,
        received,
        submitter_plan.key_spec,
        reference_label="제출 대상 명단",
        target_label="실제 제출 명단",
    )
    attach_auto_summary(submitter_reconciliation, submitter_plan, "제출 대상자 누락 확인")
    writer.write_xlsx(submitter_reconciliation, output_dir / "06_제출대상_누락확인.xlsx")

    department_status = ConsolidationEngine().consolidate_folder(
        merge_samples / "row_merge_messy_headers",
        template="department_status_consolidation",
    )
    writer.write_xlsx(
        department_status,
        output_dir / "07_부서별현황_수합결과.xlsx",
        mark_result_cells=False,
        include_privacy_scan=False,
    )

    monthly = load_table(extra_samples / "horizontal_table" / "monthly_budget_wide.csv")
    converted = HorizontalTableEngine().wide_to_long(monthly, id_columns=["기관명", "항목"])
    horizontal = JobResult(
        result_frame=converted,
        summary={"workflow": "월별 가로표 세로 변환", "row_count": len(converted)},
    )
    writer.write_xlsx(horizontal, output_dir / "08_월별가로표_세로변환.xlsx")
    return sorted(output_dir.glob("*.xlsx"))


def split_columns(text: str) -> list[str]:
    return [part.strip() for part in text.split(",") if part.strip()]


class LocalApp:
    def __init__(self, root: Tk) -> None:
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1020x760")
        self.root.minsize(900, 680)
        self.root.configure(bg="#F4F7FA")
        self.base_dir = app_base_dir()
        self.output_dir = self.base_dir / "outputs"
        self.output_dir.mkdir(exist_ok=True)
        self.last_output_dir = self.output_dir

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("App.TFrame", background="#F4F7FA")
        style.configure("Panel.TFrame", background="#FFFFFF", relief="flat")
        style.configure("Header.TFrame", background="#1F6F5F")
        style.configure("HeaderTitle.TLabel", background="#1F6F5F", foreground="#FFFFFF", font=("", 20, "bold"), padding=4)
        style.configure("HeaderSub.TLabel", background="#1F6F5F", foreground="#D7F3EE", font=("", 10), padding=4)
        style.configure("Section.TLabel", background="#F4F7FA", foreground="#17324D", font=("", 12, "bold"))
        style.configure("Action.TFrame", background="#FFFFFF")
        style.configure("ActionTitle.TLabel", background="#FFFFFF", foreground="#17324D", font=("", 10, "bold"))
        style.configure("ActionDesc.TLabel", background="#FFFFFF", foreground="#526173", font=("", 9))
        style.configure("Status.TLabel", background="#F4F7FA", foreground="#17324D", font=("", 9, "bold"))
        style.configure("TButton", padding=(10, 7), font=("", 9))
        style.configure("Action.TButton", padding=(12, 8), font=("", 9, "bold"))
        style.map("Action.TButton", background=[("active", "#DCEFEA")])

        self.status = StringVar(value="준비됨")
        self._build()

    def _build(self) -> None:
        top = ttk.Frame(self.root, padding=14, style="App.TFrame")
        top.pack(fill="both", expand=True)

        header = ttk.Frame(top, padding=(18, 16), style="Header.TFrame")
        header.pack(fill="x")
        ttk.Label(header, text=APP_DISPLAY_NAME, style="HeaderTitle.TLabel").pack(anchor="w")
        ttk.Label(
            header,
            text="엑셀 수합, 대조, 검증, 피벗 요약을 로컬 PC에서 처리합니다. 원본 파일은 수정하지 않습니다.",
            style="HeaderSub.TLabel",
        ).pack(anchor="w", pady=(4, 0))

        ttk.Label(top, text="실행할 작업", style="Section.TLabel").pack(anchor="w", pady=(14, 6))

        actions = ttk.Frame(top, padding=10, style="Panel.TFrame")
        actions.pack(fill="x")
        self._action_button(
            actions,
            "1. 개인정보 마스킹",
            "이름, 고유식별번호, 연락처, 이메일, 계좌번호, 주소, 성별, 나이, 생년월일을 가린 새 엑셀을 만듭니다.",
            self.quick_privacy_mask,
        )
        self._action_button(
            actions,
            "2. 분류별 시트 나누기",
            "부서, 기관명, 상태 같은 열을 기준으로 분류별 시트를 앞쪽에 만듭니다.",
            self.quick_split_sheets,
        )
        self._action_button(
            actions,
            "3. 엑셀/CSV 파일 여러 개 합치기",
            "파일 구조를 보고 행/열 합치기를 자동 추천하고, 합쳐진 결과를 첫 시트로 만듭니다.",
            self.quick_consolidate,
        )
        self._action_button(
            actions,
            "4. 전/후 파일 검증",
            "수정 전/후 파일을 비교하고, 달라진 셀에 색과 메모를 남깁니다.",
            self.quick_diff_workbooks,
        )
        self._action_button(
            actions,
            "5. 피벗 요약표 만들기",
            "부서/기관/월별 건수, 합계, 평균 요약표를 드롭박스로 선택해 만듭니다.",
            self.quick_pivot_summary,
        )

        ttk.Label(top, text="실행 상태", style="Section.TLabel").pack(anchor="w", pady=(14, 6))
        log_frame = ttk.Frame(top, padding=10, style="Panel.TFrame")
        log_frame.pack(fill="both", expand=True)
        self.log = Text(
            log_frame,
            height=8,
            wrap="word",
            relief="flat",
            bg="#FBFCFE",
            fg="#17324D",
            insertbackground="#17324D",
            padx=10,
            pady=8,
        )
        self.log.pack(fill="both", expand=True)
        bottom = ttk.Frame(top, style="App.TFrame")
        bottom.pack(fill="x", pady=(8, 0))
        ttk.Label(bottom, textvariable=self.status, style="Status.TLabel").pack(side="left")
        ttk.Button(bottom, text="최근 저장 폴더 열기", command=lambda: self.open_folder(self.last_output_dir)).pack(side="right")

    def _action_button(self, parent, title: str, description: str, command) -> None:
        row = ttk.Frame(parent, padding=(8, 7), style="Action.TFrame")
        row.pack(fill="x", pady=3)
        button = ttk.Button(row, text=title, command=command, width=30, style="Action.TButton")
        button.pack(side="left", anchor="n")
        text_box = ttk.Frame(row, style="Action.TFrame")
        text_box.pack(side="left", fill="x", expand=True, padx=(12, 0))
        ttk.Label(text_box, text=title.split(". ", 1)[-1], style="ActionTitle.TLabel").pack(anchor="w")
        ttk.Label(text_box, text=description, wraplength=650, justify="left", style="ActionDesc.TLabel").pack(anchor="w", pady=(2, 0))

    def _timestamped_output(self, base_name: str) -> Path:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return self.output_dir / f"{base_name}_{stamp}.xlsx"

    def _ask_save_path_or_none(self, title: str, base_name: str) -> Path | None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        selected = filedialog.asksaveasfilename(
            title=title,
            initialdir=self.output_dir,
            initialfile=f"{base_name}_{stamp}.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not selected:
            return None
        output = Path(selected)
        if output.suffix.lower() != ".xlsx":
            output = output.with_suffix(".xlsx")
        return output

    def _ask_folder_or_none(self, title: str) -> Path | None:
        selected = filedialog.askdirectory(title=title, initialdir=self.base_dir)
        return Path(selected) if selected else None

    def _ask_file_or_none(self, title: str) -> Path | None:
        selected = filedialog.askopenfilename(
            title=title,
            initialdir=self.base_dir,
            filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv"), ("All files", "*.*")],
        )
        return Path(selected) if selected else None

    def _ask_files_or_none(self, title: str) -> list[Path]:
        selected = filedialog.askopenfilenames(
            title=title,
            initialdir=self.base_dir,
            filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv"), ("All files", "*.*")],
        )
        return [Path(path) for path in selected]

    def quick_privacy_mask(self) -> None:
        file_path = self._ask_file_or_none("개인정보를 마스킹할 엑셀/CSV 파일을 선택하세요")
        if not file_path:
            return
        output = self._ask_save_path_or_none("마스킹 결과를 저장할 위치를 선택하세요", "개인정보_마스킹결과")
        if not output:
            return

        def job():
            PrivacyMaskingEngine().write_xlsx(file_path, output)
            return [output]

        self._run("개인정보 마스킹", job, open_path=output.parent)

    def quick_split_sheets(self) -> None:
        file_path = self._ask_file_or_none("분류별로 나눌 엑셀/CSV 파일을 선택하세요")
        if not file_path:
            return

        tools = AdminWorkbookTools()
        try:
            table, _metadata = tools.load_table(file_path)
            inferred_column = tools.infer_split_column(table)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, str(exc))
            return

        selected_column = self._choose_column_dialog(
            title="분류 기준 열 선택",
            message="어떤 열 기준으로 시트를 나눌까요?",
            columns=[str(column) for column in table.columns],
            initial=inferred_column,
        )
        if not selected_column:
            return
        output = self._ask_save_path_or_none("분류별 시트 결과를 저장할 위치를 선택하세요", "분류별_시트나누기")
        if not output:
            return

        def job():
            split_result = tools.write_split_workbook(file_path, output, selected_column)
            self.log_insert(f"분류 기준 열: {split_result.split_column}")
            self.log_insert(f"생성 시트 수: {split_result.sheet_count}")
            return [output]

        self._run("분류별 시트 나누기", job, open_path=output.parent)

    def _choose_column_dialog(
        self,
        title: str,
        message: str,
        columns: list[str],
        initial: str | None = None,
    ) -> str | None:
        dialog = Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("460x180")
        dialog.transient(self.root)
        dialog.grab_set()

        selected = StringVar(value=initial if initial in columns else (columns[0] if columns else ""))
        result: dict[str, str | None] = {"value": None}

        frame = ttk.Frame(dialog, padding=14)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text=message, font=("", 11, "bold")).pack(anchor="w", pady=(0, 8))
        combo = ttk.Combobox(frame, textvariable=selected, values=columns, state="readonly")
        combo.pack(fill="x")
        if selected.get():
            combo.set(selected.get())

        button_row = ttk.Frame(frame)
        button_row.pack(fill="x", pady=(16, 0))

        def confirm() -> None:
            result["value"] = selected.get()
            dialog.destroy()

        def cancel() -> None:
            result["value"] = None
            dialog.destroy()

        ttk.Button(button_row, text="취소", command=cancel).pack(side="right")
        ttk.Button(button_row, text="확인", command=confirm).pack(side="right", padx=(0, 8))
        dialog.protocol("WM_DELETE_WINDOW", cancel)
        combo.focus_set()
        self.root.wait_window(dialog)
        return result["value"]

    def _write_preview_tree(self, tree: ttk.Treeview, frame: pd.DataFrame, limit: int = 10) -> None:
        for item in tree.get_children():
            tree.delete(item)
        preview = frame.head(limit).copy()
        columns = [str(column) for column in preview.columns]
        tree.configure(columns=columns)
        tree["show"] = "headings"
        for column in columns:
            tree.heading(column, text=column)
            tree.column(column, width=120, minwidth=80, stretch=True)
        for _, row in preview.iterrows():
            tree.insert(END, values=[self._format_preview_value(value) for value in row.tolist()])

    def _write_preview_text(self, widget: Text, frame: pd.DataFrame, limit: int = 10) -> None:
        preview = frame.head(limit).copy()
        if preview.empty:
            body = "미리볼 데이터가 없습니다."
        else:
            display_preview = preview.apply(lambda column: column.map(self._format_preview_value))
            body = display_preview.to_string(index=False, max_rows=limit, max_cols=20)
        header = f"미리보기: 총 {len(frame)}행, {len(frame.columns)}열 중 앞 {min(len(frame), limit)}행 표시\n"
        widget.configure(state="normal")
        widget.delete("1.0", END)
        widget.insert(END, header)
        widget.insert(END, "-" * 80 + "\n")
        widget.insert(END, body)
        widget.configure(state="disabled")
        widget.see("1.0")

    def _format_preview_value(self, value: object) -> str:
        if pd.isna(value):
            return ""
        if isinstance(value, Real) and not isinstance(value, (bool, Integral)):
            rounded = round(float(value), 2)
            if rounded.is_integer():
                return f"{int(rounded):,}"
            return f"{rounded:,.2f}".rstrip("0").rstrip(".")
        if isinstance(value, Integral) and not isinstance(value, bool):
            return f"{int(value):,}"
        return str(value)

    def _scrollable_frame(self, parent) -> tuple[ttk.Frame, Canvas]:
        canvas = Canvas(parent, highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)
        window_id = canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def configure_inner(_event=None) -> None:
            canvas.configure(scrollregion=canvas.bbox("all"))

        def configure_canvas(event) -> None:
            canvas.itemconfigure(window_id, width=event.width)

        inner.bind("<Configure>", configure_inner)
        canvas.bind("<Configure>", configure_canvas)
        return inner, canvas

    def _confirm_row_merge_preview(self, files: list[Path]) -> dict[str, dict[str, str]] | None:
        try:
            tables = [(file, load_table(file)) for file in files]
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"미리보기를 만들 수 없습니다.\n{exc}")
            return None

        base_columns = [str(column) for column in tables[0][1].columns]
        engine = ConsolidationEngine()
        variables: dict[tuple[str, str], StringVar] = {}
        result: dict[str, dict[str, str]] | None = None

        dialog = Toplevel(self.root)
        dialog.title("행 합치기 미리보기 / 컬럼 매칭")
        dialog.geometry("980x680")
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding=12)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="예상 결과를 확인하고, 컬럼 매칭이 틀리면 드롭박스로 수정하세요.", font=("", 11, "bold")).pack(
            anchor="w", pady=(0, 8)
        )

        preview_tree = ttk.Treeview(frame, height=8)
        preview_tree.pack(fill="both", expand=True)

        notebook = ttk.Notebook(frame)
        notebook.pack(fill="both", expand=True, pady=(10, 0))

        for file, table in tables:
            tab = ttk.Frame(notebook)
            notebook.add(tab, text=file.name[:24])
            inner, _canvas = self._scrollable_frame(tab)
            source_columns = [str(column) for column in table.columns]
            auto_mapping = engine.mapper.map_columns(source_columns, base_columns).target_to_source
            for row_index, target_column in enumerate(base_columns):
                ttk.Label(inner, text=target_column, width=24).grid(row=row_index, column=0, sticky="w", padx=(4, 8), pady=3)
                selected_source = auto_mapping.get(target_column, target_column if target_column in source_columns else "")
                variable = StringVar(value=selected_source)
                variables[(str(file.resolve()), target_column)] = variable
                combo = ttk.Combobox(inner, textvariable=variable, values=["", *source_columns], state="readonly", width=36)
                combo.grid(row=row_index, column=1, sticky="ew", padx=(0, 4), pady=3)
            inner.columnconfigure(1, weight=1)

        def current_mappings() -> dict[str, dict[str, str]]:
            mappings: dict[str, dict[str, str]] = {}
            for file, _table in tables:
                file_key = str(file.resolve())
                mappings[file_key] = {
                    target_column: variable.get()
                    for (path_key, target_column), variable in variables.items()
                    if path_key == file_key and variable.get()
                }
            return mappings

        def preview_frame() -> pd.DataFrame:
            mappings = current_mappings()
            rows = []
            for file, table in tables:
                file_mapping = mappings.get(str(file.resolve()), {})
                for row_number, (_, source_row) in enumerate(table.head(5).iterrows(), start=1):
                    row = {"원본 파일명": file.name, "미리보기 행": row_number}
                    for target_column in base_columns:
                        source_column = file_mapping.get(target_column)
                        row[target_column] = source_row[source_column] if source_column in table.columns else pd.NA
                    rows.append(row)
            return pd.DataFrame(rows).head(10)

        def refresh_preview() -> None:
            self._write_preview_tree(preview_tree, preview_frame(), limit=10)

        def cancel() -> None:
            nonlocal result
            result = None
            dialog.destroy()

        def confirm() -> None:
            nonlocal result
            result = current_mappings()
            dialog.destroy()

        button_row = ttk.Frame(frame)
        button_row.pack(fill="x", pady=(10, 0))
        ttk.Button(button_row, text="미리보기 새로고침", command=refresh_preview).pack(side="left")
        ttk.Button(button_row, text="취소", command=cancel).pack(side="right")
        ttk.Button(button_row, text="이 설정으로 실행", command=confirm).pack(side="right", padx=(0, 8))
        dialog.protocol("WM_DELETE_WINDOW", cancel)
        refresh_preview()
        self.root.wait_window(dialog)
        return result

    def _infer_column_merge_settings(self, files: list[Path]) -> tuple[list[tuple[Path, pd.DataFrame]], dict[str, dict[str, object]]]:
        tables = [(file, load_table(file)) for file in files]
        if len(tables) < 2:
            return tables, {}
        base_frame = tables[0][1]
        planner = AutoLookupPlanner()
        settings: dict[str, dict[str, object]] = {}
        for file, table in tables[1:]:
            try:
                plan = planner.infer_lookup_plan(table, base_frame)
                settings[str(file.resolve())] = {
                    "base_key_columns": plan.key_spec.target_columns(),
                    "source_key_columns": plan.key_spec.reference_key_columns,
                    "value_columns": plan.value_columns,
                }
            except Exception:
                source_columns = tuple(str(column) for column in table.columns)
                base_columns = tuple(str(column) for column in base_frame.columns)
                settings[str(file.resolve())] = {
                    "base_key_columns": base_columns[:1],
                    "source_key_columns": source_columns[:1],
                    "value_columns": source_columns[1:] or source_columns[:1],
                }
        return tables, settings

    def _confirm_column_merge_preview(self, files: list[Path]) -> dict[str, dict[str, object]] | None:
        try:
            tables, auto_settings = self._infer_column_merge_settings(files)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"미리보기를 만들 수 없습니다.\n{exc}")
            return None
        if len(tables) < 2:
            messagebox.showwarning(APP_TITLE, "열 합치기는 최소 2개 파일이 필요합니다.")
            return None

        base_columns = [str(column) for column in tables[0][1].columns]
        controls: dict[str, dict[str, object]] = {}
        result: dict[str, dict[str, object]] | None = None

        dialog = Toplevel(self.root)
        dialog.title("열 합치기 미리보기 / 키 컬럼 확인")
        dialog.geometry("980x700")
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding=12)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="첫 번째 파일을 기준으로, 나머지 파일의 키 컬럼과 붙일 컬럼을 확인하세요.", font=("", 11, "bold")).pack(
            anchor="w", pady=(0, 8)
        )

        preview_tree = ttk.Treeview(frame, height=8)
        preview_tree.pack(fill="both", expand=True)

        notebook = ttk.Notebook(frame)
        notebook.pack(fill="both", expand=True, pady=(10, 0))

        for file, table in tables[1:]:
            source_columns = [str(column) for column in table.columns]
            setting = auto_settings.get(str(file.resolve()), {})
            base_key = next(iter(setting.get("base_key_columns", base_columns[:1])), base_columns[0] if base_columns else "")
            source_key = next(iter(setting.get("source_key_columns", source_columns[:1])), source_columns[0] if source_columns else "")
            value_columns = set(setting.get("value_columns", source_columns[1:]))

            tab = ttk.Frame(notebook, padding=10)
            notebook.add(tab, text=file.name[:24])
            ttk.Label(tab, text="기준 파일 키 컬럼").grid(row=0, column=0, sticky="w", pady=4)
            base_var = StringVar(value=base_key)
            ttk.Combobox(tab, textvariable=base_var, values=base_columns, state="readonly").grid(row=0, column=1, sticky="ew", pady=4)
            ttk.Label(tab, text="이 파일 키 컬럼").grid(row=1, column=0, sticky="w", pady=4)
            source_var = StringVar(value=source_key)
            ttk.Combobox(tab, textvariable=source_var, values=source_columns, state="readonly").grid(row=1, column=1, sticky="ew", pady=4)
            ttk.Label(tab, text="오른쪽에 붙일 컬럼").grid(row=2, column=0, sticky="nw", pady=(10, 4))
            value_list = Listbox(tab, selectmode="multiple", height=10)
            value_list.grid(row=2, column=1, sticky="nsew", pady=(10, 4))
            for source_column in source_columns:
                value_list.insert(END, source_column)
                if source_column in value_columns:
                    value_list.selection_set(END)
            tab.columnconfigure(1, weight=1)
            tab.rowconfigure(2, weight=1)
            controls[str(file.resolve())] = {
                "base": base_var,
                "source": source_var,
                "values": value_list,
                "columns": source_columns,
            }

        def current_settings() -> dict[str, dict[str, object]]:
            settings: dict[str, dict[str, object]] = {}
            for file, _table in tables[1:]:
                key = str(file.resolve())
                control = controls[key]
                value_list = control["values"]
                source_columns = control["columns"]
                selected_values = tuple(source_columns[index] for index in value_list.curselection())
                settings[key] = {
                    "base_key_columns": (control["base"].get(),),
                    "source_key_columns": (control["source"].get(),),
                    "value_columns": selected_values,
                }
            return settings

        def refresh_preview() -> None:
            try:
                preview = ConsolidationEngine().merge_files_by_columns(files, merge_plans_by_file=current_settings()).result_frame
                self._write_preview_tree(preview_tree, preview, limit=10)
            except Exception as exc:
                self._write_preview_tree(preview_tree, pd.DataFrame([{"미리보기 오류": str(exc)}]), limit=10)

        def cancel() -> None:
            nonlocal result
            result = None
            dialog.destroy()

        def confirm() -> None:
            nonlocal result
            settings = current_settings()
            for setting in settings.values():
                if not setting["base_key_columns"][0] or not setting["source_key_columns"][0] or not setting["value_columns"]:
                    messagebox.showwarning(APP_TITLE, "키 컬럼과 붙일 컬럼을 모두 선택해주세요.")
                    return
            result = settings
            dialog.destroy()

        button_row = ttk.Frame(frame)
        button_row.pack(fill="x", pady=(10, 0))
        ttk.Button(button_row, text="미리보기 새로고침", command=refresh_preview).pack(side="left")
        ttk.Button(button_row, text="취소", command=cancel).pack(side="right")
        ttk.Button(button_row, text="이 설정으로 실행", command=confirm).pack(side="right", padx=(0, 8))
        dialog.protocol("WM_DELETE_WINDOW", cancel)
        refresh_preview()
        self.root.wait_window(dialog)
        return result

    def _confirm_diff_preview(self, before_path: Path, after_path: Path) -> dict[str, object] | None:
        try:
            before_frame = load_table(before_path)
            after_frame = load_table(after_path)
            engine = WorkbookDiffEngine()
            auto_mapping = engine.suggest_column_mapping(before_frame, after_frame)
            initial_result = engine.compare_files(before_path, after_path)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"미리보기를 만들 수 없습니다.\n{exc}")
            return None

        before_columns = [str(column) for column in before_frame.columns]
        after_columns = [str(column) for column in after_frame.columns]
        variables: dict[str, StringVar] = {}
        result: dict[str, object] | None = None

        dialog = Toplevel(self.root)
        dialog.title("전/후 파일 검증 미리보기 / 컬럼 매칭")
        dialog.geometry("980x700")
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding=12)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="비교 기준열과 전/후 컬럼 매칭을 확인한 뒤 실행하세요.", font=("", 11, "bold")).pack(
            anchor="w", pady=(0, 8)
        )

        basis_row = ttk.Frame(frame)
        basis_row.pack(fill="x", pady=(0, 8))
        ttk.Label(basis_row, text="비교 기준열", width=14).pack(side="left")
        initial_basis = str(initial_result.summary.get("compare_basis", ""))
        initial_key = initial_basis.replace("키 컬럼:", "").strip() if initial_basis.startswith("키 컬럼:") else ""
        key_var = StringVar(value=initial_key)
        ttk.Combobox(basis_row, textvariable=key_var, values=["", *before_columns], state="readonly").pack(
            side="left", fill="x", expand=True
        )
        ttk.Label(basis_row, text="비워두면 행 순서 기준").pack(side="left", padx=(8, 0))

        preview_tree = ttk.Treeview(frame, height=8)
        preview_tree.pack(fill="both", expand=True)

        mapping_box = ttk.LabelFrame(frame, text="컬럼 매칭", padding=6)
        mapping_box.pack(fill="both", expand=True, pady=(10, 0))
        inner, _canvas = self._scrollable_frame(mapping_box)
        for row_index, before_column in enumerate(before_columns):
            ttk.Label(inner, text=before_column, width=26).grid(row=row_index, column=0, sticky="w", padx=(4, 8), pady=3)
            selected_after = auto_mapping.get(before_column, before_column if before_column in after_columns else "")
            variable = StringVar(value=selected_after)
            variables[before_column] = variable
            combo = ttk.Combobox(inner, textvariable=variable, values=["", *after_columns], state="readonly", width=36)
            combo.grid(row=row_index, column=1, sticky="ew", padx=(0, 4), pady=3)
        inner.columnconfigure(1, weight=1)

        def current_mapping() -> dict[str, str]:
            return {before_column: variable.get() for before_column, variable in variables.items() if variable.get()}

        def current_key_columns() -> tuple[str, ...]:
            return (key_var.get(),) if key_var.get() else ()

        def refresh_preview() -> None:
            try:
                preview_result = WorkbookDiffEngine().compare_files(
                    before_path,
                    after_path,
                    column_mapping_override=current_mapping(),
                    key_columns=current_key_columns(),
                )
                preview = preview_result.diff_frame
                if preview.empty:
                    preview = pd.DataFrame([{"미리보기": "값 차이가 없습니다.", "비교 기준": preview_result.summary.get("compare_basis", "")}])
                self._write_preview_tree(preview_tree, preview, limit=10)
            except Exception as exc:
                self._write_preview_tree(preview_tree, pd.DataFrame([{"미리보기 오류": str(exc)}]), limit=10)

        def cancel() -> None:
            nonlocal result
            result = None
            dialog.destroy()

        def confirm() -> None:
            nonlocal result
            result = {"mapping": current_mapping(), "key_columns": current_key_columns()}
            dialog.destroy()

        button_row = ttk.Frame(frame)
        button_row.pack(fill="x", pady=(10, 0))
        ttk.Button(button_row, text="미리보기 새로고침", command=refresh_preview).pack(side="left")
        ttk.Button(button_row, text="취소", command=cancel).pack(side="right")
        ttk.Button(button_row, text="이 설정으로 실행", command=confirm).pack(side="right", padx=(0, 8))
        dialog.protocol("WM_DELETE_WINDOW", cancel)
        refresh_preview()
        self.root.wait_window(dialog)
        return result

    def quick_consolidate(self) -> None:
        dialog = Toplevel(self.root)
        dialog.title("엑셀/CSV 파일 여러 개 합치기")
        dialog.geometry("980x720")
        dialog.transient(self.root)
        dialog.grab_set()

        files: list[Path] = []
        merge_mode = StringVar(value="auto")
        mode_status = StringVar(value="자동 선택: 파일을 올리면 행/열 합치기를 추천합니다.")
        manual_row_mappings: dict[str, dict[str, str]] | None = None
        manual_column_plans: dict[str, dict[str, object]] | None = None

        frame = ttk.Frame(dialog, padding=14)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="합칠 파일을 올리고 방향을 선택한 뒤 실행하세요.", font=("", 11, "bold")).pack(
            anchor="w", pady=(0, 8)
        )

        list_frame = ttk.Frame(frame)
        list_frame.pack(fill="x")
        file_list = Listbox(list_frame, height=5)
        file_list.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=file_list.yview)
        scrollbar.pack(side="right", fill="y")
        file_list.configure(yscrollcommand=scrollbar.set)

        def refresh_files() -> None:
            file_list.delete(0, END)
            for index, file in enumerate(files, start=1):
                file_list.insert(END, f"{index}. {file}")

        def add_files() -> None:
            selected = filedialog.askopenfilenames(
                parent=dialog,
                title="합칠 엑셀/CSV 파일들을 선택하세요",
                initialdir=self.base_dir,
                filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv"), ("All files", "*.*")],
            )
            for path in selected:
                file_path = Path(path)
                if file_path not in files:
                    files.append(file_path)
            refresh_files()
            refresh_preview()

        def clear_files() -> None:
            nonlocal manual_row_mappings, manual_column_plans
            files.clear()
            manual_row_mappings = None
            manual_column_plans = None
            refresh_files()
            refresh_preview()

        button_row = ttk.Frame(frame)
        button_row.pack(fill="x", pady=(8, 10))
        ttk.Button(button_row, text="파일 올리기", command=add_files).pack(side="left")
        ttk.Button(button_row, text="목록 비우기", command=clear_files).pack(side="left", padx=(8, 0))

        mode_frame = ttk.LabelFrame(frame, text="합치기 방향", padding=10)
        mode_frame.pack(fill="x")
        ttk.Radiobutton(
            mode_frame,
            text="자동 선택: 파일 구조를 보고 행/열 합치기를 고릅니다.",
            value="auto",
            variable=merge_mode,
        ).pack(anchor="w")
        ttk.Radiobutton(
            mode_frame,
            text="행 합치기: 같은 양식의 여러 파일을 아래로 이어 붙입니다.",
            value="rows",
            variable=merge_mode,
        ).pack(anchor="w", pady=(4, 0))
        ttk.Radiobutton(
            mode_frame,
            text="열 합치기: 공통 키를 찾아 다른 파일의 열을 오른쪽에 붙입니다.",
            value="columns",
            variable=merge_mode,
        ).pack(anchor="w", pady=(4, 0))
        ttk.Label(mode_frame, textvariable=mode_status, foreground="#374151").pack(anchor="w", pady=(8, 0))

        preview_box = ttk.LabelFrame(frame, text="미리보기", padding=8)
        preview_box.pack(fill="both", expand=True, pady=(10, 0))
        preview_box.rowconfigure(0, weight=1)
        preview_box.columnconfigure(0, weight=1)
        preview_text = Text(preview_box, height=14, wrap="none")
        preview_text.grid(row=0, column=0, sticky="nsew")
        preview_scroll_y = ttk.Scrollbar(preview_box, orient="vertical", command=preview_text.yview)
        preview_scroll_y.grid(row=0, column=1, sticky="ns")
        preview_scroll_x = ttk.Scrollbar(preview_box, orient="horizontal", command=preview_text.xview)
        preview_scroll_x.grid(row=1, column=0, sticky="ew")
        preview_text.configure(
            yscrollcommand=preview_scroll_y.set,
            xscrollcommand=preview_scroll_x.set,
            state="disabled",
        )

        def effective_merge_mode() -> str:
            selected = merge_mode.get()
            if selected != "auto":
                return selected
            if len(files) < 2:
                return "rows"
            return ConsolidationEngine().recommend_merge_mode(files)

        def build_preview_frame() -> pd.DataFrame:
            if not files:
                return pd.DataFrame([{"안내": "파일 올리기를 눌러 합칠 파일을 선택하세요."}])
            mode = effective_merge_mode()
            if mode == "columns":
                result = ConsolidationEngine().merge_files_by_columns(files, merge_plans_by_file=manual_column_plans)
            else:
                result = ConsolidationEngine().consolidate_files(
                    files,
                    template=None,
                    saved_mappings_by_file=manual_row_mappings,
                )
            if result.result_frame.empty:
                return pd.DataFrame([{"안내": "미리볼 데이터가 없습니다.", "확인 필요": len(result.issues)}])
            return result.result_frame.head(10)

        def refresh_preview() -> None:
            try:
                if files:
                    mode = effective_merge_mode()
                    mode_status.set("자동 선택 결과: 열 합치기" if mode == "columns" else "자동 선택 결과: 행 합치기")
                else:
                    mode_status.set("자동 선택: 파일을 올리면 행/열 합치기를 추천합니다.")
                self._write_preview_text(preview_text, build_preview_frame(), limit=10)
            except Exception as exc:
                self._write_preview_text(preview_text, pd.DataFrame([{"미리보기 오류": str(exc)}]), limit=10)

        merge_mode.trace_add("write", lambda *_args: refresh_preview())

        def edit_mapping() -> None:
            nonlocal manual_row_mappings, manual_column_plans
            if not files:
                messagebox.showwarning(APP_TITLE, "먼저 합칠 파일을 올려주세요.")
                return
            if effective_merge_mode() == "columns":
                selected = self._confirm_column_merge_preview(files)
                if selected is not None:
                    manual_column_plans = selected
            else:
                selected = self._confirm_row_merge_preview(files)
                if selected is not None:
                    manual_row_mappings = selected
            refresh_preview()

        run_row = ttk.Frame(frame)
        run_row.pack(fill="x", pady=(14, 0))

        def cancel() -> None:
            dialog.destroy()

        def execute() -> None:
            if not files:
                messagebox.showwarning(APP_TITLE, "먼저 합칠 파일을 올려주세요.")
                return
            selected_files = list(files)
            selected_mode = effective_merge_mode()
            output = self._ask_save_path_or_none(
                "합치기 결과를 저장할 위치를 선택하세요",
                "파일열합치기_결과" if selected_mode == "columns" else "파일행합치기_결과",
            )
            if not output:
                return
            dialog.destroy()

            def job():
                if selected_mode == "columns":
                    result = ConsolidationEngine().merge_files_by_columns(
                        selected_files,
                        merge_plans_by_file=manual_column_plans,
                    )
                else:
                    result = ConsolidationEngine().consolidate_files(
                        selected_files,
                        template=None,
                        saved_mappings_by_file=manual_row_mappings,
                    )
                ReportWriter().write_xlsx(result, output, mark_result_cells=True, include_privacy_scan=False)
                return [output]

            label = "열 방향 파일 합치기" if selected_mode == "columns" else "행 방향 파일 합치기"
            self._run(label, job, open_path=output.parent)

        ttk.Button(run_row, text="미리보기 새로고침", command=refresh_preview).pack(side="left")
        ttk.Button(run_row, text="컬럼 매칭 수정", command=edit_mapping).pack(side="left", padx=(8, 0))
        ttk.Button(run_row, text="취소", command=cancel).pack(side="right")
        ttk.Button(run_row, text="실행", command=execute).pack(side="right", padx=(0, 8))
        dialog.protocol("WM_DELETE_WINDOW", cancel)
        refresh_preview()
        self.root.wait_window(dialog)

    def quick_diff_workbooks(self) -> None:
        dialog = Toplevel(self.root)
        dialog.title("전/후 파일 검증")
        dialog.geometry("980x680")
        dialog.transient(self.root)
        dialog.grab_set()

        before_file = StringVar()
        after_file = StringVar()
        manual_settings: dict[str, object] | None = None

        frame = ttk.Frame(dialog, padding=14)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="수정 전 파일과 수정 후 파일을 올린 뒤 실행하세요.", font=("", 11, "bold")).pack(
            anchor="w", pady=(0, 10)
        )

        def choose(variable: StringVar, title: str) -> None:
            nonlocal manual_settings
            selected = filedialog.askopenfilename(
                parent=dialog,
                title=title,
                initialdir=self.base_dir,
                filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv"), ("All files", "*.*")],
            )
            if selected:
                variable.set(selected)
                manual_settings = None
                refresh_preview()

        def file_row(label: str, variable: StringVar, title: str) -> None:
            row = ttk.Frame(frame)
            row.pack(fill="x", pady=5)
            ttk.Label(row, text=label, width=12).pack(side="left")
            ttk.Entry(row, textvariable=variable).pack(side="left", fill="x", expand=True)
            ttk.Button(row, text="파일 선택", command=lambda: choose(variable, title)).pack(side="left", padx=(8, 0))

        file_row("수정 전", before_file, "수정 전 엑셀/CSV 파일을 선택하세요")
        file_row("수정 후", after_file, "수정 후 엑셀/CSV 파일을 선택하세요")

        ttk.Label(
            frame,
            text="결과의 후파일_메모 시트에서 노란색 셀을 열면 전 값/후 값 메모를 볼 수 있습니다.",
            wraplength=620,
        ).pack(anchor="w", pady=(12, 0))

        preview_box = ttk.LabelFrame(frame, text="미리보기", padding=8)
        preview_box.pack(fill="both", expand=True, pady=(10, 0))
        preview_box.rowconfigure(0, weight=1)
        preview_box.columnconfigure(0, weight=1)
        preview_text = Text(preview_box, height=14, wrap="none")
        preview_text.grid(row=0, column=0, sticky="nsew")
        preview_scroll_y = ttk.Scrollbar(preview_box, orient="vertical", command=preview_text.yview)
        preview_scroll_y.grid(row=0, column=1, sticky="ns")
        preview_scroll_x = ttk.Scrollbar(preview_box, orient="horizontal", command=preview_text.xview)
        preview_scroll_x.grid(row=1, column=0, sticky="ew")
        preview_text.configure(
            yscrollcommand=preview_scroll_y.set,
            xscrollcommand=preview_scroll_x.set,
            state="disabled",
        )

        def current_preview_settings() -> dict[str, object]:
            return manual_settings or {"mapping": {}, "key_columns": ()}

        def build_preview_frame() -> pd.DataFrame:
            if not before_file.get() or not after_file.get():
                return pd.DataFrame([{"안내": "수정 전 파일과 수정 후 파일을 모두 선택하세요."}])
            settings = current_preview_settings()
            result = WorkbookDiffEngine().compare_files(
                Path(before_file.get()),
                Path(after_file.get()),
                column_mapping_override=settings["mapping"],
                key_columns=settings["key_columns"],
            )
            if result.diff_frame.empty:
                return pd.DataFrame(
                    [
                        {
                            "미리보기": "값 차이가 없습니다.",
                            "비교 기준": result.summary.get("compare_basis", ""),
                            "행 차이": result.summary.get("missing_row_count", 0)
                            + result.summary.get("added_row_count", 0),
                            "컬럼 차이": result.summary.get("before_only_column_count", 0)
                            + result.summary.get("after_only_column_count", 0),
                        }
                    ]
                )
            return result.diff_frame.head(10)

        def refresh_preview() -> None:
            try:
                self._write_preview_text(preview_text, build_preview_frame(), limit=10)
            except Exception as exc:
                self._write_preview_text(preview_text, pd.DataFrame([{"미리보기 오류": str(exc)}]), limit=10)

        def edit_mapping() -> None:
            nonlocal manual_settings
            if not before_file.get() or not after_file.get():
                messagebox.showwarning(APP_TITLE, "수정 전 파일과 수정 후 파일을 모두 선택해주세요.")
                return
            selected = self._confirm_diff_preview(Path(before_file.get()), Path(after_file.get()))
            if selected is not None:
                manual_settings = selected
            refresh_preview()

        button_row = ttk.Frame(frame)
        button_row.pack(fill="x", pady=(18, 0))

        def cancel() -> None:
            dialog.destroy()

        def execute() -> None:
            if not before_file.get() or not after_file.get():
                messagebox.showwarning(APP_TITLE, "수정 전 파일과 수정 후 파일을 모두 선택해주세요.")
                return
            before_path = Path(before_file.get())
            after_path = Path(after_file.get())
            preview_settings = current_preview_settings()
            output = self._ask_save_path_or_none("전/후 검증 결과를 저장할 위치를 선택하세요", "전후파일_검증결과")
            if not output:
                return
            dialog.destroy()

            def job():
                result = WorkbookDiffEngine().compare_files(
                    before_path,
                    after_path,
                    column_mapping_override=preview_settings["mapping"],
                    key_columns=preview_settings["key_columns"],
                )
                WorkbookDiffReportWriter().write_xlsx(result, output)
                return [output]

            self._run("전/후 파일 검증", job, open_path=output.parent)

        ttk.Button(button_row, text="미리보기 새로고침", command=refresh_preview).pack(side="left")
        ttk.Button(button_row, text="컬럼 매칭 수정", command=edit_mapping).pack(side="left", padx=(8, 0))
        ttk.Button(button_row, text="취소", command=cancel).pack(side="right")
        ttk.Button(button_row, text="실행", command=execute).pack(side="right", padx=(0, 8))
        dialog.protocol("WM_DELETE_WINDOW", cancel)
        refresh_preview()
        self.root.wait_window(dialog)

    def quick_pivot_summary(self) -> None:
        no_column_label = "(선택 안 함)"
        row_count_label = "(행 개수)"

        dialog = Toplevel(self.root)
        dialog.title("피벗 요약표 만들기")
        dialog.geometry("980x720")
        dialog.transient(self.root)
        dialog.grab_set()

        tools = AdminWorkbookTools()
        file_path = StringVar()
        row_column = StringVar()
        column_column = StringVar(value=no_column_label)
        value_column = StringVar(value=row_count_label)
        aggregation = StringVar(value="건수")
        table_frame: pd.DataFrame | None = None
        table_metadata: dict[str, object] = {}

        frame = ttk.Frame(dialog, padding=14)
        frame.pack(fill="both", expand=True)
        ttk.Label(
            frame,
            text="파일을 올린 뒤 행 기준, 열 기준, 값 열, 집계 방식을 선택하세요.",
            font=("", 11, "bold"),
        ).pack(anchor="w", pady=(0, 10))

        file_row = ttk.Frame(frame)
        file_row.pack(fill="x", pady=5)
        ttk.Label(file_row, text="원본 파일", width=12).pack(side="left")
        ttk.Entry(file_row, textvariable=file_path).pack(side="left", fill="x", expand=True)

        option_box = ttk.LabelFrame(frame, text="요약 기준", padding=10)
        option_box.pack(fill="x", pady=(10, 0))
        for column_index in range(4):
            option_box.columnconfigure(column_index, weight=1)

        ttk.Label(option_box, text="행 기준").grid(row=0, column=0, sticky="w")
        ttk.Label(option_box, text="열 기준").grid(row=0, column=1, sticky="w")
        ttk.Label(option_box, text="값 열").grid(row=0, column=2, sticky="w")
        ttk.Label(option_box, text="집계 방식").grid(row=0, column=3, sticky="w")

        row_combo = ttk.Combobox(option_box, textvariable=row_column, state="readonly")
        column_combo = ttk.Combobox(option_box, textvariable=column_column, state="readonly")
        value_combo = ttk.Combobox(option_box, textvariable=value_column, state="readonly")
        aggregation_combo = ttk.Combobox(
            option_box,
            textvariable=aggregation,
            state="readonly",
            values=["건수", "합계", "평균", "최대", "최소"],
        )
        row_combo.grid(row=1, column=0, sticky="ew", padx=(0, 8), pady=(4, 0))
        column_combo.grid(row=1, column=1, sticky="ew", padx=(0, 8), pady=(4, 0))
        value_combo.grid(row=1, column=2, sticky="ew", padx=(0, 8), pady=(4, 0))
        aggregation_combo.grid(row=1, column=3, sticky="ew", pady=(4, 0))

        ttk.Label(
            frame,
            text="건수는 값 열 없이 행 개수를 셉니다. 합계/평균/최대/최소는 숫자 값 열을 선택하세요.",
            wraplength=780,
        ).pack(anchor="w", pady=(8, 0))

        preview_box = ttk.LabelFrame(frame, text="미리보기", padding=8)
        preview_box.pack(fill="both", expand=True, pady=(10, 0))
        preview_box.rowconfigure(0, weight=1)
        preview_box.columnconfigure(0, weight=1)
        preview_text = Text(preview_box, height=14, wrap="none")
        preview_text.grid(row=0, column=0, sticky="nsew")
        preview_scroll_y = ttk.Scrollbar(preview_box, orient="vertical", command=preview_text.yview)
        preview_scroll_y.grid(row=0, column=1, sticky="ns")
        preview_scroll_x = ttk.Scrollbar(preview_box, orient="horizontal", command=preview_text.xview)
        preview_scroll_x.grid(row=1, column=0, sticky="ew")
        preview_text.configure(
            yscrollcommand=preview_scroll_y.set,
            xscrollcommand=preview_scroll_x.set,
            state="disabled",
        )

        def selected_column_value() -> str | None:
            return None if column_column.get() == no_column_label else column_column.get()

        def selected_value_column() -> str | None:
            if aggregation.get() == "건수" or value_column.get() == row_count_label:
                return None
            return value_column.get()

        def refresh_combos(columns: list[str]) -> None:
            row_combo.configure(values=columns)
            column_combo.configure(values=[no_column_label, *columns])
            value_combo.configure(values=[row_count_label, *columns])

        def build_preview_frame() -> pd.DataFrame:
            if table_frame is None:
                return pd.DataFrame([{"안내": "파일 선택을 눌러 요약할 엑셀/CSV 파일을 선택하세요."}])
            if not row_column.get():
                return pd.DataFrame([{"안내": "행 기준 열을 선택하세요."}])
            result = tools.build_pivot_summary(
                table_frame,
                row_column=row_column.get(),
                column_column=selected_column_value(),
                value_column=selected_value_column(),
                aggregation=aggregation.get(),
                metadata=table_metadata,
            )
            return result.pivot_frame.head(10)

        def refresh_preview() -> None:
            try:
                self._write_preview_text(preview_text, build_preview_frame(), limit=10)
            except Exception as exc:
                self._write_preview_text(preview_text, pd.DataFrame([{"미리보기 오류": str(exc)}]), limit=10)

        def choose_file() -> None:
            nonlocal table_frame, table_metadata
            selected = filedialog.askopenfilename(
                parent=dialog,
                title="피벗 요약표를 만들 엑셀/CSV 파일을 선택하세요",
                initialdir=self.base_dir,
                filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv"), ("All files", "*.*")],
            )
            if not selected:
                return
            try:
                table_frame, table_metadata = tools.prepare_pivot_source(selected)
                defaults = tools.infer_pivot_defaults(table_frame)
            except Exception as exc:
                messagebox.showerror(APP_TITLE, f"파일을 읽을 수 없습니다.\n{exc}")
                return
            columns = [str(column) for column in table_frame.columns]
            file_path.set(selected)
            refresh_combos(columns)
            row_column.set(defaults["row_column"] if defaults["row_column"] in columns else (columns[0] if columns else ""))
            column_column.set(defaults["column_column"] if defaults["column_column"] in columns else no_column_label)
            value_column.set(defaults["value_column"] if defaults["value_column"] in columns else row_count_label)
            aggregation.set(defaults["aggregation"])
            refresh_preview()

        def aggregation_changed(_event=None) -> None:
            if aggregation.get() != "건수" and value_column.get() == row_count_label and table_frame is not None:
                default_value = tools.infer_pivot_defaults(table_frame).get("value_column", "")
                if default_value:
                    value_column.set(default_value)
            refresh_preview()

        ttk.Button(file_row, text="파일 선택", command=choose_file).pack(side="left", padx=(8, 0))
        row_combo.bind("<<ComboboxSelected>>", lambda _event: refresh_preview())
        column_combo.bind("<<ComboboxSelected>>", lambda _event: refresh_preview())
        value_combo.bind("<<ComboboxSelected>>", lambda _event: refresh_preview())
        aggregation_combo.bind("<<ComboboxSelected>>", aggregation_changed)

        button_row = ttk.Frame(frame)
        button_row.pack(fill="x", pady=(18, 0))

        def cancel() -> None:
            dialog.destroy()

        def execute() -> None:
            if not file_path.get() or table_frame is None:
                messagebox.showwarning(APP_TITLE, "먼저 요약할 파일을 선택하세요.")
                return
            if not row_column.get():
                messagebox.showwarning(APP_TITLE, "행 기준 열을 선택하세요.")
                return
            selected_file = Path(file_path.get())
            selected_row = row_column.get()
            selected_column = selected_column_value()
            selected_value = selected_value_column()
            selected_aggregation = aggregation.get()
            output = self._ask_save_path_or_none("피벗 요약표 결과를 저장할 위치를 선택하세요", "피벗요약표_결과")
            if not output:
                return
            dialog.destroy()

            def job():
                tools.write_pivot_workbook(
                    selected_file,
                    output,
                    row_column=selected_row,
                    column_column=selected_column,
                    value_column=selected_value,
                    aggregation=selected_aggregation,
                )
                return [output]

            self._run("피벗 요약표 만들기", job, open_path=output.parent)

        ttk.Button(button_row, text="미리보기 새로고침", command=refresh_preview).pack(side="left")
        ttk.Button(button_row, text="취소", command=cancel).pack(side="right")
        ttk.Button(button_row, text="실행", command=execute).pack(side="right", padx=(0, 8))
        dialog.protocol("WM_DELETE_WINDOW", cancel)
        refresh_preview()
        self.root.wait_window(dialog)

    def quick_lookup(self) -> None:
        reference = self._ask_file_or_none("기준표 파일을 선택하세요")
        if not reference:
            return
        target = self._ask_file_or_none("값을 붙일 대상 파일을 선택하세요")
        if not target:
            return
        output = self._ask_save_path_or_none("값 붙이기 결과를 저장할 위치를 선택하세요", "기준표_값붙이기")
        if not output:
            return

        def job():
            reference_frame = load_table(reference)
            target_frame = load_table(target)
            plan = AutoLookupPlanner().infer_lookup_plan(reference_frame, target_frame)
            result = MergeEngine().merge_lookup(reference_frame, target_frame, plan.key_spec, list(plan.value_columns))
            attach_auto_summary(result, plan, "기준표에서 값 붙이기")
            ReportWriter().write_xlsx(result, output)
            return [output]

        self._run("기준표 값 붙이기", job, open_path=output.parent)

    def quick_reconcile(self) -> None:
        reference = self._ask_file_or_none("기준 명단 파일을 선택하세요")
        if not reference:
            return
        target = self._ask_file_or_none("비교할 제출/이수 명단 파일을 선택하세요")
        if not target:
            return
        output = self._ask_save_path_or_none("누락 확인 결과를 저장할 위치를 선택하세요", "누락확인_결과")
        if not output:
            return

        def job():
            reference_frame = load_table(reference)
            target_frame = load_table(target)
            plan = AutoLookupPlanner().infer_reconciliation_key_spec(reference_frame, target_frame)
            result = ReconciliationEngine().compare_lists(reference_frame, target_frame, plan.key_spec)
            attach_auto_summary(result, plan, "빠진 사람/누락자료 찾기")
            ReportWriter().write_xlsx(result, output)
            return [output]

        self._run("빠진 사람/기관 찾기", job, open_path=output.parent)

    def quick_horizontal(self) -> None:
        file_path = self._ask_file_or_none("월별 가로표 파일을 선택하세요")
        if not file_path:
            return
        output = self._ask_save_path_or_none("가로표 변환 결과를 저장할 위치를 선택하세요", "월별가로표_세로변환")
        if not output:
            return

        def job():
            table = load_table(file_path)
            engine = HorizontalTableEngine()
            detection = engine.detect(table)
            id_columns = [column for column in table.columns if column not in detection.value_columns]
            converted = engine.wide_to_long(table, id_columns=id_columns)
            result = JobResult(
                result_frame=converted,
                summary={"workflow": "월별 가로표 세로 변환", "row_count": len(converted)},
            )
            ReportWriter().write_xlsx(result, output)
            return [output]

        self._run("월별 가로표 세로 변환", job, open_path=output.parent)

    def _build_consolidate_tab(self, notebook: ttk.Notebook) -> None:
        frame = ttk.Frame(notebook, padding=12)
        notebook.add(frame, text="제출자료 수합")
        self.consolidate_folder = StringVar()
        self.consolidate_output = StringVar(value=str(self.output_dir / "제출자료_수합결과.xlsx"))
        self.consolidate_columns = StringVar(value=", ".join(get_template("school_submission_consolidation").standard_columns))
        self._path_row(frame, "수합 폴더", self.consolidate_folder, self.choose_folder)
        self._entry_row(frame, "표준 컬럼", self.consolidate_columns)
        self._path_row(frame, "결과 파일", self.consolidate_output, self.choose_save_file)
        ttk.Button(frame, text="수합 실행", command=self.run_consolidate).pack(anchor="w", pady=10)

    def _build_lookup_tab(self, notebook: ttk.Notebook) -> None:
        frame = ttk.Frame(notebook, padding=12)
        notebook.add(frame, text="값 붙이기")
        self.lookup_reference = StringVar()
        self.lookup_target = StringVar()
        self.lookup_key = StringVar()
        self.lookup_target_key = StringVar()
        self.lookup_values = StringVar()
        self.lookup_output = StringVar(value=str(self.output_dir / "기준표_값붙이기.xlsx"))
        self._path_row(frame, "기준 파일", self.lookup_reference, self.choose_file)
        self._path_row(frame, "대상 파일", self.lookup_target, self.choose_file)
        self._entry_row(frame, "기준 키 컬럼", self.lookup_key)
        self._entry_row(frame, "대상 키 컬럼", self.lookup_target_key)
        self._entry_row(frame, "가져올 컬럼", self.lookup_values)
        self._path_row(frame, "결과 파일", self.lookup_output, self.choose_save_file)
        ttk.Label(frame, text="키/가져올 컬럼을 비워도 컬럼명과 데이터 겹침으로 자동 추천합니다.").pack(anchor="w")
        ttk.Button(frame, text="값 붙이기 실행", command=self.run_lookup).pack(anchor="w", pady=10)

    def _build_reconcile_tab(self, notebook: ttk.Notebook) -> None:
        frame = ttk.Frame(notebook, padding=12)
        notebook.add(frame, text="누락 확인")
        self.reconcile_reference = StringVar()
        self.reconcile_target = StringVar()
        self.reconcile_key = StringVar()
        self.reconcile_target_key = StringVar()
        self.reconcile_output = StringVar(value=str(self.output_dir / "누락확인_결과.xlsx"))
        self._path_row(frame, "기준 명단", self.reconcile_reference, self.choose_file)
        self._path_row(frame, "대조 명단", self.reconcile_target, self.choose_file)
        self._entry_row(frame, "기준 키 컬럼", self.reconcile_key)
        self._entry_row(frame, "대상 키 컬럼", self.reconcile_target_key)
        self._path_row(frame, "결과 파일", self.reconcile_output, self.choose_save_file)
        ttk.Label(frame, text="키 컬럼을 비우면 자동 추천합니다.").pack(anchor="w")
        ttk.Button(frame, text="누락 확인 실행", command=self.run_reconcile).pack(anchor="w", pady=10)

    def _build_horizontal_tab(self, notebook: ttk.Notebook) -> None:
        frame = ttk.Frame(notebook, padding=12)
        notebook.add(frame, text="가로표 변환")
        self.horizontal_file = StringVar()
        self.horizontal_id_columns = StringVar()
        self.horizontal_output = StringVar(value=str(self.output_dir / "월별가로표_세로변환.xlsx"))
        self._path_row(frame, "가로표 파일", self.horizontal_file, self.choose_file)
        self._entry_row(frame, "행 기준 컬럼", self.horizontal_id_columns)
        self._path_row(frame, "결과 파일", self.horizontal_output, self.choose_save_file)
        ttk.Label(frame, text="행 기준 컬럼을 비우면 월/분기 컬럼을 제외한 나머지를 기준으로 사용합니다.").pack(anchor="w")
        ttk.Button(frame, text="세로 변환 실행", command=self.run_horizontal).pack(anchor="w", pady=10)

    def _entry_row(self, parent, label: str, variable: StringVar) -> None:
        row = ttk.Frame(parent)
        row.pack(fill="x", pady=4)
        ttk.Label(row, text=label, width=14).pack(side="left")
        ttk.Entry(row, textvariable=variable).pack(side="left", fill="x", expand=True)

    def _path_row(self, parent, label: str, variable: StringVar, chooser) -> None:
        row = ttk.Frame(parent)
        row.pack(fill="x", pady=4)
        ttk.Label(row, text=label, width=14).pack(side="left")
        ttk.Entry(row, textvariable=variable).pack(side="left", fill="x", expand=True)
        ttk.Button(row, text="선택", command=lambda: chooser(variable)).pack(side="left", padx=(6, 0))

    def choose_folder(self, variable: StringVar) -> None:
        selected = filedialog.askdirectory(initialdir=self.base_dir)
        if selected:
            variable.set(selected)

    def choose_file(self, variable: StringVar) -> None:
        selected = filedialog.askopenfilename(
            initialdir=self.base_dir,
            filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv"), ("All files", "*.*")],
        )
        if selected:
            variable.set(selected)

    def choose_save_file(self, variable: StringVar) -> None:
        selected = filedialog.asksaveasfilename(
            initialdir=self.output_dir,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if selected:
            variable.set(selected if selected.lower().endswith(".xlsx") else f"{selected}.xlsx")

    def run_consolidate(self) -> None:
        def job():
            result = ConsolidationEngine().consolidate_folder(
                self.consolidate_folder.get(),
                standard_columns=split_columns(self.consolidate_columns.get()) or None,
                template="school_submission_consolidation",
            )
            output = Path(self.consolidate_output.get())
            ReportWriter().write_xlsx(result, output, mark_result_cells=False, include_privacy_scan=False)
            return [output]

        self._run("제출자료 수합", job, open_path=Path(self.consolidate_output.get()).parent)

    def run_lookup(self) -> None:
        def job():
            reference = load_table(Path(self.lookup_reference.get()))
            target = load_table(Path(self.lookup_target.get()))
            plan = AutoLookupPlanner().infer_lookup_plan(
                reference,
                target,
                preferred_reference_key_columns=tuple(split_columns(self.lookup_key.get())),
                preferred_target_key_columns=tuple(split_columns(self.lookup_target_key.get())),
                preferred_value_columns=tuple(split_columns(self.lookup_values.get())),
            )
            result = MergeEngine().merge_lookup(reference, target, plan.key_spec, list(plan.value_columns))
            attach_auto_summary(result, plan, "기준표에서 값 붙이기")
            output = Path(self.lookup_output.get())
            ReportWriter().write_xlsx(result, output)
            return [output]

        self._run("값 붙이기", job, open_path=Path(self.lookup_output.get()).parent)

    def run_reconcile(self) -> None:
        def job():
            reference = load_table(Path(self.reconcile_reference.get()))
            target = load_table(Path(self.reconcile_target.get()))
            plan = AutoLookupPlanner().infer_reconciliation_key_spec(
                reference,
                target,
                preferred_reference_key_columns=tuple(split_columns(self.reconcile_key.get())),
                preferred_target_key_columns=tuple(split_columns(self.reconcile_target_key.get())),
            )
            result = ReconciliationEngine().compare_lists(reference, target, plan.key_spec)
            attach_auto_summary(result, plan, "빠진 사람/누락자료 찾기")
            output = Path(self.reconcile_output.get())
            ReportWriter().write_xlsx(result, output)
            return [output]

        self._run("누락 확인", job, open_path=Path(self.reconcile_output.get()).parent)

    def run_horizontal(self) -> None:
        def job():
            table = load_table(Path(self.horizontal_file.get()))
            engine = HorizontalTableEngine()
            detection = engine.detect(table)
            id_columns = split_columns(self.horizontal_id_columns.get()) or [
                column for column in table.columns if column not in detection.value_columns
            ]
            converted = engine.wide_to_long(table, id_columns=id_columns)
            result = JobResult(
                result_frame=converted,
                summary={"workflow": "월별 가로표 세로 변환", "row_count": len(converted)},
            )
            output = Path(self.horizontal_output.get())
            ReportWriter().write_xlsx(result, output)
            return [output]

        self._run("가로표 변환", job, open_path=Path(self.horizontal_output.get()).parent)

    def _run(self, label: str, job, open_path: Path | None = None) -> None:
        try:
            self.status.set(f"{label} 실행 중...")
            self.root.update_idletasks()
            outputs = job()
            self.log_insert(f"[완료] {label}")
            for output in outputs:
                self.log_insert(f"  - {output}")
            if outputs:
                self.last_output_dir = Path(outputs[0]).parent
            self.status.set(f"{label} 완료")
            if open_path:
                self.open_folder(open_path)
            messagebox.showinfo(APP_TITLE, f"{label}이 완료되었습니다.")
        except Exception as exc:
            self.status.set(f"{label} 실패")
            self.log_insert(f"[실패] {label}: {exc}")
            self.log_insert(traceback.format_exc())
            messagebox.showerror(APP_TITLE, str(exc))

    def log_insert(self, text: str) -> None:
        self.log.insert(END, text + "\n")
        self.log.see(END)

    def open_folder(self, path: Path) -> None:
        path.mkdir(parents=True, exist_ok=True)
        os.startfile(path)


def main() -> int:
    if "--smoke" in sys.argv:
        samples = resource_path("samples/public_admin")
        if not samples.exists():
            raise SystemExit(f"sample data missing: {samples}")
        return 0
    root = Tk()
    LocalApp(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
