from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from vhlookup_core import (
    AutoLookupPlanner,
    ExcelLoader,
    HeaderDetector,
    ReportWriter,
    SheetDetector,
    get_template,
    template_catalog_frame,
    template_usage_frame,
)
from vhlookup_core.consolidation import ConsolidationEngine
from vhlookup_core.horizontal import HorizontalTableEngine
from vhlookup_core.merge import MergeEngine
from vhlookup_core.models import JobResult
from vhlookup_core.reconciliation import ReconciliationEngine


SAMPLES = ROOT / "samples" / "public_admin"
OUTPUT = ROOT / "demo_output"


DEMO_TOUR_ROWS = [
    {
        "순서": "01",
        "결과 파일": "01_제출자료_수합결과.xlsx",
        "업무 상황": "학교/부서/기관에서 받은 제출자료 수합",
        "열어볼 시트": "결과, 확인사항",
        "확인 포인트": "제목 행과 안내문을 건너뛰고 컬럼명을 자동으로 맞춘 통합본, 담당자/연락처 점검",
    },
    {
        "순서": "02",
        "결과 파일": "02_오류검증_제출자료.xlsx",
        "업무 상황": "제출자료 오류 검토",
        "열어볼 시트": "결과, 확인사항",
        "확인 포인트": "필수값 누락, 숫자 오류, 날짜 오류, 중복 제출 의심, 담당자/연락처 점검",
    },
    {
        "순서": "03",
        "결과 파일": "03_교육이수_명단대조.xlsx",
        "업무 상황": "교육이수 명단에 직원 기본정보 붙이기",
        "열어볼 시트": "결과, 확인사항",
        "확인 포인트": "사번과 직원번호를 자동 연결하고 부서/직급/소속을 붙이며 성명 컬럼을 점검",
    },
    {
        "순서": "04",
        "결과 파일": "04_교육누락자_대조결과.xlsx",
        "업무 상황": "교육 대상자 중 빠진 사람 찾기",
        "열어볼 시트": "결과, 확인사항",
        "확인 포인트": "기준명단에만 있음, 교육이수 명단에만 있음, 양쪽 모두 있음, 성명 컬럼 점검",
    },
    {
        "순서": "05",
        "결과 파일": "05_수당예산_대조결과.xlsx",
        "업무 상황": "수당/예산 지급대상자 기준표 대조",
        "열어볼 시트": "결과, 확인사항",
        "확인 포인트": "사번+지급월 복합 키로 단가/지급기준/예산과목을 붙이고 성명 컬럼을 점검",
    },
    {
        "순서": "06",
        "결과 파일": "06_제출대상_누락확인.xlsx",
        "업무 상황": "제출 대상 기관 누락 확인",
        "열어볼 시트": "결과, 확인사항",
        "확인 포인트": "제출 대상인데 미제출한 기관과 대상이 아닌 제출 기관 분리",
    },
    {
        "순서": "07",
        "결과 파일": "07_부서별현황_수합결과.xlsx",
        "업무 상황": "부서별 현황자료 수합",
        "열어볼 시트": "결과, 확인사항",
        "확인 포인트": "담당 부서/소속, 현원/인원수처럼 다른 표현을 표준 컬럼으로 통합",
    },
    {
        "순서": "08",
        "결과 파일": "08_월별가로표_세로변환.xlsx",
        "업무 상황": "월별 가로표 세로 변환",
        "열어볼 시트": "결과",
        "확인 포인트": "1월, 2월, 3월 컬럼을 세로형 자료로 변환",
    },
]


DEMO_SECURITY_ROWS = [
    {"항목": "AI 사용", "내용": "사용 안 함"},
    {"항목": "외부 API", "내용": "호출 안 함"},
    {"항목": "클라우드 업로드", "내용": "없음"},
    {"항목": "원본 파일", "내용": "수정하지 않음"},
    {"항목": "결과 저장", "내용": "demo_output 폴더에 새 xlsx 파일로 저장"},
    {"항목": "검토 방식", "내용": "첫 시트는 실제 결과, 확인사항 시트는 자동 매칭 근거와 확인 항목"},
    {"항목": "개인정보 점검", "내용": "이름, 연락처, 이메일, 주소, 계좌, 주민등록번호 패턴을 실제 값 저장 없이 유형/건수로 표시"},
]


def load_table(path: Path):
    loader = ExcelLoader()
    header_detector = HeaderDetector()
    sheet_detector = SheetDetector(header_detector)
    source = loader.load(path)
    sheet = sheet_detector.select(source)
    detection = header_detector.detect(sheet)
    return header_detector.apply(sheet, detection)


def attach_auto_summary(result: JobResult, plan, workflow: str) -> JobResult:
    result.mapping_records.extend(plan.evidence_rows)
    result.summary["workflow"] = workflow
    result.summary["auto_key_columns"] = " + ".join(plan.key_spec.reference_key_columns)
    result.summary["auto_target_key_columns"] = " + ".join(plan.key_spec.target_columns())
    if plan.value_columns:
        result.summary["auto_value_columns"] = ", ".join(plan.value_columns)
    return result


def write_demo_tour(output_dir: Path) -> None:
    path = output_dir / "00_샘플_둘러보기.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(DEMO_TOUR_ROWS).to_excel(writer, sheet_name="샘플순서", index=False)
        pd.DataFrame(DEMO_SECURITY_ROWS).to_excel(writer, sheet_name="보안원칙", index=False)
        template_catalog_frame().to_excel(writer, sheet_name="업무템플릿", index=False)
        template_usage_frame().to_excel(writer, sheet_name="명령예시", index=False)
        pd.DataFrame(
            [
                {"단계": 1, "할 일": "00_샘플_둘러보기.xlsx에서 샘플순서 시트를 봅니다."},
                {"단계": 2, "할 일": "02_오류검증_제출자료.xlsx의 확인사항 시트를 봅니다."},
                {"단계": 3, "할 일": "03_교육이수_명단대조.xlsx의 확인사항 시트에서 자동 매칭 근거를 봅니다."},
                {"단계": 4, "할 일": "05_수당예산_대조결과.xlsx에서 복합 키 대조 결과를 봅니다."},
                {"단계": 5, "할 일": "01_제출자료_수합결과.xlsx의 확인사항 시트에서 담당자/연락처 점검 결과를 봅니다."},
                {"단계": 6, "할 일": "원본 CSV 파일이 그대로 남아 있는지 확인합니다."},
            ]
        ).to_excel(writer, sheet_name="추천동선", index=False)
        style_workbook(writer.book)


def style_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F6F5F")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row > 1 and sheet.max_column > 0:
            sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value), 48))
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            sheet.column_dimensions[column_letter].width = max(10, min(max_length + 2, 50))


def main() -> int:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    for old_output in OUTPUT.glob("*.xlsx"):
        old_output.unlink()
    writer = ReportWriter()
    write_demo_tour(OUTPUT)
    merge_samples = SAMPLES / "03_merge_files"
    hr_samples = merge_samples / "column_merge_hr_training"
    extra_samples = SAMPLES / "90_extra_cli_samples"

    consolidated = ConsolidationEngine().consolidate_folder(
        merge_samples / "row_merge_school_submissions",
        template="school_submission_consolidation",
    )
    writer.write_xlsx(consolidated, OUTPUT / "01_제출자료_수합결과.xlsx", include_privacy_scan=False)

    invalid_consolidated = ConsolidationEngine().consolidate_folder(
        merge_samples / "row_merge_submission_errors",
        template="school_submission_consolidation",
    )
    writer.write_xlsx(invalid_consolidated, OUTPUT / "02_오류검증_제출자료.xlsx", include_privacy_scan=False)

    employee_master = load_table(hr_samples / "hr_employee_master.csv")
    training = load_table(hr_samples / "hr_training_completion.csv")
    lookup = ConsolidationEngine().merge_files_by_columns(
        [hr_samples / "hr_employee_master.csv", hr_samples / "hr_training_completion.csv"]
    )
    writer.write_xlsx(lookup, OUTPUT / "03_교육이수_명단대조.xlsx", include_privacy_scan=False)

    reconciliation_plan = AutoLookupPlanner().infer_reconciliation_key_spec(employee_master, training)
    reconciliation = ReconciliationEngine().compare_lists(
        employee_master,
        training,
        reconciliation_plan.key_spec,
    )
    attach_auto_summary(reconciliation, reconciliation_plan, "빠진 사람/누락 제출자 찾기")
    writer.write_xlsx(reconciliation, OUTPUT / "04_교육누락자_대조결과.xlsx")

    allowance_template = get_template("allowance_budget_lookup")
    allowance_samples = merge_samples / "column_merge_allowance_budget"
    rate_reference = load_table(allowance_samples / "rate_reference.csv")
    payment_requests = load_table(allowance_samples / "payment_requests.csv")
    allowance_plan = AutoLookupPlanner().infer_lookup_plan(
        rate_reference,
        payment_requests,
        preferred_reference_key_columns=allowance_template.key_columns,
        preferred_target_key_columns=("직원번호", "지급월"),
        preferred_value_columns=allowance_template.value_columns,
    )
    allowance = MergeEngine().merge_lookup(
        rate_reference,
        payment_requests,
        allowance_plan.key_spec,
        value_columns=list(allowance_plan.value_columns),
    )
    attach_auto_summary(allowance, allowance_plan, allowance_template.name)
    writer.write_xlsx(allowance, OUTPUT / "05_수당예산_대조결과.xlsx")

    submitter_samples = extra_samples / "submission_reconciliation"
    expected = load_table(submitter_samples / "expected_submitters.csv")
    received = load_table(submitter_samples / "received_submitters.csv")
    submitter_plan = AutoLookupPlanner().infer_reconciliation_key_spec(
        expected,
        received,
        preferred_reference_key_columns=("기관코드",),
        preferred_target_key_columns=("제출기관코드",),
    )
    submitter_reconciliation = ReconciliationEngine().compare_lists(
        expected,
        received,
        submitter_plan.key_spec,
        reference_label="제출 대상 명단",
        target_label="실제 제출 명단",
    )
    attach_auto_summary(submitter_reconciliation, submitter_plan, "제출 대상자 누락 확인")
    writer.write_xlsx(submitter_reconciliation, OUTPUT / "06_제출대상_누락확인.xlsx")

    department_status = ConsolidationEngine().consolidate_folder(
        merge_samples / "row_merge_messy_headers",
        template="department_status_consolidation",
    )
    writer.write_xlsx(department_status, OUTPUT / "07_부서별현황_수합결과.xlsx", include_privacy_scan=False)

    monthly = load_table(extra_samples / "horizontal_table" / "monthly_budget_wide.csv")
    converted = HorizontalTableEngine().wide_to_long(monthly, id_columns=["기관명", "항목"])
    horizontal = JobResult(
        result_frame=converted,
        summary={"workflow": "월별 가로표 세로 변환", "row_count": len(converted)},
    )
    writer.write_xlsx(horizontal, OUTPUT / "08_월별가로표_세로변환.xlsx")

    for path in sorted(OUTPUT.glob("*.xlsx")):
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
