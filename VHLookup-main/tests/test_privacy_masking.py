from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

from vhlookup_core.privacy_masking import PrivacyMaskingEngine, SHEET_MASKED, SHEET_RECORDS


def test_privacy_masking_masks_common_personal_information(tmp_path):
    source = tmp_path / "privacy.xlsx"
    output = tmp_path / "masked.xlsx"
    pd.DataFrame(
        {
            "성명": ["홍길동"],
            "주민등록번호": ["900101-1234567"],
            "외국인등록번호": ["880811-5234567"],
            "연락처": ["010-1234-5678"],
            "이메일": ["hong@example.go.kr"],
            "계좌번호": ["110123456789"],
            "주소": ["서울특별시 중구 세종대로 110"],
            "성별": ["남"],
            "나이": [38],
            "생년월일": ["1988-08-11"],
            "비고": ["주민번호 8505052234567 확인 생년월일 1900/01/01 93년8월11일생 생년 930811"],
        }
    ).to_excel(source, index=False)

    PrivacyMaskingEngine().write_xlsx(source, output)

    sheets = pd.read_excel(output, sheet_name=None)
    workbook = load_workbook(output)
    masked = sheets[SHEET_MASKED]
    records = sheets[SHEET_RECORDS]

    assert workbook.sheetnames[0] == SHEET_MASKED
    assert masked.loc[0, "성명"] == "***"
    assert masked.loc[0, "주민등록번호"] == "******-*******"
    assert masked.loc[0, "외국인등록번호"] == "******-*******"
    assert masked.loc[0, "연락처"] == "***"
    assert masked.loc[0, "이메일"] == "***"
    assert masked.loc[0, "계좌번호"] == "***"
    assert masked.loc[0, "주소"] == "***"
    assert masked.loc[0, "성별"] == "***"
    assert masked.loc[0, "나이"] == "***"
    assert masked.loc[0, "생년월일"] == "***"
    assert (
        masked.loc[0, "비고"]
        == "주민번호 ******-******* 확인 생년월일 *** ***생 생년 ***"
    )
    assert set(records["마스킹 유형"]) >= {
        "이름",
        "주민등록번호",
        "연락처",
        "이메일",
        "계좌번호",
        "주소",
        "성별",
        "나이",
        "생년월일",
    }
    assert "홍길동" not in records.to_string()
    assert "1234567" not in records.to_string()
    assert "880811" not in masked.to_string()
    assert "010" not in masked.to_string()
    assert "5678" not in masked.to_string()
    assert "110123456789" not in masked.to_string()
    assert "1988-08-11" not in masked.to_string()
    assert "1900/01/01" not in masked.to_string()
    assert "93년8월11일" not in masked.to_string()
    assert "930811" not in masked.to_string()
    assert workbook[SHEET_MASKED]["A2"].comment is not None

    with ZipFile(output) as archive:
        drawing_xml = "\n".join(
            archive.read(name).decode("utf-8", errors="ignore")
            for name in archive.namelist()
            if "commentsDrawing" in name
        )
    assert "width:420px" in drawing_xml
