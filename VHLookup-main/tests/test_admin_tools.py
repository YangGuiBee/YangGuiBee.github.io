import re
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

from vhlookup_core.admin_tools import AdminWorkbookTools
from vhlookup_core.report import ReportWriter


def test_admin_clean_file_preserves_columns_and_adds_review_metadata(tmp_path):
    source = tmp_path / "dirty.xlsx"
    pd.DataFrame(
        [
            ["보고서", None, None, None],
            ["Employee_Name", "Department", "Salary", "Memo"],
            [" 홍길동 ", "IT", 100, " 확인 "],
            [None, None, None, None],
            ["김영희", "Sales", 200, None],
        ]
    ).to_excel(source, header=False, index=False)

    result = AdminWorkbookTools().clean_file(source)

    assert list(result.result_frame.columns) == [
        "원본 파일명",
        "원본 시트명",
        "원본 행 번호",
        "Employee_Name",
        "Department",
        "Salary",
        "Memo",
    ]
    assert len(result.result_frame) == 2
    assert result.result_frame.loc[0, "Employee_Name"] == "홍길동"
    assert result.result_frame.loc[0, "Memo"] == "확인"
    assert result.summary["row_count"] == 2


def test_admin_split_workbook_infers_department_column(tmp_path):
    source = tmp_path / "hr.xlsx"
    output = tmp_path / "split.xlsx"
    pd.DataFrame(
        {
            "Employee_Name": ["A", "B", "C"],
            "Department": ["IT", "Sales", "IT"],
            "Salary": [100, 200, 300],
        }
    ).to_excel(source, index=False)

    tools = AdminWorkbookTools()
    frame, _metadata = tools.load_table(source)
    assert tools.infer_split_column(frame) == "Department"

    split_result = tools.write_split_workbook(source, output)
    sheets = pd.read_excel(output, sheet_name=None)
    workbook = load_workbook(output)
    first_sheet = workbook[workbook.sheetnames[0]]

    assert split_result.split_column == "Department"
    assert {"확인사항", "전체", "IT", "Sales"} <= set(sheets)
    assert workbook.sheetnames[0] == "IT"
    assert first_sheet["B1"].comment is None
    assert first_sheet["B2"].comment is None
    assert len(sheets["IT"]) == 2


def test_clean_file_report_can_be_written(tmp_path):
    source = tmp_path / "dirty.xlsx"
    output = tmp_path / "cleaned.xlsx"
    pd.DataFrame({"성명": ["홍길동"], "부서": ["총무"]}).to_excel(source, index=False)

    result = AdminWorkbookTools().clean_file(source)
    ReportWriter().write_xlsx(result, output)
    sheets = pd.read_excel(output, sheet_name=None)

    assert set(sheets) == {"결과", "확인사항"}
    assert "파일 정리" in set(sheets["확인사항"]["구분"])


def test_pivot_summary_builds_cross_tab_and_workbook(tmp_path):
    source = tmp_path / "budget.xlsx"
    output = tmp_path / "pivot.xlsx"
    pd.DataFrame(
        {
            "부서": ["총무", "총무", "예산", "예산"],
            "월": ["1월", "2월", "1월", "2월"],
            "사업명": ["교육", "교육", "홍보", "홍보"],
            "금액": [100, 150, 200, 50],
        }
    ).to_excel(source, index=False)

    tools = AdminWorkbookTools()
    frame, _metadata = tools.prepare_pivot_source(source)
    defaults = tools.infer_pivot_defaults(frame)
    assert defaults["row_column"] == "부서"
    assert defaults["value_column"] == "금액"

    summary = tools.build_pivot_summary(frame, row_column="부서", column_column="월", value_column="금액", aggregation="합계")
    totals = summary.pivot_frame.set_index("부서")

    assert totals.loc["총무", "1월"] == 100
    assert totals.loc["총무", "2월"] == 150
    assert totals.loc["총무", "합계"] == 250
    assert totals.loc["합계", "합계"] == 500

    result = tools.write_pivot_workbook(source, output, row_column="부서", column_column="월", value_column="금액", aggregation="합계")
    sheets = pd.read_excel(output, sheet_name=None)
    workbook = load_workbook(output)

    assert result.row_column == "부서"
    assert set(sheets) == {"피벗요약", "확인사항"}
    assert workbook.sheetnames[0] == "피벗요약"
    assert "상위 목록" in set(sheets["확인사항"]["구분"])
    assert workbook["피벗요약"]["A1"].comment is not None
    with ZipFile(output) as archive:
        drawing_xml = "\n".join(
            archive.read(name).decode("utf-8", errors="ignore")
            for name in archive.namelist()
            if "commentsDrawing" in name
        )
    assert "width:420px" in drawing_xml
    heights = [int(value) for value in re.findall(r"height:(\d+)px", drawing_xml)]
    assert heights
    assert min(heights) >= 160
    assert sheets["피벗요약"].set_index("부서").loc["예산", "합계"] == 250


def test_pivot_summary_count_and_invalid_numeric_rows(tmp_path):
    frame = pd.DataFrame(
        {
            "기관명": ["A기관", "A기관", "B기관"],
            "상태": ["제출", "미제출", "제출"],
            "금액": ["1,000원", "오류", "50"],
        }
    )
    tools = AdminWorkbookTools()

    count_summary = tools.build_pivot_summary(frame, row_column="기관명", column_column="상태", aggregation="건수")
    count_totals = count_summary.pivot_frame.set_index("기관명")
    assert count_totals.loc["A기관", "합계"] == 2
    assert count_totals.loc["합계", "합계"] == 3

    sum_summary = tools.build_pivot_summary(frame, row_column="기관명", value_column="금액", aggregation="합계")
    sum_totals = sum_summary.pivot_frame.set_index("기관명")
    assert sum_totals.loc["A기관", "금액_합계"] == 1000
    assert len(sum_summary.invalid_rows) == 1
    assert sum_summary.invalid_rows.iloc[0]["기관명"] == "A기관"


def test_pivot_average_numbers_are_rounded_and_formatted(tmp_path):
    source = tmp_path / "scores.xlsx"
    output = tmp_path / "pivot_average.xlsx"
    pd.DataFrame(
        {
            "부서": ["총무", "총무", "예산"],
            "월": ["1월", "2월", "1월"],
            "점수": [100, 101, 10 / 3],
        }
    ).to_excel(source, index=False)

    tools = AdminWorkbookTools()
    frame, _metadata = tools.prepare_pivot_source(source)
    summary = tools.build_pivot_summary(frame, row_column="부서", value_column="점수", aggregation="평균")

    averages = summary.pivot_frame.set_index("부서")
    assert averages.loc["총무", "점수_평균"] == 100.5
    assert averages.loc["예산", "점수_평균"] == 3.33

    tools.write_pivot_workbook(source, output, row_column="부서", value_column="점수", aggregation="평균")
    workbook = load_workbook(output)
    sheet = workbook["피벗요약"]

    assert sheet["B2"].value == 100.5
    assert sheet["B2"].number_format == "#,##0.##"
    assert sheet["B3"].value == 3.33
