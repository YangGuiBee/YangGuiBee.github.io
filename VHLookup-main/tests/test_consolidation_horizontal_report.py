from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from vhlookup_core.consolidation import ConsolidationEngine
from vhlookup_core.horizontal import HorizontalTableEngine
from vhlookup_core.models import JobResult
from vhlookup_core.report import (
    ReportWriter,
    SHEET_REVIEW,
    SHEET_RESULT,
)
from vhlookup_core.privacy import PrivacyScanner


SAMPLES = Path("samples/public_admin")
MERGE_SAMPLES = SAMPLES / "03_merge_files"
HR_SAMPLES = MERGE_SAMPLES / "column_merge_hr_training"


def test_consolidation_folder_merges_files_with_different_headers(tmp_path):
    file_one = tmp_path / "one.xlsx"
    file_two = tmp_path / "two.xlsx"
    pd.DataFrame(
        [
            ["직원 현황", None, None],
            ["사번", "성명", "부서"],
            ["001", "홍길동", "총무"],
        ]
    ).to_excel(file_one, header=False, index=False)
    pd.DataFrame(
        [
            ["제출자료", None, None],
            ["직원번호", "이름", "소속"],
            ["002", "김영희", "인사"],
        ]
    ).to_excel(file_two, header=False, index=False)

    result = ConsolidationEngine().consolidate_folder(
        tmp_path,
        standard_columns=["사번", "성명", "부서"],
    )

    assert len(result.result_frame) == 2
    assert list(result.result_frame.columns) == ["사번", "성명", "부서"]
    assert set(result.result_frame["성명"]) == {"홍길동", "김영희"}


def test_consolidation_preserves_unmapped_source_columns(tmp_path):
    file_one = tmp_path / "one.xlsx"
    file_two = tmp_path / "two.xlsx"
    pd.DataFrame(
        [
            ["직원 현황", None, None, None],
            ["사번", "성명", "부서", "첨부파일"],
            ["001", "홍길동", "총무", "있음"],
        ]
    ).to_excel(file_one, header=False, index=False)
    pd.DataFrame(
        [
            ["제출자료", None, None, None],
            ["직원번호", "이름", "소속", "비고2"],
            ["002", "김영희", "인사", "추가 확인"],
        ]
    ).to_excel(file_two, header=False, index=False)

    result = ConsolidationEngine().consolidate_folder(
        tmp_path,
        standard_columns=["사번", "성명", "부서"],
    )

    assert {"첨부파일", "비고2"} <= set(result.result_frame.columns)
    assert result.result_frame.loc[0, "첨부파일"] == "있음"
    assert result.result_frame.loc[1, "비고2"] == "추가 확인"
    assert "원본 컬럼 보존" in {row["역할"] for row in result.mapping_records}


def test_consolidation_accepts_user_corrected_column_mapping(tmp_path):
    file_one = tmp_path / "one.xlsx"
    file_two = tmp_path / "two.xlsx"
    pd.DataFrame({"사번": ["001"], "성명": ["홍길동"], "부서": ["총무"]}).to_excel(file_one, index=False)
    pd.DataFrame({"직원번호": ["002"], "이름": ["김영희"], "팀명": ["인사"]}).to_excel(file_two, index=False)

    result = ConsolidationEngine().consolidate_files(
        [file_one, file_two],
        standard_columns=["사번", "성명", "부서"],
        saved_mappings_by_file={str(file_two.resolve()): {"사번": "직원번호", "성명": "이름", "부서": "팀명"}},
    )

    assert result.result_frame.loc[1, "사번"] == "002"
    assert result.result_frame.loc[1, "성명"] == "김영희"
    assert result.result_frame.loc[1, "부서"] == "인사"


def test_column_merge_attaches_columns_by_common_key(tmp_path):
    file_one = tmp_path / "base.xlsx"
    file_two = tmp_path / "extra.xlsx"
    pd.DataFrame(
        {
            "사번": ["001", "002"],
            "성명": ["홍길동", "김영희"],
        }
    ).to_excel(file_one, index=False)
    pd.DataFrame(
        {
            "사번": ["001", "002"],
            "직급": ["주무관", "팀장"],
            "부서": ["총무", "인사"],
        }
    ).to_excel(file_two, index=False)

    result = ConsolidationEngine().merge_files_by_columns([file_one, file_two])

    assert len(result.result_frame) == 2
    assert {"사번", "성명", "직급", "부서"} <= set(result.result_frame.columns)
    assert result.result_frame.loc[0, "직급"] == "주무관"
    assert result.summary["workflow"] == "열 방향 파일 합치기"
    assert "키 컬럼" in {row["역할"] for row in result.mapping_records}


def test_column_merge_marks_unmatched_attached_cells_with_comment(tmp_path):
    output = tmp_path / "result.xlsx"
    file_one = HR_SAMPLES / "hr_employee_master.csv"
    file_two = HR_SAMPLES / "hr_training_completion.csv"

    result = ConsolidationEngine().merge_files_by_columns([file_one, file_two])
    ReportWriter().write_xlsx(result, output, include_privacy_scan=False)

    assert len(result.result_frame) == 4
    assert result.result_frame.loc[2, "사번"] == "00125"
    assert pd.isna(result.result_frame.loc[2, "교육명"])
    assert result.result_frame.loc[3, "사번"] == "00999"
    assert pd.isna(result.result_frame.loc[3, "성명"])
    issue = next(issue for issue in result.issues if issue.issue_type == "match_failed")
    assert issue.message == "2번째 파일(hr_training_completion.csv)에 해당 기준값이 없습니다."
    reference_only_issue = next(issue for issue in result.issues if issue.issue_type == "reference_only_unmatched")
    assert reference_only_issue.message == "2번째 파일(hr_training_completion.csv)에만 있는 기준값입니다. 다른 파일에는 없으니 확인하세요."

    workbook = load_workbook(output)
    result_sheet = workbook[SHEET_RESULT]
    headers = [cell.value for cell in result_sheet[1]]
    education_column = headers.index("교육명") + 1
    name_column = headers.index("성명") + 1
    base_column = headers.index("직급") + 1
    missing_cell = result_sheet.cell(row=4, column=education_column)
    reference_only_cell = result_sheet.cell(row=5, column=name_column)
    base_cell = result_sheet.cell(row=4, column=base_column)
    existing_name_cell = result_sheet.cell(row=2, column=name_column)

    assert missing_cell.value is None
    assert missing_cell.fill.fgColor.rgb in {"00FDE68A", "FDE68A"}
    assert missing_cell.comment is not None
    assert "2번째 파일(hr_training_completion.csv)에 해당 기준값이 없습니다." in missing_cell.comment.text
    assert reference_only_cell.value is None
    assert reference_only_cell.fill.fgColor.rgb in {"00FDE68A", "FDE68A"}
    assert reference_only_cell.comment is not None
    assert "2번째 파일(hr_training_completion.csv)에만 있는 기준값입니다." in reference_only_cell.comment.text
    assert base_cell.comment is None
    assert existing_name_cell.comment is None
    assert existing_name_cell.fill.fgColor.rgb not in {"00FDE68A", "FDE68A"}


def test_report_writer_extracts_error_rows(tmp_path):
    output = tmp_path / "result.xlsx"
    result = ConsolidationEngine().consolidate_folder(
        MERGE_SAMPLES / "row_merge_submission_errors",
        template="school_submission_consolidation",
    )

    ReportWriter().write_xlsx(result, output)
    sheets = pd.read_excel(output, sheet_name=None)

    assert set(sheets) == {SHEET_RESULT, SHEET_REVIEW}
    assert not sheets[SHEET_REVIEW].empty
    assert {"구분", "판단 내용", "확인할 점"} <= set(sheets[SHEET_REVIEW].columns)
    assert {"필수값 누락", "숫자 오류"} & set(sheets[SHEET_REVIEW]["구분"])


def test_report_writer_can_leave_merge_result_unmarked(tmp_path):
    output = tmp_path / "result.xlsx"
    result = ConsolidationEngine().consolidate_folder(
        MERGE_SAMPLES / "row_merge_submission_errors",
        template="school_submission_consolidation",
    )

    ReportWriter().write_xlsx(result, output, mark_result_cells=False, include_privacy_scan=False)
    workbook = load_workbook(output)
    result_sheet = workbook[SHEET_RESULT]

    assert workbook.sheetnames[0] == SHEET_RESULT
    assert all(cell.comment is None for row in result_sheet.iter_rows() for cell in row)


def test_horizontal_table_engine_detects_and_converts_month_columns():
    frame = pd.DataFrame(
        {
            "사번": ["001", "002"],
            "성명": ["홍길동", "김영희"],
            "1월": [10, 20],
            "2월": [11, 21],
            "3월": [12, 22],
        }
    )
    engine = HorizontalTableEngine()

    detection = engine.detect(frame)
    converted = engine.wide_to_long(frame, id_columns=["사번", "성명"])

    assert detection.is_horizontal
    assert set(detection.value_columns) == {"1월", "2월", "3월"}
    assert len(converted) == 6
    assert set(converted.columns) == {"사번", "성명", "열 기준", "값"}


def test_report_writer_creates_expected_sheets(tmp_path):
    output = tmp_path / "result.xlsx"
    result = JobResult(
        result_frame=pd.DataFrame({"성명": ["홍길동"]}),
        summary={"row_count": 1},
        mapping_records=[
            {
                "역할": "키 컬럼",
                "기준표 컬럼": "사번",
                "대상표 컬럼": "직원번호",
                "신뢰도": 0.9,
                "추천 방식": "테스트",
                "검토 메모": "",
            }
        ],
    )

    ReportWriter().write_xlsx(result, output)
    sheets = pd.read_excel(output, sheet_name=None)

    assert set(sheets) == {SHEET_RESULT, SHEET_REVIEW}
    assert list(sheets)[0] == SHEET_RESULT
    assert sheets[SHEET_RESULT].loc[0, "성명"] == "홍길동"
    assert "키 컬럼" in set(sheets[SHEET_REVIEW]["구분"])
    assert "개인정보 의심" in set(sheets[SHEET_REVIEW]["구분"])


def test_privacy_scanner_flags_sensitive_columns_without_values():
    frame = pd.DataFrame(
        {
            "성명": ["홍길동"],
            "연락처": ["010-1234-5678"],
            "외국인등록번호": ["880811-5234567"],
            "성별": ["남"],
            "나이": [38],
            "생년월일": ["1988-08-11"],
            "비고": ["주민번호 900101-1234567 확인 생년월일 1900/01/01 93년8월11일생 생년 930811"],
        }
    )

    records = PrivacyScanner().scan_frame(frame)

    assert {record["점검 유형"] for record in records} >= {
        "개인 식별 이름",
        "연락처",
        "고유식별정보",
        "고유식별번호 패턴",
        "생년월일 패턴",
        "성별",
        "나이",
        "생년월일",
    }
    serialized = str(records)
    assert "홍길동" not in serialized
    assert "010-1234-5678" not in serialized
    assert "900101-1234567" not in serialized
    assert "880811-5234567" not in serialized
    assert "1988-08-11" not in serialized
    assert "1900/01/01" not in serialized
    assert "93년8월11일" not in serialized
    assert "930811" not in serialized
