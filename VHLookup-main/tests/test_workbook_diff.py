from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

from vhlookup_core.workbook_diff import WorkbookDiffEngine, WorkbookDiffReportWriter


def test_workbook_diff_writes_comments_on_changed_after_cells(tmp_path):
    before = tmp_path / "before.xlsx"
    after = tmp_path / "after.xlsx"
    output = tmp_path / "diff.xlsx"

    pd.DataFrame(
        {
            "사번": ["001", "002"],
            "성명": ["홍길동", "김영희"],
            "금액": [100, 200],
        }
    ).to_excel(before, index=False)
    pd.DataFrame(
        {
            "사번": ["001", "002"],
            "성명": ["홍길동", "김영희"],
            "금액": [150, 200],
        }
    ).to_excel(after, index=False)

    result = WorkbookDiffEngine().compare_files(before, after)
    WorkbookDiffReportWriter().write_xlsx(result, output)

    sheets = pd.read_excel(output, sheet_name=None)
    workbook = load_workbook(output)
    memo_sheet = workbook["후파일_메모"]

    assert result.summary["changed_cell_count"] == 1
    assert workbook.sheetnames[0] == "후파일_메모"
    assert set(sheets) == {"후파일_메모", "확인사항"}
    assert sheets["확인사항"].loc[0, "컬럼명"] == "금액"
    assert memo_sheet["C2"].comment is not None
    assert "전 값: 100" in memo_sheet["C2"].comment.text
    assert "후 값: 150" in memo_sheet["C2"].comment.text
    with ZipFile(output) as archive:
        drawing_xml = "\n".join(
            archive.read(name).decode("utf-8", errors="ignore")
            for name in archive.namelist()
            if "commentsDrawing" in name
        )
    assert "width:420px" in drawing_xml
    assert "height:160px" in drawing_xml


def test_workbook_diff_accepts_user_corrected_column_mapping(tmp_path):
    before = tmp_path / "before.xlsx"
    after = tmp_path / "after.xlsx"

    pd.DataFrame({"사번": ["001"], "지급액": [100]}).to_excel(before, index=False)
    pd.DataFrame({"직원번호": ["001"], "최종금액": [120]}).to_excel(after, index=False)

    result = WorkbookDiffEngine().compare_files(
        before,
        after,
        column_mapping_override={"사번": "직원번호", "지급액": "최종금액"},
        key_columns=("사번",),
    )

    assert result.summary["changed_cell_count"] == 1
    assert result.diff_frame.loc[0, "전 값"] == "100"
    assert result.diff_frame.loc[0, "후 값"] == "120"


def test_workbook_diff_reports_missing_rows_with_original_values(tmp_path):
    before = tmp_path / "before.xlsx"
    after = tmp_path / "after.xlsx"
    output = tmp_path / "diff.xlsx"

    pd.DataFrame(
        {
            "사번": ["E001", "E002", "E003"],
            "성명": ["홍길동", "김영희", "박철수"],
            "부서": ["총무", "회계", "민원"],
            "지급월": ["2026-06", "2026-06", "2026-06"],
            "금액": [100000, 150000, 120000],
            "비고": ["정상", "정상", "정상"],
        }
    ).to_excel(before, index=False)
    pd.DataFrame(
        {
            "사번": ["E001", "E002", "E004"],
            "성명": ["홍길동", "김영희", "최민수"],
            "부서": ["총무", "회계", "복지"],
            "지급월": ["2026-06", "2026-06", "2026-06"],
            "금액": [100000, 170000, 130000],
            "비고": ["정상", "금액 수정", "추가"],
        }
    ).to_excel(after, index=False)

    result = WorkbookDiffEngine().compare_files(before, after)
    WorkbookDiffReportWriter().write_xlsx(result, output)
    sheets = pd.read_excel(output, sheet_name=None)
    workbook = load_workbook(output)
    memo_sheet = workbook["후파일_메모"]

    assert result.summary["missing_row_count"] == 1
    assert result.summary["added_row_count"] == 1
    assert memo_sheet["A4"].value == "E004"
    assert memo_sheet["A5"].value == "E003"
    assert all(memo_sheet.cell(row=4, column=column).fill.fgColor.rgb in {"00BFDBFE", "BFDBFE"} for column in range(1, 7))
    assert all(memo_sheet.cell(row=5, column=column).fill.fgColor.rgb in {"00FECACA", "FECACA"} for column in range(1, 7))
    assert memo_sheet["A4"].comment is not None
    assert "전 파일에 없던 행" in memo_sheet["A4"].comment.text
    assert memo_sheet["A5"].comment is not None
    assert "후 파일에서 사라진 행" in memo_sheet["A5"].comment.text
    assert memo_sheet["E3"].fill.fgColor.rgb in {"00FDE68A", "FDE68A"}
    review = sheets["확인사항"]
    missing_row = review[review["비교 기준"] == "E003"].iloc[0]
    assert "박철수" in missing_row["전 값"]
    assert "후 파일에 행 없음" in set(review["상태"])
    assert "전 파일에 없던 행" in set(review["상태"])


def test_workbook_diff_auto_maps_renamed_columns_by_data_overlap_and_loose_keys(tmp_path):
    before = tmp_path / "before.xlsx"
    after = tmp_path / "after.xlsx"

    pd.DataFrame(
        {
            "관리번호": ["001", "002", "003"],
            "지급액": [100, 200, 300],
            "담당부서": ["총무", "회계", "민원"],
        }
    ).to_excel(before, index=False)
    pd.DataFrame(
        {
            "접수ID": ["1", "002", "004"],
            "최종지급액": [100, 250, 400],
            "소속부서": ["총무", "회계", "복지"],
        }
    ).to_excel(after, index=False)

    result = WorkbookDiffEngine().compare_files(before, after)

    assert result.summary["compare_basis"] == "키 컬럼: 관리번호"
    assert result.summary["missing_row_count"] == 1
    assert result.summary["added_row_count"] == 1
    assert result.summary["changed_cell_count"] == 1
    assert result.diff_frame.loc[result.diff_frame["상태"] == "값 다름", "컬럼명"].tolist() == ["지급액"]
    assert result.row_frame.loc[result.row_frame["비교 기준"] == "3", "상태"].item() == "후 파일에 행 없음"
